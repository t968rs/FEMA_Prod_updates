"""Performs a QC check of the attributes of the database tables"""

import os
import sys
import arcpy
import string
from arcpy.da import SearchCursor

try:
    import openpyxl
except ImportError:
    arcpy.AddError("Unable to import 'openpyxl'.  "
                   "Install 'openpyxl using PIP or contact the Software Developer.")
    sys.exit(1)
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

# Contains dictionaries of the domains found in a standard FEMA DFIRM database.  The variable
# name is the name of the domain.  The KEY is the coded value of the domain.  The VALUE is
# the text value of the domain.
D_Area_Units = {
    "1000": "Acres", "1010": "Hectares", "1020": "Square Feet", "1030": "Square Meters",
    "1040": "Square Yards", "1050": "Square Miles", "1060": "Square Kilometers", "NP": "NP"}
D_Basemap_Typ = {"1000": "Orthophoto", "2000": "Vector", "NP": "NP"}
D_Beachset = {
    "1000": "Sandy Beach Backed By Low Sand Berm or High Sand Dune Formation",
    "1010": "Sandy Beach Backed By Shore Protection Structures",
    "1020": "Erodible Coastal Bluffs", "1030": "Non-Erodible Coastal Cliffs and Bluffs",
    "1040": "Tidal Flats and Wetlands",
    "1050": "Cobble, Gravel, Shingle, or Mixed Grain Sized Beach and Berms", "NP": "NP"}
D_Carto_Trans_Code = {
    "0100": "Interstates", "0200": "US Highways", "0300": "State Highways",
    "0400": "County Roads", "0500": "Local Roads", "0700": "Railroads", "0800": "Airports",
    "NP": "NP"}
D_Cert_Status = {
    "1000": "Meets Existing FEMA Accreditation Policy",
    "1010": "Certified Prior to Existing FEMA Policy",
    "1020": "Certified by Another Federal Agency",
    "2000": "Not Certified: Failure of Vertical Structure",
    "2010": "Not Certified: Failure of Revetment", "2030": "Not Certified: Failure of Levee",
    "3000": "Not Applicable", "NP": "NP"}
D_Cst_Struct = {
    "1000": "Beach Stabilization Structure", "2000": "Coastal Armoring Structure",
    "3000": "Miscellaneous Structure", "NP": "NP"}
D_Cst_Typ = {
    "1000": "Open Coast", "2000": "Sheltered Waters",
    "3000": "Combined Open Coast and Sheltered Waters", "NP": "NP"}
D_Discharge_Units = {"CFS": "CFS", "CMS": "CMS", "NP": "NP"}
D_Erosion = {
    "1010": "Not Applied", "1020": "Dune Removal", "1030": "Dune Retreat", "1040": "K-D",
    "1050": "MK-A", "1060": "SBEACH", "9000": "Other", "NP": "NP"}
D_Event = {
    "01minus": "1 Percent Minus Chance", "01pct": "1 Percent Chance",
    "01pctfut": "1 Percent Chance Future Conditions", "01plus": "1 Percent Plus Chance",
    "02pct": "2 Percent Chance", "04pct": "4 Percent Chance", "0_2pct": "0.2 Percent Chance",
    "10pct": "10 Percent Chance", "50pct": "50 Percent Chance"}
D_Flood_Typ = {
    "1000": "Riverine", "1010": "Coastal", "1020": "Alluvial", "1030": "Lacustrine",
    "1040": "Ponding"}
D_Font = {
    "1000": "Arial", "1010": "Arial Narrow", "2000": "Franklin Gothic Book",
    "2010": "Franklin Gothic Medium", "2020": "Franklin Gothic Medium Cond",
    "3000": "Times New Roman", "4000": "Verdana", "NP": "NP"}
D_Font_Style = {
    "1000": "Regular", "2000": "Bold", "3000": "Italic", "4000": "Bold Italic", "NP": "NP"}
D_Gage_Typ = {
    "1000": "Flow", "1010": "Flow / Stage", "1020": "Stage", "1100": "Fixed Interval",
    "1110": "Instantaneous", "1120": "Tipping", "1200": "Wave Height",
    "1210": "Wind Direction", "1220": "Wind Speed", "1230": "Wind Speed and Direction",
    "1240": "Tide", "1250": "Wave Height and Direction", "NP": "NP"}
D_Horiz_Datum = {
    "83HARN": "North American Datum 1983 HARN", "NAD27": "North American Datum 1927",
    "NAD83": "North American Datum 1983", "NP": "NP", "NSRS07": "NAD83 (NSRS2007)",
    "WGS84": "World Geodetic System 1984"}
D_Hydra_Mdl = {
    "0110": "CHAN for Windows v. 2.03 (1997)",
    "0120": "Culvert Master v. 2.0 (September 2000) and up", "0140": "FAN",
    "0260": "MIKE Flood HD (2002 D and 2004)", "0262": "MIKE Flood HD v.2009 SP4",
    "0270": "NETWORK (June 2002)", "0280": "PondPack v. 8 (May 2002) and up",
    "0300": "S2DMM (Feb 2008)", "0310": "StormCAD v.4 (June 2002) and up",
    "0322": "SWMM 5 Version 5.0.005 (May 2005) and up", "0360": "XP-SWMM 8.52 and up",
    "0362": "XPSWMM 2D/XPStorm 2D v. 12.00 (May 2010)", "0370": "Xpstorm 10.0 (May 2006)",
    "0401": "SMS ADH v11.1 and up", "0402": "SMS ADCIRC v11.1 and up",
    "0403": "SMS BOUSS-2D v11.1 and up", "0404": "SMS CGWAVE v11.1 and up",
    "0405": "SMS CMS Flow v11.1 and up", "0406": "SMS CMS Wave v11.1 and up",
    "0407": "SMS FESWMS v11.1 and up", "0408": "SMS HYDRO_AS-2D",
    "0409": "SMS GENCADE v11.1 and up", "0410": "SMS PTM v11.1 and up",
    "0411": "SMS RiverFlow2D v11.2 and up", "0412": "SMS RMA2 v11.1 and up",
    "0413": "SMS RMA4 v11.1 and up", "0414": "SMS SRH-2D v11.2 and up",
    "0415": "SMS STWAVE v11.1 and up", "0416": "SMS TUFLOW v11.1 and up",
    "0417": "SMS TUFLOW AD v11.1 and up", "0418": "SMS TUFLOW Multiple Domains v11.1 and up",
    "0419": "SMS TUFLOW FV v11.1 and up", "0420": "SMS WAM v11.1 and up",
    "1000":
        "ICPR 2.20 (Oct. 2000), 3.02 (Nov. 2002), and 3.10 (April 2008) with PercPack Option",
    "1001": "DHM 21 and 34 (Aug. 1987)", "1002": "FEQ 8.92 (1999) and FEQ 9.98 (2005)",
    "1003": "FEQUTL 4.68 (1999) and FEQUTL 5.46 (2005)",
    "1004": "FESWMS 2DH 1.1 and up (Jun. 1995)", "1005": "FLDWAV (Nov. 1998)",
    "1006": "FLO-2D v. 2007.06 and 2009.06", "1007": "FLO-2D V.2003.6, 2004.10 and 2006.1",
    "1008": "Gage Analysis", "1009": "HCSWMM 4.31B (August 2000)",
    "1010": "HEC-2 4.6.2 (May 1991", "1012": "HEC-RAS 3.1.1 and up",
    "1013": "HY8 4.1 and up (Nov. 1992)", "1014": "HEC-RAS 5.0 and up",
    "1015": "MIKE 11 HD (2002 D, 2004)",
    "1017": "QUICK-2 1.0 and up (Jan. 1995)", "1021": "SWMM 4.30 (MAY 1994)",
    "1022": "SWMM 4.31 (JANUARY 1997)", "1023": "TABS RMA2 v. 4.3 and up (Oct. 1996)",
    "1024": "TABS RMA4 v. 4.5 and up (July 2000)", "1025": "UNET 4.0 (April 2001)",
    "1026": "WSPGW 12.96 (OCTOBER 2000)", "1027": "WSPRO (Jun. 1988 and up)",
    "1028": "MIKE URBAN Collection Systems (MOUSE) Release 2009, date June 2010",
    "1029": "TUFLOW Release Version 2010-10 (October 2010)", "1030": "MIKE 11 HD v.2009 SP4",
    "9000": "OTHER"}
D_Hydro_Mdl = {
    "0190": "PondPack v.8 (May 2002) and up", "0200": "PRMS Version 2.1 (Jan 1996)",
    "0222": "SWMM 5 Version 5.0.005 (May 2005) and up", "0231": "TR-20 Win 1.00 (Jan 2005)",
    "0240": "TR-55 (JUNE 1986)", "0250": "XP-SWMM 8.52 and up",
    "0260": "Xpstorm 10.0 (May 2006)", "2000": "AHYMO 97 (Aug. 1997)",
    "2001": "CUHPF/PC (May 1996 and May 2002)", "2005": "HEC-FFA 3.1 (February 1995)",
    "2006": "HEC-1 4.0.1 and up 1 (May 1991)", "2008": "HEC-HMS 3.0 and up (Dec 2005)",
    "2018": "HSPF 10.10 (Dec 1993) and up", "2022": "MIKE 11 RR (2009 SP4)",
    "2023": "MIKE 11 UHM (2002 D and 2004)", "2024": "PEAKFQ 2.4 (April 1998) and up",
    "2029": "Regression Equations", "2031": "SWMM (RUNOFF) 4.30 (May1994)",
    "2032": "SWMM (RUNOFF) 4.31 (Jan 1997)", "2033": "TR-20 Win (Feb 1992)",
    "2034": "WinTR-55 1.0.08 (Jan 2005)", "2040": "HEC-SSP 1.1 (April 2009) and up",
    "2041": "VCRat 2.6 (Dec. 2008)", "2042": "MIKE 11 (2009 SP4)", "9000": "OTHER"}
D_Jurisdiction_Typ = {
    "0100": "All Jurisdictions", "0200": "And Incorporated Areas", "0300": "Independent City",
    "0900": "Unincorporated Areas"}
D_LOMC_Status = {
    "1000": "Superseded", "1010": "Revalidated", "1020": "Incorporated",
    "1030": "Redetermined", "1040": "Effective", "NP": "NP"}
D_Label_Typ_2017 = {
    "1000": "Ortho-Transportation", "1001": "Ortho-Water", "1002": "S_Trnsport_Ln",
    "1003": "S_Wtr_Ar", "1004": "S_Wtr_Ln", "1005": "Notes and Map Collar", "1006": "S_BFE",
    "1007": "S_CBRS", "1008": "S_Cst_Tsct_Ln", "1009": "S_FIRM_Pan", "1010": "S_Fld_Haz_Ar",
    "1011": "S_Fld_Haz_Ln", "1012": "S_Gage", "1013": "S_Gen_Struct", "1014": "S_Levee",
    "1015": "S_LiMWA", "1016": "S_PFD_Ln", "1017": "S_PLSS_Ar", "1018": "S_Pol_Ar",
    "1019": "S_Profil_Basln", "1020": "S_Riv_Mrk", "1021": "S_Tsct_Basln", "1022": "S_XS",
    "NP": "NP"}
D_Label_Typ = {
    "1000": "Ortho-Transportation", "1001": "Ortho-Water", "1002": "S_Trnsport_Ln",
    "1003": "S_Wtr_Ar", "1004": "S_Wtr_Ln", "1005": "Notes and Map Collar", "1006": "S_BFE",
    "1008": "S_Cst_Tsct_Ln", "1009": "S_FIRM_Pan", "1010": "S_Fld_Haz_Ar",
    "1011": "S_Fld_Haz_Ln", "1012": "S_Gage", "1013": "S_Gen_Struct", "1014": "S_Levee",
    "1015": "S_LiMWA", "1016": "S_PFD_Ln", "1017": "S_PLSS_Ar", "1018": "S_Pol_Ar",
    "1019": "S_Profil_Basln", "1020": "S_Riv_Mrk", "1021": "S_Tsct_Basln", "1022": "S_XS",
    "NP": "NP"}
D_Length_Units = {
    "CM": "Centimeters", "FT": "Feet", "IN": "Inches", "KM": "Kilometers", "M": "Meters",
    "MI": "Miles", "MM": "Millimeters", "NP": "NP", "USFT": "U.S. Survey Feet"}
D_Levee_Analysis_Type_2017 = {
    "AL": "Accredited Levee", "FD": "Freeboard Deficient",
    "NHS": "Not Hydraulically Significant", "NV": "Natural Valley", "OTH": "Other",
    "OVR": "Overtopping", "SBI": "Structural-Based Inundation", "SR": "Sound Reach", "NP": "NP"}
D_Levee_Analysis_Type = {
    "BBT": "BFE Below Toe", "FD": "Freeboard Deficient",
    "NHS": "Not Hydraulically Significant", "NV": "Natural Valley", "OTH": "Other",
    "OVR": "Overtopping", "SBI": "Structural-Based Inundation", "SR": "Sound Reach", "NP": "NP"}
D_Levee_Scenario = {"B": "With Both Levees", "L": "With Left Levee", "R": "With Right Levee"}
D_Levee_Status_2017 = {
    "A": "Accredited", "D": "De_Accredited", "N": "Non-Accredited", "NP": "NP",
    "P": "Provisionally Accredited"}
D_Levee_Status = {
    "A": "Accredited", "N": "Non-Accredited", "NP": "NP", "P": "Provisionally Accredited"}
D_Levee_Type = {
    "CLC": "Coastal Levee Centerline", "CS": "Closure Structure", "FW": "Floodwall",
    "LC": "Levee Centerline", "D": "Dike", "NP": "NP"}
D_Ln_Typ = {
    "1010": "Limit Lines", "1020": "Other Boundary", "2034": "SFHA / Flood Zone Boundary"}
D_Loc_Accuracy = {"H": "High", "L": "Low", "M": "Medium"}
D_MTFCC = {
    "NP": "NP", "R1011": "Railroad Feature (Main, Spur, or Yard)",
    "R1051": "Carline, Streetcar Track, Monorail, Other Mass Transit Rail",
    "R1052": "Cog Rail Line, Incline Rail Line, Tram", "S1100": "Primary Road",
    "S1200": "Secondary Road", "S1400": "Local Neighborhood Road, Rural Road, City Street",
    "S1500": "Vehicular Trail (4WD)", "S1630": "Ramp",
    "S1640": "Service Drive Usually Along a Limited Access Highway",
    "S1710": "Walkway/Pedestrian Trail", "S1720": "Stairway", "S1730": "Alley",
    "S1740": "Private Road for Service Vehicles (Logging, Oil Fields, Ranches, Etc.)",
    "S1750": "Internal U.S. Census Bureau Use", "S1780": "Parking Lot Road",
    "S1820": "Bike Path or Trail", "S1830": "Bridle Path", "S2000": "Road Median"}
D_Method = {
    "1000": "Combo of Topo and Survey", "1010": "Cut from Topo", "1020": "Digitized from FIRM",
    "1030": "Field Survey", "NP": "NP"}
D_Mtg_Typ = {
    "1000": "Final CCO", "1010": "Flood Risk Review", "1020": "Initial CCO",
    "1030": "Intermediate CCO", "1040": "Project Discovery", "1050": "Resilience",
    "1060": "Scoping", "9000": "Other", "NP": "NP"}
D_Node_Typ = {
    "1000": "Diversion", "1010": "Junction", "1020": "Reservoir", "1030": "Structure",
    "1040": "Sub-Basin Outlet"}
D_Obscured = {
    "1": "Low Confidence Area", "2": "Acceptable Confidence Area", "NP": "NP"}
D_Panel_Typ = {
    "1000": "Countywide, Panel Printed", "1010": "Countywide, Not Printed",
    "1020": "Community Based, Panel Printed", "1030": "Community Based, Not Printed",
    "1040": "Unmapped Community", "1050": "Statewide, Panel Printed",
    "1060": "Statewide, Not Printed"}
D_Prof_Basln_Typ = {
    "1000": "Profile Baseline", "2000": "Profile Baseline and Stream Centerline",
    "3000": "Hydraulic Link", "UNK": "Unknown"}
D_Prof_Lbl_Adjust = {
    "B": "Bottom", "C": "Center", "L": "Left", "M": "Middle", "R": "Right", "T": "Top"}
D_Prof_Lbl_Orient = {"H": "Horizontal", "V": "Vertical"}
D_Proj_Unit = {
    "DECDEG": "Decimal Degrees", "INTLFT": "International Feet", "METER": "Meters", "NP": "NP",
    "USFT": "US Survey Feet"}
D_Projection = {
    "0101": "STATE PLANE TRANSVERSE MERCATOR ALABAMA EAST ZONE",
    "0101H": "HARN STATE PLANE TRANSVERSE MERCATOR, ALABAMA EAST ZONE",
    "0102": "STATE PLANE TRANSVERSE MERCATOR, ALABAMA WEST ZONE",
    "0102H": "HARN STATE PLANE TRANSVERSE MERCATOR, ALABAMA WEST ZONE",
    "0201": "STATE PLANE TRANSVERSE MERCATOR, ARIZONA EAST ZONE",
    "0201H": "HARN STATE PLANE TRANSVERSE MERCATOR, ARIZONA EAST ZONE",
    "0202": "STATE PLANE TRANSVERSE MERCATOR, ARIZONA CENTRAL ZONE",
    "0202H": "HARN STATE PLANE TRANSVERSE MERCATOR, ARIZONA CENTRAL ZONE",
    "0203": "STATE PLANE TRANSVERSE MERCATOR, ARIZONA WEST ZONE",
    "0203H": "HARN STATE PLANE TRANSVERSE MERCATOR, ARIZONA WEST ZONE",
    "0301": "STATE PLANE LAMBERT CONFORMAL CONIC, ARKANSAS NORTH ZONE",
    "0302": "STATE PLANE LAMBERT CONFORMAL CONIC, ARKANSAS SOUTH ZONE",
    "0401": "STATE PLANE LAMBERT CONFORMAL CONIC, CALIFORNIA I ZONE",
    "0401H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, CALIFORNIA I ZONE",
    "0402": "STATE PLANE LAMBERT CONFORMAL CONIC, CALIFORNIA II ZONE",
    "0402H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, CALIFORNIA II ZONE",
    "0403": "STATE PLANE LAMBERT CONFORMAL CONIC, CALIFORNIA III ZONE",
    "0403H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, CALIFORNIA III ZONE",
    "0404": "STATE PLANE LAMBERT CONFORMAL CONIC, CALIFORNIA IV ZONE",
    "0404H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, CALIFORNIA IV ZONE",
    "0405": "STATE PLANE LAMBERT CONFORMAL CONIC, CALIFORNIA V ZONE",
    "0405H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, CALIFORNIA V ZONE",
    "0406": "STATE PLANE LAMBERT CONFORMAL CONIC, CALIFORNIA VI ZONE",
    "0406H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, CALIFORNIA VI ZONE",
    "0501": "STATE PLANE LAMBERT CONFORMAL CONIC, COLORADO CENTRAL ZONE",
    "0501H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, COLORADO CENTRAL ZONE",
    "0502": "STATE PLANE LAMBERT CONFORMAL CONIC, COLORADO NORTH ZONE",
    "0502H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, COLORADO NORTH ZONE",
    "0503": "STATE PLANE LAMBERT CONFORMAL CONIC, COLORADO SOUTH ZONE",
    "0503H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, COLORADO SOUTH ZONE",
    "0600": "STATE PLANE LAMBERT CONFORMAL CONIC, CONNECTICUT ZONE",
    "0600H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, CONNECTICUT ZONE",
    "0700": "STATE PLANE TRANSVERSE MERCATOR, DELAWARE ZONE",
    "0700H": "HARN STATE PLANE TRANSVERSE MERCATOR, DELAWARE ZONE",
    "0901": "STATE PLANE TRANSVERSE MERCATOR, FLORIDA EAST ZONE",
    "0901H": "HARN STATE PLANE TRANSVERSE MERCATOR, FLORIDA EAST ZONE",
    "0902": "STATE PLANE TRANSVERSE MERCATOR, FLORIDA WEST ZONE",
    "0902H": "HARN STATE PLANE TRANSVERSE MERCATOR, FLORIDA WEST ZONE",
    "1001": "STATE PLANE TRANSVERSE MERCATOR, GEORGIA EAST ZONE",
    "1001H": "HARN STATE PLANE TRANSVERSE MERCATOR, GEORGIA EAST ZONE",
    "1002": "STATE PLANE TRANSVERSE MERCATOR, GEORGIA WEST ZONE",
    "1002H": "HARN STATE PLANE TRANSVERSE MERCATOR, GEORGIA WEST ZONE",
    "1101": "STATE PLANE TRANSVERSE MERCATOR, IDAHO EAST ZONE",
    "1101H": "HARN STATE PLANE TRANSVERSE MERCATOR, IDAHO EAST ZONE",
    "1102": "STATE PLANE TRANSVERSE MERCATOR, IDAHO CENTRAL ZONE",
    "1102H": "HARN STATE PLANE TRANSVERSE MERCATOR, IDAHO CENTRAL ZONE",
    "1103": "STATE PLANE TRANSVERSE MERCATOR, IDAHO WEST ZONE",
    "1103H": "HARN STATE PLANE TRANSVERSE MERCATOR, IDAHO WEST ZONE",
    "1201": "STATE PLANE TRANSVERSE MERCATOR, ILLINOIS EAST ZONE",
    "1201H": "HARN STATE PLANE TRANSVERSE MERCATOR, ILLINOIS EAST ZONE",
    "1202": "STATE PLANE TRANSVERSE MERCATOR, ILLINOIS WEST ZONE",
    "1202H": "HARN STATE PLANE TRANSVERSE MERCATOR, ILLINOIS WEST ZONE",
    "1301": "STATE PLANE TRANSVERSE MERCATOR, INDIANA EAST ZONE",
    "1301H": "HARN STATE PLANE TRANSVERSE MERCATOR, INDIANA EAST ZONE",
    "1302": "STATE PLANE TRANSVERSE MERCATOR, INDIANA WEST ZONE",
    "1302H": "HARN STATE PLANE TRANSVERSE MERCATOR, INDIANA WEST ZONE",
    "1401": "STATE PLANE LAMBERT CONFORMAL CONIC, IOWA NORTH ZONE",
    "1402": "STATE PLANE LAMBERT CONFORMAL CONIC, IOWA SOUTH ZONE",
    "1501": "STATE PLANE LAMBERT CONFORMAL CONIC, KANSAS NORTH ZONE",
    "1501H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, KANSAS NORTH ZONE",
    "1502": "STATE PLANE LAMBERT CONFORMAL CONIC, KANSAS SOUTH ZONE",
    "1502H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, KANSAS SOUTH ZONE",
    "1600": "STATE PLANE LAMBERT CONFORMAL CONIC, KENTUCKY ZONE",
    "1601": "STATE PLANE LAMBERT CONFORMAL CONIC, KENTUCKY NORTH ZONE",
    "1601H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, KENTUCKY NORTH ZONE",
    "1602": "STATE PLANE LAMBERT CONFORMAL CONIC, KENTUCKY SOUTH ZONE",
    "1602H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, KENTUCKY SOUTH ZONE",
    "1701": "STATE PLANE LAMBERT CONFORMAL CONIC, LOUISIANA NORTH ZONE",
    "1701H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, LOUISIANA NORTH ZONE",
    "1702": "STATE PLANE LAMBERT CONFORMAL CONIC, LOUISIANA SOUTH ZONE",
    "1702H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, LOUISIANA SOUTH ZONE",
    "1801": "STATE PLANE TRANSVERSE MERCATOR, MAINE EAST ZONE",
    "1801H": "HARN STATE PLANE TRANSVERSE MERCATOR, MAINE EAST ZONE",
    "1802": "STATE PLANE TRANSVERSE MERCATOR, MAINE WEST ZONE",
    "1802H": "HARN STATE PLANE TRANSVERSE MERCATOR, MAINE WEST ZONE",
    "1900": "STATE PLANE LAMBERT CONFORMAL CONIC, MARYLAND ZONE",
    "1900H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, MARYLAND ZONE",
    "2001": "STATE PLANE LAMBERT CONFORMAL CONIC, MASSACHUSETTS ISLAND ZONE",
    "2001H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, MASSACHUSETTS ISLAND ZONE",
    "2002": "STATE PLANE LAMBERT CONFORMAL CONIC, MASSACHUSETTS MLAND ZONE",
    "2002H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, MASSACHUSETTS MLAND ZONE",
    "2111": "STATE PLANE LAMBERT CONFORMAL CONIC, MICHIGAN NORTH ZONE",
    "2111H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, MICHIGAN NORTH ZONE",
    "2112": "STATE PLANE LAMBERT CONFORMAL CONIC, MICHIGAN CENTRAL ZONE",
    "2112H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, MICHIGAN CENTRAL ZONE",
    "2113": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, MICHIGAN SOUTH ZONE",
    "2201": "STATE PLANE LAMBERT CONFORMAL CONIC, MINNESOTA NORTH ZONE",
    "2202": "STATE PLANE LAMBERT CONFORMAL CONIC, MINNESOTA CENTRAL ZONE",
    "2203": "STATE PLANE LAMBERT CONFORMAL CONIC, MINNESOTA SOUTH ZONE",
    "2301": "STATE PLANE TRANSVERSE MERCATOR, MISSISSIPPI EAST ZONE",
    "2301H": "HARN STATE PLANE TRANSVERSE MERCATOR, MISSISSIPPI EAST ZONE",
    "2302": "STATE PLANE TRANSVERSE MERCATOR, MISSISSIPPI WEST ZONE",
    "2302H": "HARN STATE PLANE TRANSVERSE MERCATOR, MISSISSIPPI WEST ZONE",
    "2401": "STATE PLANE TRANSVERSE MERCATOR, MISSOURI EAST ZONE",
    "2402": "STATE PLANE TRANSVERSE MERCATOR, MISSOURI CENTRAL ZONE",
    "2403": "STATE PLANE TRANSVERSE MERCATOR, MISSOURI WEST ZONE",
    "2500": "STATE PLANE LAMBERT CONFORMAL CONIC, MONTANA ZONE",
    "2500H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, MONTANA ZONE",
    "2600": "STATE PLANE LAMBERT CONFORMAL CONIC, NEBRASKA ZONE",
    "2600H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, NEBRASKA ZONE",
    "2701": "STATE PLANE TRANSVERSE MERCATOR, NEVADA EAST ZONE",
    "2701H": "HARN STATE PLANE TRANSVERSE MERCATOR, NEVADA EAST ZONE",
    "2702": "STATE PLANE TRANSVERSE MERCATOR, NEVADA CENTRAL ZONE",
    "2702H": "HARN STATE PLANE TRANSVERSE MERCATOR, NEVADA CENTRAL ZONE",
    "2703": "STATE PLANE TRANSVERSE MERCATOR, NEVADA WEST ZONE",
    "2703H": "HARN STATE PLANE TRANSVERSE MERCATOR, NEVADA WEST ZONE",
    "2800": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, NEW HAMPSHIRE ZONE",
    "2900": "STATE PLANE TRANSVERSE MERCATOR, NEW JERSEY ZONE",
    "2900H": "HARN STATE PLANE TRANSVERSE MERCATOR, NEW JERSEY ZONE",
    "3001": "STATE PLANE TRANSVERSE MERCATOR, NEW MEXICO EAST ZONE",
    "3001H": "HARN STATE PLANE TRANSVERSE MERCATOR, NEW MEXICO EAST ZONE",
    "3002": "STATE PLANE TRANSVERSE MERCATOR, NEW MEXICO CENTRAL ZONE",
    "3002H": "HARN STATE PLANE TRANSVERSE MERCATOR, NEW MEXICO CENTRAL ZONE",
    "3003": "STATE PLANE TRANSVERSE MERCATOR, NEW MEXICO WEST ZONE",
    "3003H": "HARN STATE PLANE TRANSVERSE MERCATOR, NEW MEXICO WEST ZONE",
    "3101": "STATE PLANE TRANSVERSE MERCATOR, NEW YORK EAST ZONE",
    "3101H": "HARN STATE PLANE TRANSVERSE MERCATOR, NEW YORK EAST ZONE",
    "3102": "STATE PLANE TRANSVERSE MERCATOR, NEW YORK CENTRAL ZONE",
    "3102H": "HARN STATE PLANE TRANSVERSE MERCATOR, NEW YORK CENTRAL ZONE",
    "3103": "STATE PLANE TRANSVERSE MERCATOR, NEW YORK WEST ZONE",
    "3103H": "HARN STATE PLANE TRANSVERSE MERCATOR, NEW YORK WEST ZONE",
    "3104": "STATE PLANE LAMBERT CONFORMAL CONIC, NEW YORK LONG ISLAND ZONE",
    "3104H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, NEW YORK LONG ISLAND ZONE",
    "3200": "STATE PLANE LAMBERT CONFORMAL CONIC, NORTH CAROLINA ZONE",
    "3301": "STATE PLANE LAMBERT CONFORMAL CONIC, NORTH DAKOTA NORTH ZONE",
    "3301H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, NORTH DAKOTA NORTH ZONE",
    "3302": "STATE PLANE LAMBERT CONFORMAL CONIC, NORTH DAKOTA SOUTH ZONE",
    "3302H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, NORTH DAKOTA SOUTH ZONE",
    "3401": "STATE PLANE LAMBERT CONFORMAL CONIC, OHIO NORTH ZONE",
    "3401H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, OHIO NORTH ZONE",
    "3402": "STATE PLANE LAMBERT CONFORMAL CONIC, OHIO SOUTH ZONE",
    "3402H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, OHIO SOUTH ZONE",
    "3501": "STATE PLANE LAMBERT CONFORMAL CONIC, OKLAHOMA NORTH ZONE",
    "3501H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, OKLAHOMA NORTH ZONE",
    "3502": "STATE PLANE LAMBERT CONFORMAL CONIC, OKLAHOMA SOUTH ZONE",
    "3502H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, OKLAHOMA SOUTH ZONE",
    "3601": "STATE PLANE LAMBERT CONFORMAL CONIC, OREGON NORTH ZONE",
    "3601H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, OREGON NORTH ZONE",
    "3602": "STATE PLANE LAMBERT CONFORMAL CONIC, OREGON SOUTH ZONE",
    "3602H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, OREGON SOUTH ZONE",
    "3701": "STATE PLANE LAMBERT CONFORMAL CONIC, PENNSYLVANIA NORTH ZONE",
    "3702": "STATE PLANE LAMBERT CONFORMAL CONIC, PENNSYLVANIA SOUTH ZONE",
    "3800": "STATE PLANE TRANSVERSE MERCATOR, RHODE ISLAND ZONE",
    "3800H": "HARN STATE PLANE TRANSVERSE MERCATOR, RHODE ISLAND ZONE",
    "3900": "STATE PLANE LAMBERT CONFORMAL CONIC, SOUTH CAROLINA ZONE",
    "4001": "STATE PLANE LAMBERT CONFORMAL CONIC, SOUTH DAKOTA NORTH ZONE",
    "4001H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, SOUTH DAKOTA NORTH ZONE",
    "4002": "STATE PLANE LAMBERT CONFORMAL CONIC, SOUTH DAKOTA SOUTH ZONE",
    "4002H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, SOUTH DAKOTA SOUTH ZONE",
    "4100": "STATE PLANE LAMBERT CONFORMAL CONIC, TENNESSEE ZONE",
    "4100H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, TENNESSEE ZONE",
    "4201": "STATE PLANE LAMBERT CONFORMAL CONIC, TEXAS NORTH ZONE",
    "4201H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, TEXAS NORTH ZONE",
    "4202": "STATE PLANE LAMBERT CONFORMAL CONIC, TEXAS NORTH CENTRAL ZONE",
    "4202H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, TEXAS NORTH CENTRAL ZONE",
    "4203": "STATE PLANE LAMBERT CONFORMAL CONIC, TEXAS CENTRAL ZONE",
    "4203H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, TEXAS CENTRAL ZONE",
    "4204": "STATE PLANE LAMBERT CONFORMAL CONIC, TEXAS SOUTH CENTRAL ZONE",
    "4204H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, TEXAS SOUTH CENTRAL ZONE",
    "4205": "STATE PLANE LAMBERT CONFORMAL CONIC, TEXAS SOUTH ZONE",
    "4205H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, TEXAS SOUTH ZONE",
    "4301": "STATE PLANE LAMBERT CONFORMAL CONIC, UTAH NORTH ZONE",
    "4301H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, UTAH NORTH ZONE",
    "4302": "STATE PLANE LAMBERT CONFORMAL CONIC, UTAH CENTRAL ZONE",
    "4302H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, UTAH CENTRAL ZONE",
    "4303": "STATE PLANE LAMBERT CONFORMAL CONIC, UTAH SOUTH ZONE",
    "4303H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, UTAH SOUTH ZONE",
    "4400": "STATE PLANE TRANSVERSE MERCATOR, VERMONT ZONE",
    "4400H": "HARN STATE PLANE TRANSVERSE MERCATOR, VERMONT ZONE",
    "4501": "STATE PLANE LAMBERT CONFORMAL CONIC, VIRGINIA NORTH ZONE",
    "4501H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, VIRGINIA NORTH ZONE",
    "4502": "STATE PLANE LAMBERT CONFORMAL CONIC, VIRGINIA SOUTH ZONE",
    "4502H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, VIRGINIA SOUTH ZONE",
    "4601": "STATE PLANE LAMBERT CONFORMAL CONIC, WASHINGTON NORTH ZONE",
    "4601H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, WASHINGTON NORTH ZONE",
    "4602": "STATE PLANE LAMBERT CONFORMAL CONIC, WASHINGTON SOUTH ZONE",
    "4602H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, WASHINGTON SOUTH ZONE",
    "4701": "STATE PLANE LAMBERT CONFORMAL CONIC, WEST VIRGINIA NORTH ZONE",
    "4701H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, WEST VIRGINIA NORTH ZONE",
    "4702": "STATE PLANE LAMBERT CONFORMAL CONIC, WEST VIRGINIA SOUTH ZONE",
    "4702H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, WEST VIRGINIA SOUTH ZONE",
    "4801": "STATE PLANE LAMBERT CONFORMAL CONIC, WISCONSIN NORTH ZONE",
    "4801H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, WISCONSIN NORTH ZONE",
    "4802": "STATE PLANE LAMBERT CONFORMAL CONIC, WISCONSIN CENTRAL ZONE",
    "4802H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, WISCONSIN CENTRAL ZONE",
    "4803": "STATE PLANE LAMBERT CONFORMAL CONIC, WISCONSIN SOUTH ZONE",
    "4803H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, WISCONSIN SOUTH ZONE",
    "4901": "STATE PLANE TRANSVERSE MERCATOR, WYOMING EAST ZONE",
    "4901H": "HARN STATE PLANE TRANSVERSE MERCATOR, WYOMING EAST ZONE",
    "4902": "STATE PLANE TRANSVERSE MERCATOR, WYOMING EAST CENTRAL ZONE",
    "4902H": "HARN STATE PLANE TRANSVERSE MERCATOR, WYOMING EAST CENTRAL ZONE",
    "4903": "STATE PLANE TRANSVERSE MERCATOR, WYOMING WEST CENTRAL ZONE",
    "4903H": "HARN STATE PLANE TRANSVERSE MERCATOR, WYOMING WEST CENTRAL ZONE",
    "4904": "STATE PLANE TRANSVERSE MERCATOR, WYOMING WEST ZONE",
    "4904H": "HARN STATE PLANE TRANSVERSE MERCATOR, WYOMING WEST ZONE",
    "5001": "STATE PLANE HOTINE OBLIQUE MERCATOR, ALASKA 1 ZONE",
    "5002": "STATE PLANE TRANSVERSE MERCATOR, ALASKA 2 ZONE",
    "5003": "STATE PLANE TRANSVERSE MERCATOR, ALASKA 3 ZONE",
    "5004": "STATE PLANE TRANSVERSE MERCATOR, ALASKA 4 ZONE",
    "5005": "STATE PLANE TRANSVERSE MERCATOR, ALASKA 5 ZONE",
    "5006": "STATE PLANE TRANSVERSE MERCATOR, ALASKA 6 ZONE",
    "5007": "STATE PLANE TRANSVERSE MERCATOR, ALASKA 7 ZONE",
    "5008": "STATE PLANE TRANSVERSE MERCATOR, ALASKA 8 ZONE",
    "5009": "STATE PLANE TRANSVERSE MERCATOR, ALASKA 9 ZONE",
    "5010": "STATE PLANE LAMBERT CONFORMAL CONIC, ALASKA 10 ZONE",
    "5101": "STATE PLANE TRANSVERSE MERCATOR, HAWAII 1 ZONE",
    "5101H": "HARN STATE PLANE TRANSVERSE MERCATOR, HAWAII 1 ZONE",
    "5102": "STATE PLANE TRANSVERSE MERCATOR, HAWAII 2 ZONE",
    "5102H": "HARN STATE PLANE TRANSVERSE MERCATOR, HAWAII 2 ZONE",
    "5103": "STATE PLANE TRANSVERSE MERCATOR, HAWAII 3 ZONE",
    "5103H": "HARN STATE PLANE TRANSVERSE MERCATOR, HAWAII 3 ZONE",
    "5104": "STATE PLANE TRANSVERSE MERCATOR, HAWAII 4 ZONE",
    "5104H": "HARN STATE PLANE TRANSVERSE MERCATOR, HAWAII 4 ZONE",
    "5105": "STATE PLANE TRANSVERSE MERCATOR, HAWAII 5 ZONE",
    "5105H": "HARN STATE PLANE TRANSVERSE MERCATOR, HAWAII 5 ZONE",
    "5200": "STATE PLANE LAMBERT CONFORMAL CONIC, PR VIRGIN ISLANDS ZONE",
    "5200H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, PR VIRGIN ISLANDS ZONE",
    "5201": "STATE PLANE LAMBERT CONFORMAL CONIC, PUERTO RICO ZONE",
    "5202": "STATE PLANE LAMBERT CONFORMAL CONIC, PR VIRGIN I ST CROIX ZONE",
    "5300": "STATE PLANE LAMBERT CONFORMAL CONIC, AMERICAN SAMOA ZONE",
    "5400": "STATE PLANE POLYCONIC, GUAM ZONE", "9000": "OTHER",
    "903": "STATE PLANE LAMBERT CONFORMAL CONIC, FLORIDA NORTH ZONE",
    "903H": "HARN STATE PLANE LAMBERT CONFORMAL CONIC, FLORIDA NORTH ZONE",
    "GCS": "GEOGRAPHIC COORDINATE SYSTEM", "NP": "NP", "UTM": "UNIVERSAL TRANSVERSE MERCATOR",
    "WGS": "WGS 1984 WEB MERCATOR (AUXILIARY SPHERE)"}
D_Projzone = {
    "0101": "0101", "0101H": "0101H", "0102": "0102", "0102H": "0102H", "0201": "0201",
    "0201H": "0201H", "0202": "0202", "0202H": "0202H", "0203": "0203", "0203H": "0203H",
    "0301": "0301", "0302": "0302", "0401": "0401", "0401H": "0401H", "0402": "0402",
    "0402H": "0402H", "0403": "0403", "0403H": "0403H", "0404": "0404", "0404H": "0404H",
    "0405": "0405", "0405H": "0405H", "0406": "0406", "0406H": "0406H", "0501": "0501",
    "0501H": "0501H", "0502": "0502", "0502H": "0502H", "0503": "0503", "0503H": "0503H",
    "0600": "0600", "0600H": "0600H", "0700": "0700", "0901": "0901", "0902": "0902",
    "0903": "0903", "0903H": "0903H", "1": "1", "10": "10", "1001": "1001", "1001H": "1001H",
    "1002": "1002", "1002H": "1002H", "11": "11", "1101": "1101", "1101H": "1101H",
    "1102": "1102", "1102H": "1102H", "1103": "1103", "1103H": "1103H", "12": "12",
    "1201": "1201", "1201H": "1201H", "1202": "1202", "1202H": "1202H", "13": "13",
    "1301": "1301", "1301H": "1301H", "1302": "1302", "1302H": "1302H", "14": "14",
    "1401": "1401", "1402": "1402", "15": "15", "1501": "1501", "1501H": "1501H",
    "1502": "1502", "1502H": "1502H", "16": "16", "1600": "1600", "1601": "1601",
    "1601H": "1601H", "1602": "1602", "1602H": "1602H", "17": "17", "1701": "1701",
    "1701H": "1701H", "1702": "1702", "1702H": "1702H", "18": "18", "1801": "1801",
    "1801H": "1801H", "1802": "1802", "1802H": "1802H", "19": "19", "1900": "1900",
    "1900H": "1900H", "2": "2", "2001": "2001", "2001H": "2001H", "2002": "2002",
    "2002H": "2002H", "2111": "2111", "2111H": "2111H", "2112": "2112", "2112H": "2112H",
    "2113": "2113", "2113H": "2113H", "2201": "2201", "2202": "2202", "2203": "2203",
    "2301": "2301", "2301H": "2301H", "2302": "2302", "2302H": "2302H", "2401": "2401",
    "2402": "2402", "2403": "2403", "2500": "2500", "2500H": "2500H", "2600": "2600",
    "2600H": "2600H", "2701": "2701", "2701H": "2701H", "2702": "2702", "2702H": "2702H",
    "2703": "2703", "2703H": "2703H", "2800": "2800", "2800H": "2800H", "2900": "2900",
    "2900H": "2900H", "3": "3", "3001": "3001", "3001H": "3001H", "3002": "3002",
    "3002H": "3002H", "3003": "3003", "3003H": "3003H", "3101": "3101", "3101H": "3101H",
    "3102": "3102", "3102H": "3102H", "3103": "3103", "3103H": "3103H", "3104": "3104",
    "3104H": "3104H", "3200": "3200", "3301": "3301", "3301H": "3301H", "3302": "3302",
    "3302H": "3302H", "3401": "3401", "3401H": "3401H", "3402": "3402", "3402H": "3402H",
    "3501": "3501", "3501H": "3501H", "3502": "3502", "3502H": "3502H", "3601": "3601",
    "3601H": "3601H", "3602": "3602", "3602H": "3602H", "3701": "3701", "3702": "3702",
    "3800": "3800", "3800H": "3800H", "3900": "3900", "4": "4", "4001": "4001",
    "4001H": "4001H", "4002": "4002", "4002H": "4002H", "4100": "4100", "4100H": "4100H",
    "4201": "4201", "4201H": "4201H", "4202": "4202", "4202H": "4202H", "4203": "4203",
    "4203H": "4203H", "4204": "4204", "4204H": "4204H", "4205": "4205", "4205H": "4205H",
    "4301": "4301", "4301H": "4301H", "4302": "4302", "4302H": "4302H", "4303": "4303",
    "4303H": "4303H", "4400": "4400", "4400H": "4400H", "4501": "4501", "4501H": "4501H",
    "4502": "4502", "4502H": "4502H", "4601": "4601", "4601H": "4601H", "4602": "4602",
    "4602H": "4602H", "4701": "4701", "4701H": "4701H", "4702": "4702", "4702H": "4702H",
    "4801": "4801", "4801H": "4801H", "4802": "4802", "4802H": "4802H", "4803": "4803",
    "4803H": "4803H", "4901": "4901", "4901H": "4901H", "4902": "4902", "4902H": "4902H",
    "4903": "4903", "4903H": "4903H", "4904": "4904", "4904H": "4904H", "5": "5",
    "5001": "5001", "5002": "5002", "5003": "5003", "5004": "5004", "5005": "5005",
    "5006": "5006", "5007": "5007", "5008": "5008", "5009": "5009", "5010": "5010",
    "5101": "5101", "5101H": "5101H", "5102": "5102", "5102H": "5102H", "5103": "5103",
    "5103H": "5103H", "5104": "5104", "5104H": "5104H", "5105": "5105", "5105H": "5105H",
    "5200": "5200", "5200H": "5200H", "5201": "5201", "5202": "5202", "5300": "5300",
    "5400": "5400", "6": "6", "7": "7", "8": "8", "9": "9", "9000": "9000", "901H": "901H",
    "902H": "902H", "NP": "NP"}
D_Quad_Corner = {"NE": "NE", "NW": "NW", "SE": "SE", "SW": "SW"}
D_Runup_Mdl = {
    "1010": "ACES 1.07 (1992)", "1015": "CEM - OTHER (2003)",
    "1020": "CHAMP / RUNUP 2.0 (2007)", "1030": "EROSION (1998)", "1040": "GLWRM (1992)",
    "1050": "RUNUP 2.0 (1990)", "1060": "TAW Method", "1070": "SPM (USACE 1984)",
    "1080": "DIM", "9000": "OTHER"}
D_Scale = {"1000": "6000", "1010": "12000", "1020": "24000", "2000": "10000"}
D_Shr_Typ = {
    "1000": "Sandy Beach-Small Dune", "1010": "Sandy Beach-Large Dune",
    "1020": "Erodible Bluff", "1030": "Non-Erodible Bluff",
    "1040": "Shore Protection Structure", "1050": "Wetland"}
D_State_FIPS = {
    "01": "01", "02": "02", "04": "04", "05": "05", "06": "06", "08": "08", "09": "09",
    "10": "10", "11": "11", "12": "12", "13": "13", "15": "15", "16": "16", "17": "17",
    "18": "18", "19": "19", "20": "20", "21": "21", "22": "22", "23": "23", "24": "24",
    "25": "25", "26": "26", "27": "27", "28": "28", "29": "29", "30": "30", "31": "31",
    "32": "32", "33": "33", "34": "34", "35": "35", "36": "36", "37": "37", "38": "38",
    "39": "39", "40": "40", "41": "41", "42": "42", "44": "44", "45": "45", "46": "46",
    "47": "47", "48": "48", "49": "49", "50": "50", "51": "51", "53": "53", "54": "54",
    "55": "55", "56": "56", "60": "60", "64": "64", "66": "66", "68": "68", "69": "69",
    "70": "70", "72": "72", "74": "74", "78": "78"}
D_State_Name = {
    "AK": "Alaska", "AL": "Alabama", "AR": "Arkansas", "AS": "American Samoa", "AZ": "Arizona",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DC": "District of Columbia",
    "DE": "Delaware", "FL": "Florida", "FM": "Micronesia", "GA": "Georgia", "GU": "Guam",
    "HI": "Hawaii", "IA": "Iowa", "ID": "Idaho", "IL": "Illinois", "IN": "Indiana",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "MA": "Massachusetts",
    "MD": "Maryland", "ME": "Maine", "MH": "Marshall Islands", "MI": "Michigan",
    "MN": "Minnesota", "MO": "Missouri", "MP": "Northern Mariana Islands", "MS": "Mississippi",
    "MT": "Montana", "NC": "North Carolina", "ND": "North Dakota", "NE": "Nebraska",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NP": "NP", "NV": "Nevada",
    "NY": "New York", "OH": "Ohio", "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania",
    "PR": "Puerto Rico", "PW": "Palau", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UM": "U.S. Minor Islands",
    "UT": "Utah", "VA": "Virginia", "VI": "Virgin Islands", "VT": "Vermont",
    "WA": "Washington", "WI": "Wisconsin", "WV": "West Virginia", "WY": "Wyoming"}
D_Struct_Face = {"DN": "Downstream", "UNK": "Unknown", "UP": "Upstream"}
D_Struct_Mtl = {
    "1000": "Stone", "1010": "Asphalt", "1020": "Concrete", "1030": "Earthen",
    "1040": "Timber", "1050": "Steel", "1060": "Sand", "1070": "Other", "1080": "Unknown",
    "NP": "NP"}
D_Struct_Typ = {
    "1000": "Aqueduct", "1001": "Bridge", "1002": "Canal", "1003": "Channel",
    "1006": "Control Structure", "1007": "Culvert", "1010": "Dam", "1011": "Dike",
    "1012": "Dock", "1013": "Drop Structure", "1014": "Energy Dissipater",
    "1015": "Fish Ladder", "1017": "Flume", "1018": "Footbridge", "1019": "Gate",
    "1020": "Jetty", "1021": "Levee", "1022": "Lock", "1023": "Penstock", "1024": "Pier",
    "1025": "Pump Station", "1026": "Seawall", "1027": "Side Weir Structure",
    "1028": "Storm Sewer", "1029": "Utility Crossing", "1030": "Weir", "1031": "Wing Wall",
    "1032": "1 PCT Annual Chance Flood Discharge Contained In Structure",
    "1033": "0.2 PCT Annual Chance Flood Discharge Contained In Structure",
    "1036": "Floodway Contained In Structure", "1037": "Pipeline", "1038": "Retaining Wall",
    "1039": "Revetment", "1040": "Siphon", "9000": "Other / Misc Structure", "NP": "NP"}
D_Struct_Typ_2020 = {
    "1000": "Aqueduct", "1001": "Bridge", "1002": "Canal", "1003": "Channel",
    "1006": "Control Structure", "1007": "Culvert", "1010": "Dam",
    "1012": "Dock", "1013": "Drop Structure", "1014": "Energy Dissipater",
    "1015": "Fish Ladder", "1017": "Flume", "1018": "Footbridge", "1019": "Gate",
    "1020": "Jetty", "1021": "Levee", "1022": "Lock", "1023": "Penstock", "1024": "Pier",
    "1025": "Pump Station", "1026": "Seawall", "1027": "Side Weir Structure",
    "1028": "Storm Sewer", "1029": "Utility Crossing", "1030": "Weir", "1031": "Wing Wall",
    "1032": "1 PCT Annual Chance Flood Discharge Contained In Structure",
    "1033": "0.2 PCT Annual Chance Flood Discharge Contained In Structure",
    "1036": "Floodway Contained In Structure", "1037": "Pipeline", "1038": "Retaining Wall",
    "1039": "Revetment", "1040": "Siphon", "9000": "Other / Misc Structure", "NP": "NP"}
D_Study_Mth = {
    "1000": "New H&H", "1010": "BLE TIER A", "1020": "BLE TIER B", "1030": "BLE TIER C",
    "1040": "BLE TIER D", "1050": "BLE TIER E", "1060": "LSAE", "1100": "REDELINEATION",
    "1200": "DIGITAL CONVERSION", "1300": "SECLUSION", "NP": "NP"}
D_Study_Prefix = {
    "0100": "Borough of", "0200": "City and County of", "0300": "City of",
    "0400": "Municipality of", "0500": "Town of", "0600": "Township of", "0700": "Village of",
    "0800": "Town and Village of", "9000": "Other"}
D_Study_Typ = {
    "1000": "SFHA without BFE", "1010": "SFHA with BFE published only in FIS",
    "1030": "BLE available but unpublished", "1040": "SFHA with unpublished BFE",
    "1050": "SFHA with BFE no floodway", "1060": "SFHA with BFE and floodway",
    "1070": "Shaded Zone X with depths less than 1'", "NP": "NP"}
D_Study_Typ_2017 = {
    "1000": "SFHAs WITH LOW FLOOD RISK", "1010": "SFHAs WITH MEDIUM FLOOD RISK",
    "1020": "SFHAs WITH HIGH FLOOD RISK", "1100": "REDELINEATION",
    "1200": "DIGITAL CONVERSION", "NP": "NP"}
D_Subbasin_Typ = {"HUC8": "USGS HUC-8", "HYD": "Hydrologic Analyses", "NP": "NP"}
D_Surge_Mdl = {
    "0100": "ADCIRC (2003)", "0110": "DELFT 3D", "1010": "DYNLET", "1020": "FEMA Surge (1988)",
    "1040": "MIKE 21 HD/NHD", "1050": "New England Tide Profile",
    "1080": "TABS RMA V.4.3 (Oct 1996)", "1090": "USACE Great Lakes Tide Profile (1988)"}
D_Task_Typ_2017 = {
    "1000": "ALLUVIAL FAN", "1010": "BASE MAP", "1020": "COASTAL", "1030": "FIRM DATABASE",
    "1040": "FLOODPLAIN MAPPING", "1050": "HYDRAULIC", "1060": "HYDROLOGIC", "1070": "SURVEY",
    "1080": "TERRAIN", "1090": "DISCOVERY", "1100": "FLOOD RISK ASSESSMENT", "1200": "LOMR",
    "NP": "NP"}
D_Task_Typ = {
    "1000": "ALLUVIAL FAN", "1010": "BASE MAP", "1020": "COASTAL", "1030": "FIRM DATABASE",
    "1040": "FLOODPLAIN MAPPING", "1050": "HYDRAULIC", "1060": "HYDROLOGIC", "1070": "SURVEY",
    "1080": "NEW TOPO CAPTURE", "1081": "EXISTING TOPO CAPTURE", "1082": "TERRAIN CAPTURE",
    "1090": "DISCOVERY", "1100": "FLOOD RISK ASSESSMENT", "1200": "LOMR",
    "1300": "Levee Seclusion", "NP": "NP"}
D_Time_Units = {
    "1000": "DAYS", "1010": "HOURS", "1020": "MINUTES", "1030": "MONTHS", "1040": "SECONDS",
    "1050": "WEEKS", "1060": "YEARS"}
D_TrueFalse = {"F": "F", "T": "T", "U": "U"}
D_TsctBasln_Typ = {
    "1000": "Zero Foot Contour Field Survey", "1010": "Zero Foot Contour LiDAR / Shoals",
    "1020": "Zero Foot Contour Referenced To Tidal Datum", "9000": "Other Source", "NP": "NP"}
D_USACE_District = {
    "1001": "Alaska", "1002": "Albuquerque", "1003": "Baltimore", "1004": "Buffalo",
    "1005": "Charleston", "1006": "Chicago", "1007": "Detroit", "1008": "Fort Worth",
    "1009": "Galveston", "1010": "Gulf Region", "1011": "Honolulu", "1012": "Huntington",
    "1013": "Jacksonville", "1014": "Kansas City", "1015": "Little Rock", "1016": "Louisville",
    "1017": "Los Angeles", "1018": "Memphis", "1019": "Mobile", "1020": "Nashville",
    "1021": "New England", "1022": "New Orleans", "1023": "New York", "1024": "Norfolk",
    "1025": "Omaha", "1026": "Philadelphia", "1027": "Pittsburgh", "1028": "Portland",
    "1029": "Rock Island", "1030": "Sacramento", "1031": "San Francisco", "1032": "Savannah",
    "1033": "Seattle", "1034": "St. Louis", "1035": "St. Paul", "1036": "Tulsa",
    "1037": "Vicksburg", "1038": "Walla Walla", "1039": "Wilmington", "NP": "NP"}
D_VZone = {
    "1010": "Wave Overtopping Splash Zone", "1020": "PFD", "1030": "Runup",
    "1050": "High Velocity Flow", "1060": "Breaking Wave Ht"}
D_V_Datum = {
    "MLLW": "MLLW", "MLW": "MLW", "MSL": "MSL", "NAVD88": "NAVD88", "NGVD29": "NGVD29",
    "NP": "NP", "TIDAL": "LOCAL TIDAL DATUM"}
D_Velocity_Units = {
    "1000": "Centimeters / Day", "1010": "Centimeters / Hour", "1020": "Feet / Second",
    "1030": "Inches / Day", "1040": "Inches / Hour", "1050": "Meters / Second",
    "1060": "Micrometers / Second", "1070": "Millimeters / Day", "1080": "Millimeters / Hour",
    "NP": "NP"}
D_Wave_Mdl = {
    "1010": "DELFT 3D", "1020": "MIKE 21 (OSW)", "1030": "MIKE 21 (NSW)",
    "1040": "RCPWAVE (1986)", "1050": "WHAFIS 3.0 (1988)", "1060": "WHAFIS 3.0 GL (1983)",
    "1065": "WHAFIS 4.0 (2007)", "1070": "STWAVE (Latest Version)", "1080": "SWAN (2008)"}
D_XS_Ln_Typ = {
    "1010": "LETTERED, MAPPED", "1020": "NOT LETTERED, MAPPED",
    "1030": "NOT LETTERED, NOT MAPPED"}
D_Zone = {
    "A": "A", "A99": "A99", "AE": "AE", "AH": "AH", "ANI": "AREA NOT INCLUDED", "AO": "AO",
    "AR": "AR", "D": "D", "OW": "OPEN WATER", "V": "V", "VE": "VE", "X": "X"}
D_Zone_Subtyp = {
    "0100": "COASTAL FLOODPLAIN",
    "0110": "RIVERINE FLOODPLAIN",
    "0120": "COMBINED RIVERINE AND COASTAL FLOODPLAIN",
    "0200": "1 PCT ANNUAL CHANCE FLOOD HAZARD CONTAINED IN STRUCTURE",
    "0210": "1 PCT ANNUAL CHANCE FLOOD HAZARD CONTAINED IN CHANNEL",
    "0300": "1 PCT FUTURE CONDITIONS",
    "0310": "1 PCT FUTURE CONDITIONS CONTAINED IN STRUCTURE",
    "0400": "1 PCT DEPTH LESS THAN 1 FOOT",
    "0410": "1 PCT DRAINAGE AREA LESS THAN 1 SQUARE MILE",
    "0500": "0.2 PCT ANNUAL CHANCE FLOOD HAZARD",
    "0510": "0.2 PCT ANNUAL CHANCE FLOOD HAZARD CONTAINED IN STRUCTURE",
    "0520": "0.2 PCT ANNUAL CHANCE FLOOD HAZARD CONTAINED IN CHANNEL",
    "0530": "0.2 PCT ANNUAL CHANCE FLOOD HAZARD IN COASTAL ZONE",
    "0540": "0.2 PCT ANNUAL CHANCE FLOOD HAZARD IN COMBINED RIVERINE AND COASTAL ZONE",
    "1000": "AREA WITH REDUCED FLOOD RISK DUE TO LEVEE",
    "1010": "ADMINISTRATIVE FLOODWAY",
    "1020": "AREA OF SPECIAL CONSIDERATION",
    "1030": "COMMUNITY ENCROACHMENT AREA",
    "1040": "COLORADO RIVER FLOODWAY",
    "1050": "DENSITY FRINGE AREA",
    "1100": "FLOODWAY",
    "1110": "FLOODWAY CONTAINED IN STRUCTURE",
    "1120": "FLOODWAY CONTAINED IN CHANNEL",
    "1200": "FLOWAGE EASEMENT AREA",
    "1210": "STATE ENCROACHMENT AREA",
    "1220": "RIVERINE FLOODWAY SHOWN IN COASTAL ZONE",
    "1230": "NARROW FLOODWAY",
    "1240": "RIVERINE FLOODWAY IN COMBINED RIVERINE AND COASTAL ZONE",
    "2000": "AREA OF MINIMAL FLOOD HAZARD",
    "3000": "AREA WITH FLOOD RISK DUE TO LEVEE"}
D_Zone_Subtyp_2020 = D_Zone_Subtyp.copy()
D_Zone_Subtyp_2020["3000"] = \
    "AREA WITH FLOOD HAZARD DUE TO NON-ACCREDITED LEVEE SYSTEM"
D_Zone_Subtyp_2020["3010"] = \
    "AREA WITH REDUCED FLOOD HAZARD DUE TO PROVISIONALLY ACCREDITED LEVEE SYSTEM"
D_Zone_Subtyp_2020["3020"] = \
    "AREA WITH UNDETERMINED FLOOD HAZARD DUE TO NON-ACCREDITED LEVEE SYSTEM"
D_Zone_Subtyp_2021 = D_Zone_Subtyp_2020.copy()
D_Zone_Subtyp_2021["3030"] = \
    "AREA WITH REDUCED FLOOD HAZARD DUE TO ACCREDITED LEVEE"

# Contains dictionaries of the FEMA tables, the fields that are the keys and the
# domains associated with that field.
L_Comm_Info_domains = {
    'REPOS_ST': D_State_Name, 'REVISIONS': D_TrueFalse,
    'MULTICO_TF': D_TrueFalse, 'FLOODPRONE': D_TrueFalse,
    'FIS_INCLUD': D_TrueFalse}
L_Comm_Revis_domains = {}
L_Cst_Model_domains = {
    'STUDY_TYP': D_Study_Typ, 'SURGE_MDL': D_Surge_Mdl,
    'WAVEHT_MDL': D_Wave_Mdl, 'RUNUP_MDL': D_Runup_Mdl,
    'EROS_METH': D_Erosion, 'EROS_TF': D_TrueFalse,
    'PFD_TF': D_TrueFalse}
L_Cst_Model_domains_2017 = {
    'STUDY_TYP': D_Study_Typ_2017, 'SURGE_MDL': D_Surge_Mdl,
    'WAVEHT_MDL': D_Wave_Mdl, 'RUNUP_MDL': D_Runup_Mdl,
    'EROS_METH': D_Erosion, 'EROS_TF': D_TrueFalse,
    'PFD_TF': D_TrueFalse}
L_Cst_Struct_domains = {
    'CERT_STAT': D_Cert_Status, 'LEN_UNIT': D_Length_Units,
    'STRUCT_MTL': D_Struct_Mtl}
L_Cst_Tsct_Elev_domains = {'EVENT_TYP': D_Event}
L_ManningsN_domains = {}
L_Meetings_domains = {'MTG_TYP': D_Mtg_Typ}
L_MT2_LOMR_domains = {
    'STATUS': D_LOMC_Status, 'SCALE': D_Scale}
L_Mtg_POC_domains = {
    'CEO': D_TrueFalse, 'FPA': D_TrueFalse,
    'SHMO': D_TrueFalse, 'GIS': D_TrueFalse,
    'STATE': D_State_Name}
L_Pan_Revis_domains = {}
L_Pol_FHBM_domains = {}
L_Profil_Bkwtr_El_domains = {
    'EVENT_TYP': D_Event, 'LEN_UNIT': D_Length_Units,
    'V_DATUM': D_V_Datum}
L_Profil_Label_domains = {
    'ORIENT': D_Prof_Lbl_Orient,
    'ADJUSTED': D_Prof_Lbl_Adjust, 'UNDERLINE': D_TrueFalse,
    'LEN_UNIT': D_Length_Units, 'V_DATUM': D_V_Datum}
L_Profil_Panel_domains = {
    'LEN_UNIT': D_Length_Units, 'V_DATUM': D_V_Datum}
L_Source_Cit_domains = {}
L_Summary_Discharges_domains = {
    'AREA_UNIT': D_Area_Units, 'EVENT_TYP': D_Event,
    'DISCH_UNIT': D_Discharge_Units,
    'WSEL_UNIT': D_Length_Units, 'V_DATUM': D_V_Datum,
    'SHOWN_FIS': D_TrueFalse}
L_Summary_Elevations_domains = {
    'EVENT_TYP': D_Event, 'WSEL_UNIT': D_Length_Units,
    'V_DATUM': D_V_Datum, 'SHOWN_FIS': D_TrueFalse}
L_Survey_Pt_domains = {
    'ELEV_UNIT': D_Length_Units, 'H_DATUM': D_Horiz_Datum,
    'V_DATUM': D_V_Datum, 'PROJECTION': D_Projection,
    'PROJ_ZONE': D_Projzone, 'PROJ_UNIT': D_Proj_Unit}
L_XS_Elev_domains = {
    'AREA_UNIT': D_Area_Units, 'VEL_UNIT': D_Velocity_Units,
    'EVENT_TYP': D_Event, 'LEN_UNIT': D_Length_Units,
    'V_DATUM': D_V_Datum, 'LEVEE_TF': D_TrueFalse,
    'LVSCENARIO': D_Levee_Scenario, 'CALC_WO_BW': D_TrueFalse,
    'EVAL_LN': D_TrueFalse}
L_XS_Struct_domains = {
    'STRUCT_TYP': D_Struct_Typ, 'STRUC_FACE': D_Struct_Face,
    'LEN_UNIT': D_Length_Units, 'V_DATUM': D_V_Datum}
L_XS_Struct_domains_2020 = {
    'STRUCT_TYP': D_Struct_Typ_2020, 'STRUC_FACE': D_Struct_Face,
    'LEN_UNIT': D_Length_Units, 'V_DATUM': D_V_Datum}
S_Alluvial_Fan_domains = {
    'ACTIVE_FAN': D_TrueFalse, 'AREA_UNITS': D_Area_Units,
    'DISCH_UNIT': D_Discharge_Units,
    'VEL_UNIT': D_Velocity_Units,
    'DEPTH_UNIT': D_Length_Units, 'FLD_ZONE': D_Zone,
    'ZONE_SUBTY': D_Zone_Subtyp}
S_Base_Index_domains = {}
S_BFE_domains = {
    'LEN_UNIT': D_Length_Units, 'V_DATUM': D_V_Datum}
S_Datum_Conv_Pt_domains = {
    'QUAD_COR': D_Quad_Corner, 'FROM_DATUM': D_V_Datum,
    'TO_DATUM': D_V_Datum, 'LEN_UNIT': D_Length_Units}
S_Cst_Gage_domains = {
    'TIME_UNIT': D_Time_Units, 'GAGE_TYPE': D_Gage_Typ,
    'V_DATUM': D_V_Datum, 'TIDE_TF': D_TrueFalse,
    'WVDIR_TF': D_TrueFalse, 'WVSPEC_TF': D_TrueFalse,
    'WDSPD_TF': D_TrueFalse, 'WDDIR_TF': D_TrueFalse}
S_Cst_Tsct_Ln_domains = {
    'METHOD': D_Method, 'V_DATUM': D_V_Datum,
    'CSTLN_TYP': D_Cst_Typ, 'BEACH_SET': D_Beachset,
    'SHORE_TYP': D_Shr_Typ, 'EVENT_TYP': D_Event,
    'FTCHLNUNIT': D_Length_Units, 'EROS_METH': D_Erosion,
    'ELEV_UNIT': D_Length_Units, 'WHAFIS_TF': D_TrueFalse,
    'OVERTOP_TF': D_TrueFalse, 'BW_HGT_TF': D_TrueFalse,
    'HVFLOW_TF': D_TrueFalse, 'VZONE_EXT': D_VZone,
    'WAVE_02PCT': D_TrueFalse, 'LEN_UNIT': D_Length_Units,
    'TIME_UNIT': D_Time_Units}
S_Fld_Haz_Ln_domains = {'LN_TYP': D_Ln_Typ}
S_Fld_Haz_Ar_domains = {
    'STUDY_TYP': D_Study_Typ, 'FLD_ZONE': D_Zone,
    'ZONE_SUBTY': D_Zone_Subtyp, 'SFHA_TF': D_TrueFalse,
    'V_DATUM': D_V_Datum, 'LEN_UNIT': D_Length_Units,
    'VEL_UNIT': D_Velocity_Units, 'AR_REVERT': D_Zone,
    'AR_SUBTRV': D_Zone_Subtyp, 'DUAL_ZONE': D_TrueFalse}
# Apply 2020 Updates to S_Fld_Haz_Ar Domains.
S_Fld_Haz_Ar_domains_2020 = S_Fld_Haz_Ar_domains.copy()
S_Fld_Haz_Ar_domains_2020['ZONE_SUBTY'] = D_Zone_Subtyp_2020
# Apply 2021 Updates to S_Fld_Haz_Ar Domains.
S_Fld_Haz_Ar_domains_2021 = S_Fld_Haz_Ar_domains_2020.copy()
S_Fld_Haz_Ar_domains_2021['ZONE_SUBTY'] = D_Zone_Subtyp_2021
S_Fld_Haz_Ar_domains_2017 = {
    'STUDY_TYP': D_Study_Typ_2017, 'FLD_ZONE': D_Zone,
    'ZONE_SUBTY': D_Zone_Subtyp, 'SFHA_TF': D_TrueFalse,
    'V_DATUM': D_V_Datum, 'LEN_UNIT': D_Length_Units,
    'VEL_UNIT': D_Velocity_Units, 'AR_REVERT': D_Zone,
    'AR_SUBTRV': D_Zone_Subtyp, 'DUAL_ZONE': D_TrueFalse}
S_Gage_domains = {
    'GAGE_TYP': D_Gage_Typ, 'TIME_UNIT': D_Time_Units,
    'AREA_UNIT': D_Area_Units}
S_Gen_Struct_domains = {
    'STRUCT_TYP': D_Struct_Typ, 'CST_STRUCT': D_Cst_Struct,
    'SHOWN_FIRM': D_TrueFalse}
S_Gen_Struct_domains_2020 = {
    'STRUCT_TYP': D_Struct_Typ_2020, 'CST_STRUCT': D_Cst_Struct,
    'SHOWN_FIRM': D_TrueFalse}
S_Hydro_Reach_domains = {}
S_HWM_domains = {
    'LEN_UNIT': D_Length_Units, 'V_DATUM': D_V_Datum}
S_FIRM_Pan_domains = {
    'ST_FIPS': D_State_FIPS, 'PANEL_TYP': D_Panel_Typ,
    'SCALE': D_Scale, 'BASE_TYP': D_Basemap_Typ}
S_Label_Ld_domains = {
    'LABEL_TYPE': D_Label_Typ, 'SCALE': D_Scale}
S_Label_Ld_domains_2017 = {
    'LABEL_TYPE': D_Label_Typ_2017, 'SCALE': D_Scale}
S_Label_Pt_domains = {
    'LABEL_TYPE': D_Label_Typ, 'FONT_TYPE': D_Font,
    'FONT_STYLE': D_Font_Style, 'SCALE': D_Scale}
S_Label_Pt_domains_2017 = {
    'LABEL_TYPE': D_Label_Typ_2017, 'FONT_TYPE': D_Font,
    'FONT_STYLE': D_Font_Style, 'SCALE': D_Scale}
S_Levee_domains_2017 = {
    'LEVEE_TYP': D_Levee_Type, 'USACE_LEV': D_TrueFalse,
    'DISTRICT': D_USACE_District, 'PL84_99TF': D_TrueFalse,
    'LEVEE_STAT': D_Levee_Status_2017,
    'LEV_AN_TYP': D_Levee_Analysis_Type_2017,
    'LEN_UNIT': D_Length_Units}
S_Levee_domains = {
    'LEVEE_TYP': D_Levee_Type, 'USACE_LEV': D_TrueFalse,
    'DISTRICT': D_USACE_District, 'PL84_99TF': D_TrueFalse,
    'LEVEE_STAT': D_Levee_Status,
    'LEV_AN_TYP': D_Levee_Analysis_Type,
    'LEN_UNIT': D_Length_Units}
S_LiMWA_domains = {'SHOWN_FIRM': D_TrueFalse}
S_LOMR_domains = {'SCALE': D_Scale, 'STATUS': D_LOMC_Status}
S_Nodes_domains = {'NODE_TYP': D_Node_Typ}
S_PFD_Ln_domains = {'VZONE_LIMT': D_TrueFalse}
S_Pol_Ar_domains = {
    'ST_FIPS': D_State_FIPS,
    'ANI_TF': D_TrueFalse}
S_Profil_Basln_domains = {
    'WATER_TYP': D_Prof_Basln_Typ, 'STUDY_TYP': D_Study_Typ,
    'SHOWN_FIRM': D_TrueFalse, 'DATUM_UNIT': D_Length_Units}
S_Profil_Basln_domains_2017 = {
    'WATER_TYP': D_Prof_Basln_Typ, 'STUDY_TYP': D_Study_Typ_2017,
    'SHOWN_FIRM': D_TrueFalse, 'DATUM_UNIT': D_Length_Units}
S_PLSS_Ar_domains = {}
S_Riv_Mrk_domains = {}
S_Stn_Start_domains = {'LOC_ACC': D_Loc_Accuracy}
S_Subbasins_domains = {
    'AREA_UNIT': D_Area_Units, 'BASIN_TYP': D_Subbasin_Typ}
S_Submittal_Info_domains_2017 = {
    'TASK_TYP': D_Task_Typ_2017,
    'HYDRO_MDL': D_Hydro_Mdl, 'HYDRA_MDL': D_Hydra_Mdl,
    'STUDY_TYP': D_Study_Typ_2017}
S_Submittal_Info_domains = {
    'METHOD_TYP': D_Study_Mth, 'TASK_TYP': D_Task_Typ,
    'HYDRO_MDL': D_Hydro_Mdl, 'HYDRA_MDL': D_Hydra_Mdl}
S_Topo_Confidence_domains = {'CONF_TYPE': D_Obscured}
S_Trnsport_Ln_domains = {
    'MTFCC': D_MTFCC, 'ROUTE_TYP': D_Carto_Trans_Code}
S_Tsct_Basln_domains = {
    'TBASE_TYP': D_TsctBasln_Typ, 'V_DATUM': D_V_Datum}
S_Wtr_Ar_domains = {
    'SHOWN_FIRM': D_TrueFalse, 'SHOWN_INDX': D_TrueFalse}
S_Wtr_Ln_domains = {
    'SHOWN_FIRM': D_TrueFalse, 'SHOWN_INDX': D_TrueFalse}
S_XS_domains = {
    'XS_LN_TYP': D_XS_Ln_Typ, 'LEN_UNIT': D_Length_Units,
    'V_DATUM': D_V_Datum}
Study_Info_domains = {
    'STUDY_PRE': D_Study_Prefix, 'STATE_NM': D_State_Name,
    'JURIS_TYP': D_Jurisdiction_Typ, 'OPP_TF': D_TrueFalse,
    'H_DATUM': D_Horiz_Datum, 'V_DATUM': D_V_Datum,
    'PROJECTION': D_Projection, 'PROJ_ZONE': D_Projzone,
    'PROJ_UNIT': D_Proj_Unit, 'PROJ_SECND': D_Projection,
    'PROJ_SUNIT': D_Proj_Unit, 'PROJ_SZONE': D_Projzone,
    'CW_TF': D_TrueFalse, 'RTROFT_TF': D_TrueFalse}


class QCChecks:
    """Performs a QC check of the attributes of the database tables"""

    def __init__(self, in_workspace, in_folder, in_mip_task, in_schema, in_tables,
                 in_coded_check, in_shp_export, in_excel_export):
        """Constructor: Receives the workspace, an output folder, the MIP task, schema, the tables
           to check, and whether to check coded values or text values for the domain checks."""
        self.workspace = in_workspace  # Contains the feature classes and tables
        self.out_folder = in_folder  # The output folder to contain the error files
        self.mip_task = in_mip_task  # The MIP task to process
        self.schema = in_schema  # Year schema to apply
        self.tables = []  # The tables to be checked
        self.task_tables = []  # Tables to be checked for a specific MIP task
        self.coded_check = ''  # Determines if checking coded values or text values
        self.shp_export = ''  # Determines if the shapefiles should be created
        self.excel_export = ''  # Determines if the excel file should be created
        self.dbf_ext = ''  # Extension for DBF tables if shapefiles are used
        self.shp_ext = ''  # Extension for Shapefiles if used
        self.dataset = ''  # Populated if a feature dataset is used
        self.errors = []  # Hold the errors for the current table being processed
        self.total_errors = 0  # Total number of errors found
        self.missing_field = False  # Flag to determine if fields are missing

        # List of acceptable tables to check
        self.acceptable_tables = ['l_comm_info', 'l_comm_revis', 'l_cst_model', 'l_cst_struct',
                                  'l_cst_tsct_elev', 'l_manningsn', 'l_meetings', 'l_mt2_lomr',
                                  'l_mtg_poc', 'l_pan_revis', 'l_pol_fhbm', 'l_profil_bkwtr_el',
                                  'l_profil_label', 'l_profil_panel', 'l_source_cit',
                                  'l_summary_discharges', 'l_summary_elevations', 'l_survey_pt',
                                  'l_xs_elev', 'l_xs_struct', 'study_info', 's_alluvial_fan',
                                  's_base_index', 's_bfe', 's_cbrs', 's_cst_gage', 's_cst_tsct_ln',
                                  's_datum_conv_pt', 's_firm_pan', 's_fld_haz_ar', 's_fld_haz_ln',
                                  's_gage', 's_gen_struct', 's_hwm', 's_hydro_reach', 's_label_ld',
                                  's_label_pt', 's_levee', 's_limwa', 's_lomr', 's_nodes',
                                  's_pfd_ln', 's_plss_ar', 's_pol_ar', 's_profil_basln',
                                  's_riv_mrk', 's_stn_start', 's_subbasins', 's_submittal_info',
                                  's_topo_confidence', 's_trnsport_ln', 's_tsct_basln', 's_wtr_ar',
                                  's_wtr_ln', 's_xs']

        # Set the workspace
        arcpy.env.workspace = self.workspace

        # Determine if shapefiles or databases are the workspace
        if not (in_workspace.lower().endswith('gdb') or in_workspace.lower().endswith('mdb')):
            self.dbf_ext = ".dbf"
            self.shp_ext = ".shp"
        else:
            self.dataset = "\\FIRM_Spatial_Layers"

        # Populate the DFIRM_ID, Source Citations, and V_DATUM
        self.dfirm_id = self.__get_dfirm_id()  # Holds the DFIRM_ID value from S_Submittal_Info
        self.source_citations = self.__get_source_citations()  # source_cits from L_Source_Cit.
        self.v_datum = self.__get_v_datum()  # Vertical datum used in the study from Study_Info

        # Set the tables.  It comes a # from ArcGIS if it's empty.
        if in_tables in ["#", "", " ", False, "False"]:
            self.tables = ''
            self.__printer("No tables selected.  Exiting...", True)
        elif ";" in in_tables:
            # ArcGIS doesn't provide the list as a true list.  It's actually a string with values
            # separated by semicolons.  This part checks for that and if it finds the semicolons
            # it converts the value to a true list
            tables_names = in_tables.replace("'", "").lower()
            self.tables = tables_names.split(";")
        elif str(type(in_tables)) == "<class 'list'>" or \
                str(type(in_tables)) == "<type 'list'>":
            self.tables = in_tables
        else:  # Used for only one table entered
            self.tables.append(in_tables.lower())

        # Check if the table is a standard FIRM table.
        for table in self.tables:
            if table.lower() not in self.acceptable_tables:
                arcpy.AddWarning(table.title() + " was selected but it is not a standard table " +
                                 "(check the spelling).  This table will be skipped.")
                self.tables.remove(table.lower())

        # Set the coded_check value.  It comes has as string from ArcGIS not a boolean
        if in_coded_check in ['true', 'True', True]:
            self.coded_check = True
        else:
            self.coded_check = False

        # Set the shp_export value.  It comes has as string from ArcGIS not a boolean
        if in_shp_export in ['true', 'True', True]:
            self.shp_export = True
        else:
            self.shp_export = False

        # Set the excel_export value.  It comes has as string from ArcGIS not a boolean
        if in_excel_export in ['true', 'True', True]:
            self.excel_export = True
        else:
            self.excel_export = False

    def __applicable_null_checks(self, in_table, in_field, in_field_type, id_field):
        """Checks for appropriate null values for applicable fields"""
        # List to hold the unique id number and the error
        error_list = []

        # Check for missing fields
        if in_field not in list(field.name for field in arcpy.ListFields(in_table)):
            self.__printer(in_field + ' missing in ' + in_table, True)
            return error_list

        # Iterate through the rows
        with SearchCursor(in_table, [id_field, in_field]) as cursor:
            for row in cursor:
                error_found = False  # Flag if error is found

                # Check for an empty field
                if in_field_type in ['Text', 'String']:
                    if str(type(row[1])) == "<type 'str'>" and row[1] is not None:
                        if row[1].isspace() and len(row[1]) > 1:
                            error_found = True

                # Check for correct NULL value for Numeric field types
                if in_field_type in ['Double', 'Integer', 'SmallInteger']:
                    if row[1] == -8888:  # Should be -9999 not -8888
                        error_found = True
                    elif row[1] is None:
                        error_found = True

                # Check for correct NULL value for Date field types
                elif in_field_type == 'Date':
                    if '8888' in str(row[1]):
                        error_list.append(
                            [str(row[0]), in_field + " value of 8/8/8888" +
                             " is not an acceptable NULL value for applicable fields"])
                    elif not row[1]:  # Date field is empty
                        error_list.append([str(row[0]), in_field + " should not be NULL"])

                    elif str(row[1]).strip() == '':  # Date field is empty
                        error_list.append([str(row[0]), in_field + " should not be empty"])

                # Append the error to the list if an error is found
                if error_found:
                    error_list.append([str(row[0]), in_field + " value of \"" + str(row[1]) +
                                       "\" is not an acceptable NULL value for applicable fields"])

        # Return the errors
        return error_list

    def __table_picker(self):
        """Determines which tables to process based on the MIP task"""
        # Alluvial Fan Data Capture
        if self.mip_task == 'Alluvial Fan Data Capture':
            self.task_tables = ['s_alluvial_fan', 's_gen_struct', 's_profil_basln', 's_stn_start',
                                's_submittal_info', 's_xs', 'l_source_cit', 'l_xs_elev',
                                'l_xs_struct']

        # Base Map Data Capture
        elif self.mip_task == 'Base Map Data Capture':
            self.task_tables = ['s_base_index', 's_datum_conv_pt', 's_firm_pan', 's_gen_struct',
                                's_plss_ar', 's_pol_ar', 's_submittal_info', 's_trnsport_ln',
                                's_wtr_ar', 's_wtr_ln', 'l_source_cit']

        # Coastal Data Capture
        elif self.mip_task == 'Coastal Data Capture':
            self.task_tables = ['s_cst_gage', 's_cst_tsct_ln', 's_fld_haz_ar', 's_gen_struct',
                                's_hwm', 's_levee', 's_limwa', 's_pfd_ln', 's_submittal_info',
                                's_tsct_basln', 'l_cst_model', 'l_cst_struct', 'l_cst_tsct_elev',
                                'l_source_cit', 'l_summary_elevations']

        # Develop Final Mapping Products Data Capture, Draft FIRM Database Data Capture,
        # Produce Preliminary Products Data Capture
        elif self.mip_task in ['Develop Final Mapping Products Data Capture',
                               'Draft FIRM Database Data Capture',
                               'Produce Preliminary Products Data Capture']:
            self.task_tables = self.tables

        # Existing Topographic Data Capture, New Topographic Data Capture, Terrain Data Capture
        elif self.mip_task in ['Existing Topographic Data Capture', 'New Topographic Data Capture',
                               'Terrain Data Capture']:
            self.task_tables = ['s_submittal_info', 's_topo_confidence', 'l_source_cit']

        # Floodplain Mapping Data Capture
        elif self.mip_task == 'Floodplain Mapping Data Capture':
            self.task_tables = ['s_alluvial_fan', 's_bfe', 's_cst_gage', 's_cst_tsct_ln',
                                's_fld_haz_ar', 's_gen_struct', 's_levee', 's_limwa', 's_pfd_ln',
                                's_profil_basln', 's_riv_mrk', 's_stn_start', 's_subbasins',
                                's_submittal_info', 's_tsct_basln', 's_xs', 'l_cst_model',
                                'l_cst_struct', 'l_cst_tsct_elev', 'l_profil_bkwtr_el',
                                'l_profil_label', 'l_profil_panel', 'l_source_cit',
                                'l_summary_discharges', 'l_xs_elev', 'l_xs_struct']
            if 's_topo_confidence' in self.task_tables:
                self.task_tables.remove('s_topo_confidence')
            if 'l_survey_pt' in self.task_tables:
                self.task_tables.remove('l_survey_pt')

        # Hydraulics Data Capture
        elif self.mip_task == 'Hydraulics Data Capture':
            self.task_tables = ['s_bfe', 's_fld_haz_ar', 's_gen_struct', 's_hwm', 's_levee',
                                's_nodes', 's_profil_basln', 's_riv_mrk', 's_stn_start',
                                's_submittal_info', 's_xs', 'l_manningsn', 'l_profil_bkwtr_el',
                                'l_profil_label', 'l_profil_panel', 'l_source_cit',
                                'l_summary_elevations', 'l_xs_elev', 'l_xs_struct']

        # Hydrology Data Capture
        elif self.mip_task == 'Hydrology Data Capture':
            self.task_tables = ['s_gage', 's_hydro_reach', 's_nodes', 's_subbasins',
                                's_submittal_info', 'l_source_cit', 'l_summary_discharges',
                                'l_summary_elevations']

        # Survey Data Capture
        elif self.mip_task == 'Survey Data Capture':
            self.task_tables = ['s_submittal_info', 'l_source_cit', 'l_survey_pt']

    def __dfirm_id_check(self, in_table, id_field):
        """Checks the DFIRM_ID against the DFIRM_ID in S_Submittal_Info"""
        # List to hold the unique id number and the error
        error_list = []

        # Iterate through the rows
        with SearchCursor(in_table, [id_field, 'DFIRM_ID']) as cursor:
            for row in cursor:
                if row[1] not in self.dfirm_id:
                    error_list.append([str(row[0]), "DFIRM_ID value of \"" + str(row[1]) +
                                       "\" does not match the DFIRM_ID value in S_Submittal_Info"])

        # Return the errors
        return error_list

    def __domain_checks(self, in_table, field, domain_values, id_field, required_fields):
        """Checks the domain values of the tables"""
        # List to hold the unique id number and the error
        error_list = []

        # Iterate through the rows
        with SearchCursor(in_table, [id_field, field]) as cursor:
            for row in cursor:
                # If checking coded values, get the keys from the dictionary
                if self.coded_check:
                    check_values = domain_values.keys()
                # Else, if checking the text values, get the values from the dictionary
                else:
                    check_values = domain_values.values()
                if row[1] not in check_values:
                    # If it's a required field, it has to have a domain value.
                    # It can't be empty or Null
                    if field in required_fields:
                        error_list.append(
                            [str(row[0]),
                             field + " value of \"" + str(row[1]) + "\" is not in domain"])

                    # Else if it's an applicable field, it can be empty or Null
                    elif not (str(row[1]).strip() != '' or row[1] != "None"):
                        error_list.append(
                            [str(row[0]),
                             field + " value of \"" + str(row[1]) + "\" is not in domain"])

        # Return the errors
        return error_list

    def __get_dfirm_id(self):
        """Gets the DFIRM_ID from the S_Submittal_Info table"""
        submittal_info = self.workspace + self.dataset + '\\S_Submittal_Info' + self.shp_ext
        if arcpy.Exists(submittal_info):
            if 'DFIRM_ID' in [field.name for field in arcpy.ListFields(submittal_info)]:
                dfirm_id_value = list(set([row[0] for row in SearchCursor(submittal_info,
                                                                          "DFIRM_ID")]))

                return dfirm_id_value
            else:
                arcpy.AddError("DFIRM_ID missing in S_Submittal_Info.  Unable to proceed.")
                sys.exit(1)
        else:
            return ""

    @staticmethod
    def __get_field_dict(in_table, field_list):
        """Creates two dictionaries of required fields and applicable fields and their field types.
        The field_list parameter is a list of fields that are required for that table"""
        required_fields = {}
        applicable_fields = {}

        for field in arcpy.ListFields(in_table):
            # Skip these
            if field.name.lower() not in ['objectid', 'fid', 'shape', 'shape_length', 'shape_area']:
                if field.name in field_list:  # Required fields
                    required_fields[field.name] = field.type
                else:  # Applicable fields
                    applicable_fields[field.name] = field.type

        return required_fields, applicable_fields

    def __field_checker(self, in_table, field_list):
        """Checks if all the fields to be validated are actually in the table"""
        # List of missing fields
        missing_fields = []

        # List to hold the field names of table
        table_fields = []

        # Get a list of the field names from the table
        for field in arcpy.ListFields(in_table):
            table_fields.append(field.name)

        # Check if the fields in field_list are in the table_fields list
        for field in field_list:
            if field not in table_fields:
                missing_fields.append(field)
                self.missing_field = True  # Flip the flag

        # Display an error message if fields are missing
        if missing_fields:
            arcpy.AddWarning("Unable to validate " + os.path.basename(in_table) +
                             " - the following fields are missing: " +
                             ", ".join(missing_fields))

        # Return the list of missing fields
        return missing_fields

    def __get_source_citations(self):
        """Creates the list of source citations from L_Source_Cit"""
        # Check for the tables existence.  If it doesn't exist, return
        if arcpy.Exists(self.workspace + '\\L_Source_Cit' + self.dbf_ext):
            source_citations = list(set([row[0] for row in SearchCursor(
                self.workspace + '\\L_Source_Cit' + self.dbf_ext, "SOURCE_CIT")]))
        else:
            source_citations = []

        # Return the list
        return source_citations

    def __get_v_datum(self):
        """Gets the Vertical Datum from the Study_Info table"""
        v_datum = ''

        if arcpy.Exists(self.workspace + '\\Study_info' + self.dbf_ext):
            result = arcpy.GetCount_management(self.workspace + '\\Study_info' + self.dbf_ext)
            if int(result[0]) != 0:
                v_datum_list = list(set([row[0] for row in SearchCursor(
                    self.workspace + '\\Study_info' + self.dbf_ext, "V_DATUM")]))
                v_datum = v_datum_list[0]
        return v_datum

    @staticmethod
    def __id_table_check(primary_table, primary_field, foreign_table,
                         foreign_field, query="", error_message=""):
        """Checks if the all the values from the id field in the primary table have a
           matching value in the foreign table"""

        if error_message == '':
            error_message = "Matching ID value not found in the " + foreign_field + \
                            " field in " + os.path.basename(foreign_table) + " "

        error_list = []  # List of errors found

        # Get a list of unique values from the primary table
        primary_id_list = sorted(list(set(
            [str(row[0]).strip() for row in SearchCursor(primary_table, primary_field, query)])))

        # Set the primary_id_list to empty if only None is returned
        if len(primary_id_list) == 1 and primary_id_list[0] in ['None', '']:
            primary_id_list = []

        # Get a list of unique values from the foreign table
        if arcpy.Exists(foreign_table):
            foreign_id_list = sorted(list(set(
                [str(row[0]) for row in SearchCursor(foreign_table, foreign_field)])))

            # Check for items in primary table are in foreign table
            for primary_id in primary_id_list:
                if primary_id not in foreign_id_list and primary_id.strip() != '':
                    error_list.append([primary_id, error_message])

        else:
            for primary_id in primary_id_list:
                error_list.append([primary_id,
                                   os.path.basename(foreign_table) + " table couldn't be found"])

        # Return the error list
        return error_list

    @staticmethod
    def __printer(in_message, warning=False):
        """Prints the message to both the terminal and ArcToolbox"""
        print(in_message)

        if warning:
            arcpy.AddWarning(in_message)
        else:
            arcpy.AddMessage(in_message)

    @staticmethod
    def __required_null_checks(in_table, in_field, in_field_type, id_field):
        """Checks for appropriate null values for required fields"""
        # List to hold the unique id number and the error
        error_list = []

        # Iterate through the rows
        with SearchCursor(in_table, [id_field, in_field]) as cursor:
            for row in cursor:
                error_found = False  # Flag if error is found

                # Check for an empty field
                if in_field_type in ['Text', 'String']:
                    if row[1] is None:
                        error_found = True
                    elif str(type(row[1])) in ["<type 'str'>", "<type 'unicode'>", "<class 'str'>"]:
                        if row[1].isspace() or len(row[1]) == 0:
                            error_found = True

                elif in_field_type in ['Double', 'Integer', 'SmallInteger']:
                    if row[1]:
                        if row[1] == -9999:  # Should be -8888 not -9999
                            error_found = True
                    elif row[1] is None:
                        error_found = True

                elif in_field_type == 'Date':
                    if '9999' in str(row[1]):  # Should be 8/8/8888 not 9/9/9999
                        error_list.append([str(row[0]), in_field + " value of 9/9/9999" +
                                           " is not an acceptable NULL value for required fields"])
                    elif not row[1]:  # Date field is empty
                        error_list.append([str(row[0]), in_field + " should not be NULL"])

                    elif str(row[1]).strip() == '':  # Date field is empty
                        error_list.append([str(row[0]), in_field + " should not be empty"])

                if error_found:
                    error_list.append([str(row[0]), in_field + " value of \"" + str(row[1]) +
                                       "\" is not an acceptable NULL value for required fields"])

        # Return the errors
        return error_list

    def __source_check(self, in_table, id_field):
        """Checks for matching source citation between the input table and L_Source_Cit"""
        # List to hold the unique id number and the error
        error_list = []

        # Check if the 'SOURCE_CIT' field exists
        for field in arcpy.ListFields(in_table):
            if field.name == 'SOURCE_CIT':
                # Iterate through the rows
                with SearchCursor(in_table, [id_field, 'Source_Cit']) as cursor:
                    for row in cursor:
                        if str(row[1]) not in self.source_citations:
                            error_list.append([str(row[0]), "SOURCE_CIT value of \"" + str(row[1]) +
                                               "\" does not match any values in L_Source_Cit"])

        # Return the errors
        return error_list

    @staticmethod
    def __space_check(in_table, id_field):
        """Checks for extra spaces in each text field of the table"""
        # List to hold the unique id number and the error
        error_list = []

        # Get a list of text fields in the table
        field_names = list(field.name for field in arcpy.ListFields(in_table, "*", "String"))

        # Remove the id_field and put it at the front of the list
        field_names.remove(id_field)
        field_names.insert(0, id_field)

        # Iterate through the rows
        with SearchCursor(in_table, field_names) as cursor:
            for row in cursor:
                # Column counter
                col = 1  # Start at 1 to skip the id_field at the beginning of the list

                # Iterate through the columns
                while col < len(field_names):
                    if row[col]:
                        if len(row[col]) != len(row[col].strip()) and len(row[col]) > 1:
                            error_list.append(
                                [str(row[0]), field_names[col] + " has an extra space."])
                    col += 1

        # Return the errors
        return error_list

    def __standard_table_checks(self, in_table, id_field, field_domains,
                                required_fields, applicable_fields, check_fields):
        """Standard checks performed on all tables"""
        errors = []  # Errors found

        # Check if the field exists
        missing_fields = self.__field_checker(in_table, check_fields)

        # If there are missing fields, generate a warning and then return
        if missing_fields:
            errors.append(["", "Missing fields: " + ", ".join(missing_fields)])
            return errors

        # Check the DFIRM_ID values
        for error in self.__dfirm_id_check(in_table, id_field):
            errors.append(error)

        # Check the Source Citation values
        for error in self.__source_check(in_table, id_field):
            errors.append(error)

        # Check for unique id
        for error in self.__unique_id_check(in_table, id_field):
            errors.append(error)

        # Iterate through the field list and perform domain value checks
        for field in arcpy.ListFields(in_table):
            if field.name in field_domains.keys():
                for error in self.__domain_checks(in_table, field.name, field_domains[field.name],
                                                  id_field, required_fields):
                    errors.append(error)

        # Iterate through the required fields and perform null checks
        for key in required_fields.keys():
            for error in self.__required_null_checks(in_table, key, required_fields[key], id_field):
                errors.append(error)

        # Iterate through the applicable fields and perform null checks
        for key in applicable_fields.keys():
            for error in self.__applicable_null_checks(in_table, key,
                                                       applicable_fields[key], id_field):
                errors.append(error)

        # Check for extra spaces
        for error in self.__space_check(in_table, id_field):
            errors.append(error)

        # Return the errors found
        return errors

    @staticmethod
    def __unique_id_check(in_table, in_field, error_message="Duplicate unique id found in "):
        """Checks for unique id values"""
        # List to hold the unique id number and the error
        error_list = []

        values = [row[0] for row in SearchCursor(in_table, in_field)]
        dupes = list(set([x for n, x in enumerate(values) if x in values[:n]]))

        # Check for duplicate unique ids
        for dupe in dupes:
            error = [dupe, error_message + in_field]
            error_list.append(error)

        # Return the errors
        return error_list

    def iterate_tables(self):
        """Iterates through the tables"""
        tables_found = []  # List of tables found in the workspace

        # Determine which tables apply to the MIP Task
        self.__table_picker()

        # Find all the feature class tables with any feature datasets
        datasets = arcpy.ListDatasets("*", "Feature")
        for dataset in datasets:
            feature_classes = arcpy.ListFeatureClasses("*", "ALL", dataset)
            for feature_class in feature_classes:
                if feature_class.lower() in self.tables:
                    feature_class_path = os.path.join(self.workspace, dataset, feature_class)
                    if feature_class_path not in tables_found:
                        tables_found.append(feature_class_path)

        # Find all the stand-alone feature classes
        feature_classes = arcpy.ListFeatureClasses()
        for feature_class in feature_classes:
            if feature_class.lower().replace(".shp", "") in self.tables:
                feature_class_path = os.path.join(self.workspace, feature_class)
                if feature_class_path not in tables_found:
                    tables_found.append(feature_class_path)

        # Find all the stand-alone tables
        table_list = arcpy.ListTables()
        for table in table_list:
            if table.lower().replace(".dbf", "") in self.tables:
                table_path = os.path.join(self.workspace, table)
                if table_path not in tables_found:
                    tables_found.append(table_path)

        # Iterate through the tables found
        for table_path in sorted(tables_found):
            # Check if the table is empty
            if int(arcpy.GetCount_management(table_path)[0]) != 0:
                table_name = os.path.basename(table_path)
                self.__printer('Checking ' + table_name)
                self.errors = []
                # Remove the .shp or .dbf extensions if shapefiles are found
                table_name = table_name.replace(".shp", "")
                table_name = table_name.replace(".dbf", "")
                # Execute the function associated with the table
                # Example:
                # exec(self.s_base_index_check(r'C:\temp\test.gdb\FIRM_Spatial_Layers\S_Base_Index')
                if table_name.lower() not in self.task_tables:
                    arcpy.AddWarning(
                        table_name +
                        ' contains data but is not applicable for the choose MIP task.')
                exec("self." + table_name.lower() + "_check(r'" + table_path + "')")

                # Write out the errors to a DBF file
                if self.shp_export:
                    self.write_out_errors_dbf(sorted(self.errors), table_name)

                # Write out the errors to an Excel file
                if self.excel_export:
                    self.write_out_errors_exel(sorted(self.errors), table_name)

        # Show total errors found
        self.__printer("\nTotal errors: {}".format(self.total_errors))

    def s_alluvial_fan_check(self, in_feature_class):
        """QC check of S_Alluvial_Fan"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'ALLUVL_ID', 'ACTIVE_FAN', 'FANAPEX_DA',
                  'AREA_UNITS', 'FANAPEX_Q', 'DISCH_UNIT', 'FLD_ZONE', 'ZONE_SUBTY', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'FANAPEX_DA', 'AREA_UNITS', 'FANAPEX_Q', 'DISCH_UNIT',
                     'FAN_VEL_MN', 'FAN_VEL_MX', 'VEL_UNIT', 'DEPTH', 'DEPTH_UNIT')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Alluvial_Fan_domains, required_fields,
                                                  applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                fanapex_da = row[1]
                area_units = row[2]
                fanapex_q = row[3]
                disch_unit = row[4]
                fan_vel_mn = row[5]
                fan_vel_mx = row[6]
                vel_unit = row[7]
                depth = row[8]
                depth_unit = row[9]

                # AREA_UNITS should be populated when FANAPEX_DA is populated
                if fanapex_da != -8888 and area_units == 'NP':
                    self.errors.append(
                        [unique_id, "AREA_UNITS should be populated when FANAPEX_DA is populated"])

                # DISCH_UNIT should be populated when FANAPEX_Q is populated
                if fanapex_q != -8888 and disch_unit == 'NP':
                    self.errors.append(
                        [unique_id, "DISCH_UNIT should be populated when FANAPEX_Q is populated"])

                # DISCH_UNIT should not be populated when FANAPEX_Q is not populated
                if fanapex_q == -8888 and disch_unit != 'NP':
                    self.errors.append(
                        [unique_id,
                         "DISCH_UNIT should not be populated when FANAPEX_Q is not populated"])

                # VEL_UNIT should be populated when FAN_VEL_MN is populated
                if fan_vel_mn != -9999 and vel_unit is None:
                    self.errors.append(
                        [unique_id, "VEL_UNIT should be populated when FAN_VEL_MN is populated"])
                elif fan_vel_mn != -9999 and (vel_unit == 'NP' or vel_unit.strip() == ''):
                    self.errors.append(
                        [unique_id, "VEL_UNIT should be populated when FAN_VEL_MN is populated"])

                # VEL_UNIT should not be populated when FAN_VEL_MN is not populated
                if fan_vel_mn == -9999 and vel_unit is not None:
                    if vel_unit.strip() == '' or vel_unit != 'NP':
                        self.errors.append(
                            [unique_id,
                             "VEL_UNIT should not be populated when FAN_VEL_MN is not populated"])

                # VEL_UNIT should be populated when FAN_VEL_MX is populated
                if fan_vel_mx != -9999 and vel_unit is None:
                    self.errors.append(
                        [unique_id, "VEL_UNIT should be populated when FAN_VEL_MX is populated"])
                elif fan_vel_mx != -9999 and (vel_unit == 'NP' or vel_unit.strip() == ''):
                    self.errors.append(
                        [unique_id, "VEL_UNIT should be populated when FAN_VEL_MX is populated"])

                # VEL_UNIT should not be populated when FAN_VEL_MN is not populated
                if fan_vel_mx == -9999 and vel_unit is not None:
                    if vel_unit.strip() == '' or vel_unit != 'NP':
                        self.errors.append(
                            [unique_id,
                             "VEL_UNIT should not be populated when FAN_VEL_MN is not populated"])

                # DEPTH_UNIT should be populated when DEPTH is populated
                if depth != -9999 and depth_unit is None:
                    self.errors.append(
                        [unique_id, "DEPTH_UNIT should be populated when DEPTH is populated"])
                elif depth != -9999 and (depth_unit == 'NP' or depth_unit.strip() == ''):
                    self.errors.append(
                        [unique_id, "DEPTH_UNIT should be populated when DEPTH is populated"])

                # DEPTH_UNIT should not be populated when DEPTH is not populated
                if depth == -9999 and depth_unit is not None:
                    if depth_unit.strip() != '' or depth_unit != 'NP':
                        self.errors.append(
                            [unique_id,
                             "DEPTH_UNIT should not be populated when DEPTH is not populated"])

    def s_base_index_check(self, in_feature_class):
        """QC check of S_Base_Index"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'BASE_ID', 'FILENAME', 'BASE_DATE', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'FILENAME')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Base_Index_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                filename = row[1]

                if filename:
                    if filename[-4:] not in [".bil", ".bip", ".ecw", ".tif", ".img", ".jp2",
                                             ".jpg", ".sid", ".png"]:
                        self.errors.append([unique_id,
                                            'Proper extension is missing in the filename'])

        # Check for duplicate FILENAME values
        errors = self.__unique_id_check(in_feature_class, 'FILENAME',
                                        error_message="Duplicate names found in ")
        for error in errors:
            self.errors.append(error)

    def s_bfe_check(self, in_feature_class):
        """QC check of S_BFE"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'BFE_LN_ID', 'ELEV',
                  'LEN_UNIT', 'V_DATUM', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'LEN_UNIT', 'V_DATUM', 'ELEV')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_BFE_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                len_unit = row[1]
                v_datum = row[2]
                elev = row[3]

                # LEN_UNIT should be populated when ELEV is populated
                if len_unit == 'NP' and elev != '-8888':
                    self.errors.append([unique_id,
                                        "LEN_UNIT should be populated when ELEV is populated"])

                # V_DATUM does not match the V_DATUM value in Study_Info
                if v_datum != self.v_datum and self.v_datum != '' and \
                        self.mip_task in ['Develop Final Mapping Products Data Capture',
                                          'Draft FIRM Database Data Capture',
                                          'Produce Preliminary Products Data Capture']:
                    self.errors.append([unique_id, "V_DATUM value of " + str(row[3]) +
                                        " does not match the V_DATUM value in Study_Info"])

    def s_cst_gage_check(self, in_feature_class):
        """QC check of S_Cst_Gage"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'CSTGAGE_ID', 'GAGE_NM', 'AGENCY', 'START_PD',
                  'END_PD', 'GAGE_TYPE', 'V_DATUM', 'TIDE_TF', 'WVDIR_TF', 'WVSPEC_TF', 'WDSPD_TF',
                  'WDDIR_TF', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'REC_INTVL', 'TIME_UNIT', 'START_PD', 'START_TIME', 'END_PD',
                     'END_TIME', 'V_DATUM', 'TIDE_TF', 'TIDE_EPOCH', 'WDSPD_TF', 'WDDIR_TF',
                     'WDSTN_HT')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Cst_Gage_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in L_Cst_Model.CST_MDL_ID
        for error in self.__id_table_check(in_feature_class, 'CST_MDL_ID',
                                           self.workspace + '\\L_Cst_Model' + self.dbf_ext,
                                           'CST_MDL_ID'):
            self.errors.append(error)

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                rec_intvl = row[1]
                time_unit = row[2]
                start_pd = row[3]
                start_time = row[4]
                end_pd = row[5]
                end_time = row[6]
                v_datum = row[7]
                tide_tf = row[8]
                tide_epoch = row[9]
                wdspd_tf = row[10]
                wddir_tf = row[11]
                wdstn_ht = row[12]

                # TIME_UNIT should be populated when REC_INTVL is populated
                if str(time_unit).strip().title() in ['None', ''] and rec_intvl is not None:
                    if rec_intvl.strip() != '':
                        self.errors.append(
                            [unique_id,
                             "TIME_UNIT should be populated when REC_INTVL is populated"])

                # TIME_UNIT should be not populated when REC_INTVL is not populated
                if str(time_unit).strip().title() not in ['None', ''] and rec_intvl is None:
                    if time_unit.strip() != '':
                        self.errors.append(
                            [unique_id,
                             "TIME_UNIT should be not populated when REC_INTVL is not populated"])

                # START_TIME should be populated when START_PD is populated
                if start_time is None and '8888' not in str(start_pd):
                    self.errors.append(
                        [unique_id,
                         "START_TIME should be populated when START_PD is populated"])
                elif str(time_unit).strip().title() in ['None', ''] and '8888' not in str(start_pd):
                    if str(rec_intvl).strip() != '':
                        self.errors.append(
                            [unique_id,
                             "START_TIME should be populated when START_PD is populated"])

                # START_TIME should not be populated when START_PD is not populated
                if start_time is not None and '8888' in str(start_pd):
                    if start_time.strip() != '':
                        self.errors.append(
                            [unique_id,
                             "START_TIME should not be populated when START_PD is not populated"])

                # END_TIME should be populated when END_PD is populated
                if end_time is None and '8888' not in str(end_pd):
                    self.errors.append(
                        [unique_id, "END_TIME should be populated when END_PD is populated"])
                elif end_time.strip() == '' and '8888' not in str(end_pd):
                    if rec_intvl.strip() != '':
                        self.errors.append(
                            [unique_id, "END_TIME should be populated when END_PD is populated"])

                # END_TIME should not be populated when END_PD is not populated
                if end_time is not None and '8888' in str(end_pd):
                    if end_time.strip() != '':
                        self.errors.append(
                            [unique_id,
                             "END_TIME should not be populated when END_PD is not populated"])

                # V_DATUM does not match the V_DATUM value in Study_Info
                if v_datum != self.v_datum and self.v_datum != '' and \
                        self.mip_task in ['Develop Final Mapping Products Data Capture',
                                          'Draft FIRM Database Data Capture',
                                          'Produce Preliminary Products Data Capture']:
                    self.errors.append(
                        [unique_id, "WARNING: V_DATUM value of " + str(row[7]) +
                         " does not match the V_DATUM value in Study_Info.  Verify the value."])

                # TIDE_EPOCH should be populated when TIDE_TF is True
                if tide_epoch is None and tide_tf == 'T':
                    if str(tide_epoch).strip() != '':
                        self.errors.append([unique_id,
                                            "TIDE_EPOCH should be populated when TIDE_TF is True"])

                # WDSTN_HT should be populated when WDSPD_TF or WDDIR_TF is True
                if wdstn_ht is None and (wdspd_tf == 'T' or wddir_tf == 'T'):
                    if wdstn_ht.strip() == '':
                        self.errors.append(
                            [unique_id,
                             "WDSTN_HT should be populated when WDSPD_TF or WDDIR_TF is True"])

    def s_cst_tsct_ln_check(self, in_feature_class):
        """QC check of S_Cst_Tsct_Ln"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'TRAN_LN_ID', 'TBASELN_ID', 'TRAN_NO', 'XCOORD',
                  'YCOORD', 'WTR_NM', 'V_DATUM', 'CSTLN_TYP', 'EVENT_TYP', 'SWEL', 'LOC_DESC',
                  'RUP', 'ELEV_UNIT', 'WHAFIS_TF', 'OVERTOP_TF', 'BW_HGT_TF', 'HVFLOW_TF',
                  'WAVE_02PCT', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'V_DATUM', 'SWEL', 'RUP', 'ELEV_UNIT', 'SETUP_DPTH', 'CON_HT',
                     'SIG_HT', 'MEAN_HT', 'SIG_PD', 'CON_PD', 'MEAN_PD', 'LEN_UNIT', 'TIME_UNIT')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Cst_Tsct_Ln_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in L_Cst_Tsct_Elev.TRAN_LN_ID
        for error in self.__id_table_check(
                in_feature_class, 'TRAN_LN_ID',
                self.workspace + '\\L_Cst_Tsct_Elev' + self.dbf_ext, 'TRAN_LN_ID'):
            self.errors.append(error)

        # Check for matching ID value in S_Tsct_Basln.TBASELN_ID
        for error in self.__id_table_check(
                in_feature_class, 'TBASELN_ID',
                self.workspace + self.dataset + '\\S_Tsct_Basln' + self.shp_ext, 'TBASELN_ID'):
            self.errors.append(error)

        # Check for matching ID value in L_Cst_Model.CST_MDL_ID
        for error in self.__id_table_check(
                in_feature_class, 'CST_MDL_ID', self.workspace + '\\L_Cst_Model' + self.dbf_ext,
                'CST_MDL_ID'):
            self.errors.append(error)

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                v_datum = row[1]
                swel = row[2]
                rup = row[3]
                elev_unit = row[4]
                setup_depth = row[5]
                con_ht = row[6]
                sig_ht = row[7]
                mean_ht = row[8]
                sig_pd = row[9]
                con_pd = row[10]
                mean_pd = row[11]
                len_unit = row[12]
                time_unit = row[13]

                # V_DATUM does not match the V_DATUM value in Study_Info
                if v_datum != self.v_datum and self.v_datum != '' and \
                        self.mip_task in ['Develop Final Mapping Products Data Capture',
                                          'Draft FIRM Database Data Capture',
                                          'Produce Preliminary Products Data Capture']:
                    self.errors.append([unique_id, "V_DATUM value of " + str(row[3]) +
                                        " does not match the V_DATUM value in Study_Info"])

                # ELEV_UNIT should be populated when SWEL or RUP is populated
                if elev_unit == 'NP' and (swel != -8888 or rup != -8888):
                    self.errors.append(
                        [unique_id, "ELEV_UNIT should be populated when SWEL or RUP is populated"])

                # LEN_UNIT should be populated when SETUP_DEPTH, CON_HT,
                # SIG_HT or MEAN_HT  is populated
                if len_unit is None and (
                        setup_depth != -9999 or con_ht != -9999 or sig_ht != -9999 or
                        mean_ht != -9999):
                    self.errors.append(
                        [unique_id,
                         "LEN_UNIT should be populated when SETUP_DEPTH, CON_HT, SIG_HT or MEAN_HT "
                         " is populated"])
                elif (str(len_unit).strip == '' or len_unit == 'NP') and \
                        (setup_depth != -9999 or con_ht != -9999 or sig_ht != -9999 or
                         mean_ht != -9999):
                    self.errors.append(
                        [unique_id,
                         "LEN_UNIT should be populated when SETUP_DEPTH, CON_HT, SIG_HT or MEAN_HT "
                         " is populated"])

                # TIME_UNIT should be populated when SIG_PD, CON_PD, or MEAN_PD  is populated
                if time_unit is None and (sig_pd != -9999 or con_pd != -9999 or mean_pd != -9999):
                    self.errors.append(
                        [unique_id,
                         "TIME_UNIT should be populated when SIG_PD, CON_PD, or MEAN_PD  is "
                         "populated"])
                elif (str(time_unit).strip == '' or time_unit == 'NP') and \
                        (sig_pd != -9999 or con_pd != -9999 or mean_pd != -9999):
                    self.errors.append(
                        [unique_id,
                         "TIME_UNIT should be populated when SIG_PD, CON_PD, or MEAN_PD  is "
                         "populated"])

    def s_datum_conv_pt_check(self, in_feature_class):
        """QC check of S_Datum_Conv_Pt"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'DATCONPTID', 'FROM_DATUM', 'TO_DATUM', 'CONVFACTOR',
                  'LEN_UNIT', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'TO_DATUM', 'CONVFACTOR', 'LEN_UNIT')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Datum_Conv_Pt_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Specific checks for this feature class
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                to_datum = row[1]
                convfactor = row[2]
                len_unit = row[3]

                # TO_DATUM does not match value in Study_Info
                if to_datum != self.v_datum and self.v_datum != '' and \
                        self.mip_task in ['Develop Final Mapping Products Data Capture',
                                          'Draft FIRM Database Data Capture',
                                          'Produce Preliminary Products Data Capture']:
                    self.errors.append([unique_id, "TO_DATUM value of " + str(row[3]) +
                                        " does not match the V_DATUM value in Study_Info"])

                    # LEN_UNIT should be populated when CONVFACTOR is populated
                if len_unit == 'NP' and convfactor != -8888:
                    self.errors.append(
                        [unique_id, "LEN_UNIT should be populated when CONVFACTOR is populated"])

    def s_firm_pan_check(self, in_feature_class):
        """QC check of S_FIRM_Pan"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'FIRM_ID', 'ST_FIPS', 'PCOMM', 'PANEL', 'SUFFIX',
                  'FIRM_PAN', 'PANEL_TYP', 'SCALE', 'BASE_TYP', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'ST_FIPS', 'PCOMM', 'PANEL', 'SUFFIX', 'FIRM_PAN', 'PANEL_TYP',
                     'PNP_REASON', 'SCALE')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_FIRM_Pan_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Specific checks for this feature class
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                st_fips = row[1]
                pcomm = row[2]
                panel = row[3]
                suffix = row[4]
                firm_pan = row[5]
                panel_typ = row[6]
                pnp_reason = row[7]
                scale = row[8]

                # PCOMM ends with C for Community Panel
                if str(pcomm)[-1:].lower() == 'c' and panel_typ in \
                        ['1020', 'Community Based, Panel Printed',
                         '1030', 'Community Based, Not Printed']:
                    self.errors.append([unique_id, 'Community based panels should not end in "C"'])

                # PCOMM missing C for Countywide Panel
                if str(pcomm)[-1:].lower() != 'c' and panel_typ in \
                        ['1000', 'Countywide, Panel Printed', '1010', 'Countywide, Not Printed']:
                    self.errors.append([unique_id, 'Countywide based panels should end in "C"'])

                # PCOMM not four characters long
                if len(str(pcomm)) != 4:
                    self.errors.append([unique_id, 'PCOMM should be 4 characters long'])

                # PCOMM not 3rd - 6th characters of FIRM_PAN
                if str(pcomm) != str(firm_pan)[2:6]:
                    self.errors.append(
                        [unique_id,
                         'PCOMM value does not equal the 3rd-6th characters in FIRM_PAN'])

                # PANEL value is 0000
                if panel == '0000':
                    self.errors.append([unique_id, 'PANEL value is 0000'])

                # PANEL not four characters long
                if len(str(panel)) != 4:
                    self.errors.append([unique_id, 'PANEL should be 4 characters long'])

                # PANEL not 7th - 10th characters of FIRM_PAN
                if str(panel) != str(firm_pan)[6:10]:
                    self.errors.append(
                        [unique_id,
                         'PANEL value does not equal the 7th-10th characters in FIRM_PAN'])

                # SCALE is not correct for PANEL Number
                if panel:
                    if panel.isdigit():
                        if scale in ['1020', '24000'] and int(panel) % 25 != 0 or \
                                scale in ['1010', '12000'] and int(panel) % 5 != 0 or \
                                scale in ['1000', '6000'] and int(panel) % 5 == 0:
                            self.errors.append([unique_id, 'SCALE is incorrect of PANEL number'])

                # SUFFIX not the 11th character in FIRM_PAN
                if str(suffix) != str(firm_pan)[-1:]:
                    self.errors.append([unique_id,
                                        'SUFFIX does not equal the 11th character in FIRM_PAN'])

                # SUFFIX is not a letter
                if not str(suffix).isalpha():
                    self.errors.append([unique_id, 'SUFFIX is not a letter'])

                # SUFFIX is not O or I
                if str(suffix).lower() in ['o', 'i']:
                    self.errors.append([unique_id, 'SUFFIX should not be "O" or "I"'])

                # SUFFIX should be between A and Z
                if suffix:
                    if ord(str(suffix).lower()) not in range(ord('a'), ord('z') + 1):
                        self.errors.append([unique_id, "SUFFIX  NOT BETWEEN 'A' AND 'Z'"])

                # FIRM_PAN value is not 11 characters long
                if len(str(firm_pan)) != 11:
                    self.errors.append([unique_id, 'FIRM_PAN should be 11 characters long'])

                # FIRM_PAN value does not match ST_FIPS + PCOMM + PANEL + SUFFIX
                if str(firm_pan) != str(st_fips) + str(pcomm) + str(panel) + str(suffix):
                    self.errors.append(
                        [unique_id,
                         'FIRM_PAN value does not match ST_FIPS + PCOMM + PANEL + SUFFIX'])

                # PANEL_TYP is Printed and PNP_REASON is not empty
                if str(panel_typ) in ['1000', 'Countywide, Panel Printed', '1020',
                                      'Community Based, Panel Printed',
                                      '1050', 'Statewide, Panel Printed'] and \
                        str(pnp_reason).strip() not in ['', 'None']:
                    self.errors.append([unique_id,
                                        'PANEL_TYP is printed and PNP_REASON is not empty'])

                # PANEL_TYP Not Printed AND No PNP_REASON is given
                if str(panel_typ) in ['1010', 'Countywide, Not Printed', '1030',
                                      'Community Based, Not Printed', '1040', 'Unmapped Community',
                                      '1060', 'Statewide, Not Printed'] and \
                        str(pnp_reason).strip() in ['', 'None']:
                    self.errors.append([unique_id,
                                        'PANEL_TYP is not printed and PNP_REASON is empty'])

    def s_fld_haz_ar_check(self, in_feature_class):
        """QC check of S_Fld_Haz_Ar"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'FLD_AR_ID', 'STUDY_TYP',
                  'FLD_ZONE', 'SFHA_TF', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'STATIC_BFE', 'FLD_ZONE', 'V_DATUM', 'DEPTH', 'LEN_UNIT',
                     'VELOCITY', 'VEL_UNIT')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Change the domains to correct spec
        if self.schema == '2017':
            domains = S_Fld_Haz_Ar_domains_2017
        elif self.schema == "2020":
            domains = S_Fld_Haz_Ar_domains_2020
        elif self.schema == "2021":
            domains = S_Fld_Haz_Ar_domains_2021
        else:
            domains = S_Fld_Haz_Ar_domains

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2], domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                static_bfe = row[1]
                fld_zone = row[2]
                v_datum = row[3]
                depth = row[4]
                len_unit = row[5]
                velocity = row[6]
                vel_unit = row[7]

                # STATIC_BFE Not Normally Calculated for this FLD_ZONE Check ('AE','AH','AO','VE')
                if static_bfe != -9999 and fld_zone not in ['AE', 'AH', 'AO', 'VE']:
                    self.errors.append(
                        [unique_id, 'STATIC_BFE Not Normally Calculated for this FLD_ZONE'])

                # V_DATUM checks
                if v_datum:
                    # V_DATUM does not match value in Study_Info
                    if v_datum != self.v_datum and v_datum.strip() != '' and \
                            self.v_datum != '' and \
                            self.mip_task in ['Develop Final Mapping Products Data Capture',
                                              'Draft FIRM Database Data Capture',
                                              'Produce Preliminary Products Data Capture']:
                        self.errors.append([unique_id, "V_DATUM value of " + str(row[3]) +
                                            " does not match the V_DATUM value in Study_Info"])

                    # V_DATUM is populated without a Static BFE value
                    if v_datum.strip() != '' and static_bfe == -9999:
                        self.errors.append(
                            [unique_id, "V_DATUM is populated without a STATIC_BFE value"])

                # V_DATUM not populated for Static BFE
                if v_datum is None:
                    if static_bfe != -9999:
                        self.errors.append([unique_id, "V_DATUM not populated for STATIC_BFE"])
                elif v_datum.strip() == '' and static_bfe != -9999:
                    self.errors.append([unique_id, "V_DATUM not populated for STATIC_BFE"])

                # DEPTH field is populated for the wrong zone ('AO')
                if depth:
                    if depth != -9999 and fld_zone != 'AO':
                        self.errors.append(
                            [unique_id, "DEPTH field is populated for the wrong zone"])

                # LEN_UNIT value should be populated when STATIC_BFE or DEPTH is populated
                if static_bfe != -9999 or depth != -9999.0:
                    if not len_unit:
                        self.errors.append(
                            [unique_id, "LEN_UNIT not populated for STATIC_BFE or DEPTH"])
                    elif len_unit.strip() == '':
                        self.errors.append(
                            [unique_id, "LEN_UNIT not populated for STATIC_BFE or DEPTH"])

                # LEN_UNIT value should not be populated
                if static_bfe == -9999 and depth == -9999.0:
                    if len_unit:
                        if len_unit.strip() != '':
                            self.errors.append(
                                [unique_id,
                                 "LEN_UNIT populated for NULL STATIC_BFE and NULL DEPTH values"])

                # VELOCITY field not populated for VEL_UNIT ('AO')
                if fld_zone == 'AO':
                    if not velocity:
                        self.errors.append(
                            [unique_id, "VELOCITY field is not populated for AO FLD_ZONE"])

                # VELOCITY field populated for non-AO Zone
                if fld_zone != 'AO' and velocity:
                    if velocity != -9999:
                        self.errors.append(
                            [unique_id, "VELOCITY field is populated for non-AO FLD_ZONE"])

                # VELOCITY value not valid (VELOCITY <> -9999 and VELOCITY < 0)
                if velocity:
                    if velocity < 0 and velocity != -9999:
                        self.errors.append([unique_id, "VELOCITY is less than 0 and not -9999"])

                # VEL_UNIT field not populated for VELOCITY
                if velocity:
                    if velocity != -9999:
                        if not vel_unit:
                            self.errors.append(
                                [unique_id, "VEL_UNIT field not populated for VELOCITY"])
                        elif vel_unit.strip() == '':
                            self.errors.append(
                                [unique_id, "VEL_UNIT field not populated for VELOCITY"])

    def s_fld_haz_ln_check(self, in_feature_class):
        """QC check of S_Fld_Haz_Ln"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'FLD_LN_ID', 'LN_TYP', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Fld_Haz_Ln_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

    def s_gage_check(self, in_feature_class):
        """QC check of S_Gage"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'GAGE_ID', 'WTR_NM', 'AGENCY', 'GAGE_DESC', 'GAGE_TYP',
                  'START_PD', 'END_PD', 'DRAIN_AREA', 'AREA_UNIT', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'GAGE_TYP', 'REC_INTRVL', 'TIME_UNIT', 'START_PD', 'END_PD',
                     'DRAIN_AREA', 'AREA_UNIT')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Gage_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                gage_typ = row[1]
                rec_intrvl = row[2]
                time_unit = row[3]
                start_pd = row[4]
                end_pd = row[5]
                drain_area = row[6]
                area_unit = row[7]

                # If the GAGE_TYP is Fixed Interval, then REC_INTRVL should be populated
                if gage_typ in ['1100', 'Fixed Interval']:
                    if not rec_intrvl:
                        self.errors.append(
                            [unique_id,
                             'If the GAGE_TYP is Fixed Interval, ' +
                             'then REC_INTRVL should be populated'])
                    elif rec_intrvl.strip() == '':
                        self.errors.append(
                            [unique_id,
                             'If the GAGE_TYP is Fixed Interval, ' +
                             'then REC_INTRVL should be populated'])

                # If the GAGE_TYP is Fixed Interval, then TIME_UNIT should be populated
                if gage_typ in ['1100', 'Fixed Interval']:
                    if not time_unit:
                        self.errors.append(
                            [unique_id,
                             'If the GAGE_TYP is Fixed Interval, ' +
                             'then TIME_UNIT should be populated'])
                    elif time_unit.strip() == '':
                        self.errors.append(
                            [unique_id,
                             'If the GAGE_TYP is Fixed Interval, ' +
                             'then TIME_UNIT should be populated'])

                # START_PD date should be less than END_PD date
                if end_pd and start_pd:  # Check them only if they aren't NULL/None
                    if end_pd < start_pd:
                        self.errors.append([unique_id, 'START_PD should be earlier than END_PD'])

                # If the DRAIN_AREA is not -8888, then AREA_UNIT should not be NP
                if drain_area and drain_area != -8888 and area_unit == 'NP':
                    self.errors.append(
                        [unique_id,
                         "If the DRAIN_AREA is not -8888, then AREA_UNIT should not be NP"])

    def s_gen_struct_check(self, in_feature_class):
        """QC check of S_Gen_Struct"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'STRUCT_ID', 'STRUCT_TYP',
                  'WTR_NM', 'SHOWN_FIRM', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'STRUCT_TYP', 'STRUC_DESC', 'SHOWN_FIRM')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Change the domains to correct spec
        if self.schema == '2020':
            domains = S_Gen_Struct_domains_2020
        else:
            domains = S_Gen_Struct_domains

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  domains, required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check if STRUCT_ID matches a value in L_Cst_Struct.STRUCT_ID for coastal structures
        coastal_structure_list = ('1012', 'Dock', '1015', 'Fish Ladder', '1020', 'Jetty', '1022',
                                  'Lock', '1026', 'Seawall')
        struct_typ_field = arcpy.AddFieldDelimiters(in_feature_class, "STRUCT_TYP")
        for error in self.__id_table_check(in_feature_class, 'STRUCT_ID',
                                           self.workspace + '\\L_Cst_Struct' + self.dbf_ext,
                                           'STRUCT_ID',
                                           query=struct_typ_field + " IN " + str(
                                               coastal_structure_list)):
            self.errors.append(error)

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                struct_typ = row[1]
                struc_desc = row[2]
                shown_firm = row[3]

                # If the STRUCT_TYP is 'Contained', then STRUC_DESC should be populated
                if struct_typ in ['1033',
                                  '0.2 PCT Annual Chance Flood Discharge Contained in Structure',
                                  '1032',
                                  '1 percent Annual Chance Flood Discharge Contained in Structure']:
                    if not struc_desc:
                        self.errors.append(
                            [unique_id,
                             'If the STRUCT_TYP is \'Contained\', ' +
                             'then STRUC_DESC should be populated'])
                    elif struc_desc.strip() == '':
                        self.errors.append(
                            [unique_id,
                             'If the STRUCT_TYP is \'Contained\', ' +
                             'then STRUC_DESC should be populated'])

                # SHOWN_FIRM should be 'T' or 'F' not 'U'
                if shown_firm == 'U':
                    self.errors.append([unique_id, 'SHOWN_FIRM should be \'T\' or \'F\' not \'U\''])

    def s_hwm_check(self, in_feature_class):
        """QC check of S_HWM"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'HWM_ID', 'WTR_NM', 'LOC_DESC', 'EVENT_DT', 'ELEV',
                  'LEN_UNIT', 'V_DATUM', 'HWM_SOURCE', 'APX_FREQ', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'ELEV', 'LEN_UNIT', 'V_DATUM')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_HWM_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                elev = row[1]
                len_unit = row[2]
                v_datum = row[3]

                # LEN_UNIT should be populated for ELEV value
                if elev and len_unit:
                    if elev != -8888 and len_unit == 'NP':
                        self.errors.append(
                            [unique_id, 'LEN_UNIT should be populated for ELEV value'])

                # V_DATUM does not match STUDY_INFO.V_DATUM
                if v_datum:
                    if v_datum != self.v_datum and self.v_datum != '' and \
                            self.mip_task in ['Develop Final Mapping Products Data Capture',
                                              'Draft FIRM Database Data Capture',
                                              'Produce Preliminary Products Data Capture']:
                        self.errors.append([unique_id, "V_DATUM value of " + v_datum +
                                            " does not match the V_DATUM value in Study_Info"])

    def s_hydro_reach_check(self, in_feature_class):
        """QC check of S_Hydro_Reach"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'REACH_ID', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'UP_NODE', 'DN_NODE')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Hydro_Reach_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_Nodes.UP_NODE
        up_node_field = arcpy.AddFieldDelimiters(in_feature_class, "UP_NODE")
        for error in self.__id_table_check(
                in_feature_class, 'UP_NODE',
                self.workspace + self.dataset + '\\S_Nodes' + self.shp_ext, 'NODE_ID',
                query=up_node_field + " is not null and " + up_node_field + " not in ('', ' ')"):
            self.errors.append(error)

        # Check for matching ID value in S_Nodes.DN_NODE
        dn_node_field = arcpy.AddFieldDelimiters(in_feature_class, "DN_NODE")
        for error in self.__id_table_check(
                in_feature_class, 'DN_NODE',
                self.workspace + self.dataset + '\\S_Nodes' + self.shp_ext, 'NODE_ID',
                query=dn_node_field + " is not null and " + dn_node_field + " not in ('', ' ')"):
            self.errors.append(error)

    def s_label_ld_check(self, in_feature_class):
        """QC check of S_Label_Ld"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'LEADER_ID', 'LABEL_TYPE', 'FIRM_PAN', 'SCALE')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Change the domains if 2017 or 2018 spec is used
        if self.schema in ['2017', '2018']:
            domains = S_Label_Ld_domains_2017
        else:
            domains = S_Label_Ld_domains

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2], domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_FIRM_Pan.FIRM_PAN
        for error in self.__id_table_check(
                in_feature_class, 'FIRM_PAN',
                self.workspace + self.dataset + '\\S_FIRM_Pan' + self.shp_ext, 'FIRM_PAN'):
            self.errors.append(error)

    def s_label_pt_check(self, in_feature_class):
        """QC check of S_Label_Pt"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'LABEL_ID', 'LABEL', 'LABEL_TYPE', 'FONT_SIZE',
                  'FONT_TYPE', 'FONT_STYLE', 'DEGREES', 'FIRM_PAN', 'SCALE')

        # Fields for specific checks
        spec_list = (fields[2], 'DEGREES')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Change the domains if 2017 or 2018 spec is used
        if self.schema in ['2017', '2018']:
            domains = S_Label_Pt_domains_2017
        else:
            domains = S_Label_Pt_domains

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2], domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_FIRM_Pan.FIRM_PAN
        for error in self.__id_table_check(
                in_feature_class, 'FIRM_PAN',
                self.workspace + self.dataset + '\\S_FIRM_Pan' + self.shp_ext, 'FIRM_PAN'):
            self.errors.append(error)

        # DEGREES Field Not Between -360 and 360 Check
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                degrees = row[1]

                if degrees:
                    if degrees < 0:
                        self.errors.append([unique_id, 'DEGREES less than 0'])
                    elif degrees > 359:
                        self.errors.append([unique_id, 'DEGREES greater than 359'])

    def s_levee_check(self, in_feature_class):
        """QC check of S_Levee"""
        # Required fields
        fields = ['DFIRM_ID', 'VERSION_ID', 'LEVEE_ID', 'FC_SYS_ID', 'LEVEE_NM', 'LEVEE_TYP',
                  'WTR_NM', 'BANK_LOC', 'USACE_LEV', 'PL84_99TF', 'LEVEE_STAT', 'OWNER',
                  'LEN_UNIT', 'SOURCE_CIT']

        # Update the fields if schema 2019 or 2020 are used
        if self.schema in ['2019', '2020']:
            fields.append('LEV_AN_TYP')
            fields.append('FC_SEG_ID')

        # Fields for specific checks
        spec_list = [fields[2], 'USACE_LEV', 'DISTRICT', 'LEVEE_STAT', 'PAL_DATE', 'FREEBOARD',
                     'LEN_UNIT']

        # Update the specific fields if schema 2019 or 2020 are used
        if self.schema in ['2019', '2020']:
            spec_list.append('LEV_AN_TYP')

        # Change the domains if 2017 or 2018 spec is used
        if self.schema in ['2017', '2018']:
            domains = S_Levee_domains_2017
        else:
            domains = S_Levee_domains

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2], domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                usace_lev = row[1]
                district = row[2]
                levee_stat = row[3]
                pal_date = row[4]
                freeboard = row[5]
                len_unit = row[6]

            # If USACE_LEV is 'T', then DISTRICT should cotain a District Code
            if usace_lev == 'T':
                if not district:
                    self.errors.append(
                        [unique_id,
                         "If USACE_LEV is 'T', then DISTRICT should contain District Code"])
                elif district.strip() == '':
                    self.errors.append(
                        [unique_id,
                         "If USACE_LEV is 'T', then DISTRICT should contain District Code"])

            # If USACE_LEV is not 'T', then DISTRICT should not cotain a District Code
            if usace_lev != 'T':
                if district:
                    if district.strip() != '':
                        self.errors.append(
                            [unique_id,
                             "If USACE_LEV not is 'T', then DISTRICT should not be populated"])

            # If LEVEE_STAT is Provisional, then PAL_DATE should be populated
            if levee_stat in ['P', 'Provisionally Accredited']:
                if pal_date:
                    if '9999' in str(pal_date):
                        self.errors.append(
                            [unique_id,
                             "If LEVEE_STAT is Provisional, then PAL_DATE should be populated"])

            # If LEVEE_STAT is not Provisional, then PAL_DATE should not be populated
            if levee_stat not in ['P', 'Provisionally Accredited']:
                if pal_date:
                    if '9999' not in str(pal_date):
                        self.errors.append(
                            [unique_id,
                             "If LEVEE_STAT is not Provisional, " +
                             "then PAL_DATE should not be populated"])

            # If FREEBOARD is populated, the LEN_UNIT should be populated
            if freeboard:
                if len_unit == 'NP' and freeboard != -9999.0:
                    self.errors.append(
                        [unique_id, "If FREEBOARD is populated, the LEN_UNIT should be populated"])

        # Perform specific feature class checks is 2019 or 2020 specs are used
        if self.schema in ['2019', '2020']:
            with SearchCursor(in_feature_class, spec_list) as cursor:
                for row in cursor:
                    unique_id = row[0]
                    levee_stat = row[3]
                    lev_an_typ = row[7]

                # If LEVEE_STAT is Non-Accredited, then LEV_AN_TYP should be populated
                if levee_stat in ['N', 'Non-Accredited']:
                    if lev_an_typ:
                        if lev_an_typ == 'NP':
                            self.errors.append(
                                [unique_id,
                                 "If LEVEE_STAT is Non-Accredited, " +
                                 "then LEV_AN_TYP should be populated and not NP"])

                # If LEVEE_STAT is not Non-Accredited, then LEV_AN_TYP should not be populated
                if levee_stat not in ['N', 'Non-Accredited']:
                    if lev_an_typ:
                        if lev_an_typ.strip() != 'NP':
                            self.errors.append(
                                [unique_id,
                                 "If LEVEE_STAT is not Non-Accredited, " +
                                 "then LEV_AN_TYP should not be populated"])

    def s_limwa_check(self, in_feature_class):
        """QC check of S_LiMWA"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'LIMWA_ID', 'SHOWN_FIRM', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'SHOWN_FIRM')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_LiMWA_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                shown_firm = row[1]

                if str(shown_firm) == 'U':
                    self.errors.append([unique_id, "SHOWN_FIRM should not be 'U'"])

    def s_lomr_check(self, in_feature_class):
        """QC check of S_LOMR"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'LOMR_ID', 'EFF_DATE',
                  'CASE_NO', 'STATUS', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_LOMR_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

    def s_nodes_check(self, in_feature_class):
        """QC check of S_Nodes"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'NODE_ID', 'WTR_NM',
                  'NODE_DESC', 'MODEL_ID', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Nodes_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check if the NODE_ID can't be found in L_Summary_Discharges,
        # L_Summary_Elevations, S_Hydro_Reach or in S_Subbasins
        discharges_id_list = []
        elevations_id_list = []
        reach_up_id_list = []
        reach_down_id_list = []
        subbasins_id_list = []

        if arcpy.Exists(self.workspace + '\\L_Summary_Discharges' + self.dbf_ext):
            discharges_id_list = sorted(
                list(set([str(row[0]) for row in SearchCursor(
                    self.workspace + '\\L_Summary_Discharges' + self.dbf_ext, 'NODE_ID')])))

        if arcpy.Exists(self.workspace + '\\L_Summary_Elevations' + self.dbf_ext):
            elevations_id_list = sorted(
                list(set([str(row[0]) for row in SearchCursor(
                    self.workspace + '\\L_Summary_Elevations' + self.dbf_ext, 'NODE_ID')])))

        if arcpy.Exists(self.workspace + self.dataset + '\\S_Hydro_Reach' + self.shp_ext):
            reach_up_id_list = sorted(
                list(set([str(row[0]) for row in
                          SearchCursor(self.workspace + self.dataset +
                                       '\\S_Hydro_Reach' + self.shp_ext, 'UP_NODE')])))

        if arcpy.Exists(self.workspace + self.dataset + '\\S_Hydro_Reach' + self.shp_ext):
            reach_down_id_list = sorted(
                list(set([str(row[0]) for row in
                          SearchCursor(self.workspace + self.dataset +
                                       '\\S_Hydro_Reach' + self.shp_ext, 'DN_NODE')])))

        if arcpy.Exists(self.workspace + self.dataset + '\\S_Subbasins' + self.shp_ext):
            subbasins_id_list = sorted(
                list(set([str(row[0]) for row in
                          SearchCursor(self.workspace + self.dataset +
                                       '\\S_Subbasins' + self.shp_ext, 'NODE_ID')])))

        node_id_list = sorted(
            list(set([str(row[0]) for row in SearchCursor(in_feature_class, 'NODE_ID')])))

        foreign_id_list = \
            discharges_id_list + elevations_id_list + reach_up_id_list + \
            reach_down_id_list + subbasins_id_list

        for node_id in node_id_list:
            if node_id not in foreign_id_list:
                self.errors.append([node_id,
                                    "ID not found in either L_Summary_Discharges (NODE_ID), "
                                    "L_Summary_Elevations (NODE_ID), S_Hydro_Reach "
                                    "(UP_NODE or DN_NODE) or S_Subbasins (NODE_ID)"])

    def s_pfd_ln_check(self, in_feature_class):
        """QC check of S_PFD_Ln"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'PFD_ID', 'VZONE_LIMT', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_PFD_Ln_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

    def s_plss_ar_check(self, in_feature_class):
        """QC check of S_PLSS_Ar"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'PLSS_AR_ID', 'SECT_NO', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'RANGE', 'SECT_NO', 'TWP')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_PLSS_Ar_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                range_num = row[1]
                sect_no = row[2]
                twp = row[3]

                # RANGE field is empty while SECT_NO is not 0
                if str(range_num).strip() in ['', 'None'] and \
                        str(sect_no).strip() not in ['', 'None', '0']:
                    self.errors.append([unique_id, "RANGE field is empty while SECT_NO is not 0"])

                # RANGE has a value when SECT_NO has a 0 value
                if str(range_num).strip() not in ['', 'None'] and str(sect_no).strip() in ['0']:
                    self.errors.append([unique_id, "RANGE has a value when SECT_NO has a 0 value"])

                # RANGE value does not end with E or W
                if not (str(range_num).endswith('E') or str(range_num).endswith('W')):
                    if str(sect_no).strip() != '0':
                        self.errors.append([unique_id, "RANGE value does not end with E or W"])

                # TWP field is empty while SECT_NO is not 0
                if str(twp).strip() in ['', 'None'] and \
                        str(sect_no).strip() not in ['', 'None', '0']:
                    self.errors.append([unique_id, "TWP field is empty while SECT_NO is not 0"])

                # TWP has a value when SECT_NO has a 0 value
                if str(twp).strip() not in ['', 'None'] and str(sect_no).strip() in ['0']:
                    self.errors.append([unique_id, "TWP has a value when SECT_NO has a 0 value"])

                # TWP value does not end with N or S
                if not (str(twp).endswith('N') or str(twp).endswith('S')):
                    if str(sect_no).strip() != '0':
                        self.errors.append([unique_id, "TWP value does not end with N or S"])

                # SECT_NO value is not between 0 and 36
                if sect_no:
                    if sect_no.isdigit():
                        if int(sect_no) not in range(0, 37):
                            self.errors.append([unique_id, "SECT_NO should be between 0 and 36"])
                    else:
                        self.errors.append([unique_id, "SECT_NO should be between 0 and 36"])

    def s_pol_ar_check(self, in_feature_class):
        """QC check of S_Pol_Ar"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'POL_AR_ID', 'POL_NAME1', 'CO_FIPS', 'ST_FIPS',
                  'COMM_NO', 'CID', 'ANI_TF', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'COMM_NO', 'CID', 'ST_FIPS', 'ANI_TF', 'ANI_FIRM', 'COM_NFO_ID')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Pol_Ar_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in L_Comm_Info.COM_NFO_ID
        if self.mip_task in ['Develop Final Mapping Products Data Capture',
                             'Draft FIRM Database Data Capture',
                             'Produce Preliminary Products Data Capture']:

            ani_tf_field = arcpy.AddFieldDelimiters(in_feature_class, "ANI_TF")
            comm_no_field = arcpy.AddFieldDelimiters(in_feature_class, "COMM_NO")

            for error in self.__id_table_check(
                    in_feature_class, 'COM_NFO_ID',
                    self.workspace + '\\L_Comm_Info' + self.dbf_ext, 'COM_NFO_ID',
                    query=ani_tf_field + " = 'F' AND " + comm_no_field + " NOT IN ('ST', 'FED', 'OTHR')"):
                self.errors.append(error)

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                comm_no = row[1]
                cid = row[2]
                st_fips = row[3]
                ani_tf = row[4]
                ani_firm = row[5]
                com_nfo_id = row[6]

                # COMM_NO value not valid
                if comm_no:
                    if comm_no not in ['ST', 'FED', 'OTHR']:
                        if not comm_no.isdigit():
                            self.errors.append([unique_id, "COMM_NO value not valid"])

                # CID Field not populated by ST_FIPS and COMM_NO
                if cid and st_fips and comm_no:
                    if (cid[0:2] != st_fips) or (cid[2:] != comm_no):
                        self.errors.append(
                            [unique_id, "CID Field not populated by ST_FIPS and COMM_NO"])

                # ANI_TF value is not be 'U'
                if str(ani_tf) == 'U':
                    self.errors.append([unique_id, "ANI_TF value should not be 'U'"])

                # ANI_FIRM not populated when ANI_TF is T
                if str(ani_tf) == 'T' and str(ani_firm).strip() in ['None', '']:
                    self.errors.append([unique_id, "ANI_FIRM not populated when ANI_TF is T"])

                # COM_NFO_ID field should be populated when COMM_NO is not ST, FED or OTHR
                if str(com_nfo_id).strip() in ['None', ''] and \
                        str(comm_no) not in ['ST', 'FED', 'OTHR']:
                    self.errors.append(
                        [unique_id,
                         "COM_NFO_ID field should be populated when COMM_NO is" +
                         " not ST, FED or OTHR"])

    def s_profil_basln_check(self, in_feature_class):
        """QC check of S_Profil_Basln"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'BASELN_ID', 'WTR_NM', 'WATER_TYP', 'STUDY_TYP',
                  'SHOWN_FIRM', 'R_ST_DESC', 'R_END_DESC', 'START_ID', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'SHOWN_FIRM', 'V_DATM_OFF', 'DATUM_UNIT', 'WATER_TYP', 'STUDY_TYP')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Change the domains if 2017 spec is used
        if self.schema == '2017':
            domains = S_Profil_Basln_domains_2017
        else:
            domains = S_Profil_Basln_domains

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2], domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_Stn_Start.START_ID
        for error in self.__id_table_check(
                in_feature_class, 'START_ID',
                self.workspace + self.dataset + '\\S_Stn_Start' + self.shp_ext, 'START_ID'):
            self.errors.append(error)

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                shown_firm = row[1]
                v_datm_off = row[2]
                datum_unit = row[3]
                water_typ = row[4]
                study_typ = row[5]

                # SHOWN_FIRM value should not be 'U'
                if shown_firm == 'U':
                    self.errors.append([unique_id, "SHOWN_FIRM value should not be 'U'"])

                # V_DATM_OFF should be populated if DATUM_UNIT is populated
                if datum_unit:
                    if datum_unit.strip() != '':
                        if v_datm_off is None:
                            self.errors.append(
                                [unique_id,
                                 "V_DATM_OFF should be populated if DATUM_UNIT is populated"])
                        elif v_datm_off.strip() == '':
                            self.errors.append(
                                [unique_id,
                                 "V_DATM_OFF should be populated if DATUM_UNIT is populated"])

                # DATUM_UNIT should be populated if V_DATM_OFF is populated
                if v_datm_off:
                    if v_datm_off.strip() != '':
                        if datum_unit is None:
                            self.errors.append(
                                [unique_id,
                                 "DATUM_UNIT should be populated if V_DATM_OFF is populated"])
                        elif datum_unit.strip() == '':
                            self.errors.append(
                                [unique_id,
                                 "DATUM_UNIT should be populated if V_DATM_OFF is populated"])

                # WATER_TYP should not be 'UNK'
                if water_typ in ['UNK', 'Unknown']:
                    self.errors.append([unique_id,
                                        "WARNING: WATER_TYP should not be 'UNK' or 'Unknown'"])

                # STUDY_TYP should not be 'NP'
                if study_typ == 'NP':
                    self.errors.append([unique_id, "STUDY_TYP should not be 'NP'"])

    def s_riv_mrk_check(self, in_feature_class):
        """QC check of S_Riv_Mrk"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'RIV_MRK_ID', 'START_ID', 'RIV_MRK_NO', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Riv_Mrk_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_Stn_Start.START_ID
        for error in self.__id_table_check(
                in_feature_class, 'START_ID',
                self.workspace + self.dataset + '\\S_Stn_Start' + self.shp_ext, 'START_ID'):
            self.errors.append(error)

    def s_stn_start_check(self, in_feature_class):
        """QC check of S_Stn_Start"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'START_ID', 'START_DESC', 'LOC_ACC', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'START_DESC')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Stn_Start_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check if the START_ID can't be found in S_Profil_Basln,
        # S_XS, S_Hydro_Reach or in S_Riv_Mrk
        baseline_id_list = []
        cross_section_id_list = []
        riv_mrk_list = []

        if arcpy.Exists(self.workspace + self.dataset + '\\S_Profil_Basln' + self.shp_ext):
            baseline_id_list = sorted(
                list(set([str(row[0]) for row in SearchCursor(
                    self.workspace + self.dataset + '\\S_Profil_Basln' + self.shp_ext,
                    'START_ID')])))

        if arcpy.Exists(self.workspace + self.dataset + '\\S_XS' + self.dbf_ext):
            cross_section_id_list = sorted(
                list(set([str(row[0]) for row in SearchCursor(
                    self.workspace + self.dataset + '\\S_XS' + self.dbf_ext, 'START_ID')])))

        if arcpy.Exists(self.workspace + self.dataset + '\\S_Riv_Mrk' + self.shp_ext):
            riv_mrk_list = sorted(
                list(set([str(row[0]) for row in SearchCursor(
                    self.workspace + self.dataset + '\\S_Riv_Mrk' + self.shp_ext, 'START_ID')])))

        start_id_list = sorted(
            list(set([str(row[0]) for row in SearchCursor(in_feature_class, 'START_ID')])))

        foreign_id_list = baseline_id_list + cross_section_id_list + riv_mrk_list

        for start_id in start_id_list:
            if start_id not in foreign_id_list:
                self.errors.append(
                    [start_id, "START_ID not found in either S_Profil_Basln, S_XS, or S_Riv_Mrk"])

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                start_desc = row[1]

                # START_DESC value should not be 'NP'
                if str(start_desc) == 'NP':
                    self.errors.append([unique_id, "START_DESC value should not be 'NP'"])

    def s_subbasins_check(self, in_feature_class):
        """QC check of S_Subbasins"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'SUBBAS_ID', 'SUBBAS_NM', 'HUC8', 'WTR_NM',
                  'BASIN_DESC', 'SUB_AREA', 'AREA_UNIT', 'BASIN_TYP', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'HUC8', 'SUB_AREA', 'AREA_UNIT')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Subbasins_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_Nodes.NODE_ID
        node_id_field = arcpy.AddFieldDelimiters(in_feature_class, "NODE_ID")
        for error in self.__id_table_check(
                in_feature_class, 'NODE_ID',
                self.workspace + self.dataset + '\\S_Nodes' + self.shp_ext, 'NODE_ID',
                query=node_id_field + " is not null and " + node_id_field + " not in ('', ' ')"):
            self.errors.append(error)

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                huc8 = row[1]
                sub_area = row[2]
                area_unit = row[3]

                # HUC8 should be 8 characters in length
                if huc8:
                    if len(huc8) != 8:
                        self.errors.append([unique_id, "HUC8 should be 8 characters in length"])

                # HUC8 should not be NP
                if huc8 == 'NP':
                    self.errors.append([unique_id, "HUC8 should not be NP"])

                # SUB_AREA should not be -8888
                if sub_area == -8888:
                    self.errors.append([unique_id, "SUB_AREA should not be -8888"])

                # SUB_AREA should be greater than 0
                if sub_area < 0 and sub_area != -8888:
                    self.errors.append([unique_id, "SUB_AREA should be greater than 0"])

                # AREA_UNIT should not be 'NP'
                if area_unit == 'NP':
                    self.errors.append([unique_id, "AREA_UNIT should not be 'NP'"])

    def s_submittal_info_check(self, in_feature_class):
        """QC check of S_Submittal_Info"""
        # Required fields
        fields = ['DFIRM_ID', 'VERSION_ID', 'SUBINFO_ID', 'CASE_NO', 'CASE_DESC', 'SUBMIT_BY',
                  'METHOD_TYP', 'COMP_DATE', 'TASK_TYP', 'EFF_DATE', 'CONTRCT_NO', 'SOURCE_CIT']

        # Fields for specific checks
        spec_list = [fields[2], 'CASE_NO', 'CASE_DESC', 'SUBMIT_BY', 'COMP_DATE', 'TASK_TYP',
                     'METHOD_TYP']

        # Update fields, spec_list and domains if based on 2017 or 2018 specs
        if self.schema in ['2017', '2018']:
            fields.pop(fields.index('METHOD_TYP'))
            fields.append('STUDY_TYP')
            spec_list.pop(spec_list.index('METHOD_TYP'))
            spec_list.append('STUDY_TYP')
            domains = S_Submittal_Info_domains_2017
        else:
            domains = S_Submittal_Info_domains

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2], domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in L_Cst_Model.CST_MDL_ID
        cst_mdl_id_field = arcpy.AddFieldDelimiters(in_feature_class, "CST_MDL_ID")
        for error in self.__id_table_check(
                in_feature_class, 'CST_MDL_ID',
                self.workspace + '\\L_Cst_Model' + self.dbf_ext, 'CST_MDL_ID',
                query=cst_mdl_id_field + " is not null and " + cst_mdl_id_field + " not in ('', ' ')"):
            self.errors.append(error)

        # Specific checks for this feature class
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                case_no = row[1]
                case_desc = row[2]
                submit_by = row[3]
                comp_date = row[4]
                task_typ = row[5]
                meth_study_typ = row[6]

                # CASE_NO should not be 'NP'
                if case_no == 'NP':
                    self.errors.append([unique_id, "CASE_NO should not be 'NP'"])

                # CASE_DESC should not be 'NP'
                if case_desc == 'NP':
                    self.errors.append([unique_id, "CASE_DESC should not be 'NP'"])

                # SUBMIT_BY should not be 'NP'
                if submit_by == 'NP':
                    self.errors.append([unique_id, "SUBMIT_BY should not be 'NP'"])

                # COMP_DATE should be '8/8/8888'
                if '8888' in str(comp_date):
                    self.errors.append([unique_id, "COMP_DATE should not be '8/8/8888'"])

                # TASK_TYP should not 'NP'
                if task_typ == 'NP':
                    self.errors.append([unique_id, "TASK_TYP should not 'NP'"])

                # Only perform these checks if 'METHOD_TYP' exists in the spec_list
                if 'METHOD_TYP' in spec_list:
                    # Check for TASK_TYPs that should not have a METHOD_TYP of NP
                    if task_typ in ['1000', 'ALLUVIAL FAN', '1020', 'COASTAL',
                                    '1040', 'FLOODPLAIN MAPPING', '1050', 'HYDRAULIC',
                                    '1060', 'HYDROLOGIC', '1300', 'Levee Seclusion'] and \
                            meth_study_typ == 'NP':
                        self.errors.append(
                            [unique_id, "METHOD should not be 'NP' for this TASK_TYP"])

                    # Check for TASK_TYPs that should have a METHOD_TYP of NP
                    if task_typ in ['1010', 'BASE MAP', '1030', 'FIRM DATABASE',
                                    '1070', 'SURVEY', '1080', 'NEW TOPO CAPTURE',
                                    '1081', 'EXISTING TOPO CAPTURE',
                                    '1082', 'TERRAIN CAPTURE FIRM ',
                                    '1090', 'DISCOVERY', '1100', 'FLOOD RISK ASSESSMENT',
                                    '1200', 'LOMR', 'NP', 'NP'] and \
                            meth_study_typ != 'NP':
                        self.errors.append([unique_id, "METHOD should be 'NP' for this TASK_TYP"])

    def s_topo_confidence_check(self, in_feature_class):
        """QC check of S_Topo_Confidence"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'LOWCONF_ID', 'CONF_TYPE', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Topo_Confidence_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

    def s_trnsport_ln_check(self, in_feature_class):
        """QC check of S_Trnsport_Ln"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'TRANS_ID', 'MTFCC',
                  'FULLNAME', 'ROUTE_TYP', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Trnsport_Ln_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

    def s_tsct_basln_check(self, in_feature_class):
        """QC check of S_Tsct_Basln"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'TBASELN_ID', 'TBASE_TYP', 'R_ST_DESC', 'R_END_DESC',
                  'V_DATUM', 'WTR_NM', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Tsct_Basln_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_Cst_Tsct_Ln.TBASELN_ID
        for error in self.__id_table_check(
                in_feature_class, 'TBASELN_ID',
                self.workspace + self.dataset + '\\S_Cst_Tsct_Ln' + self.shp_ext, 'TBASELN_ID'):
            self.errors.append(error)

        # Check for matching ID value in L_Cst_Model.CST_MDL_ID
        cst_mdl_id_field = arcpy.AddFieldDelimiters(in_feature_class, "CST_MDL_ID")
        for error in self.__id_table_check(
                in_feature_class, 'CST_MDL_ID',
                self.workspace + '\\L_Cst_Model' + self.dbf_ext, 'CST_MDL_ID',
                query=cst_mdl_id_field + " is not null and " + cst_mdl_id_field + " not in ('', ' ')"):
            self.errors.append(error)

    def s_wtr_ar_check(self, in_feature_class):
        """QC check of S_Wtr_Ar"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'WTR_AR_ID', 'WTR_NM', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'SHOWN_FIRM', 'SHOWN_INDX')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Wtr_Ar_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                shown_firm = row[1]
                shown_indx = row[2]

                # Perform these checks on specific MIP Tasks only
                if self.mip_task in ['Develop Final Mapping Products Data Capture',
                                     'Draft FIRM Database Data Capture',
                                     'Produce Preliminary Products Data Capture']:
                    # SHOWN_FIRM should not be 'U'
                    if shown_firm == 'U':
                        self.errors.append([unique_id, "SHOWN_FIRM should not be 'U'"])

                    # SHOWN_INDX should not be 'U'
                    if shown_indx == 'U':
                        self.errors.append([unique_id, "SHOWN_INDX should not be 'U'"])

    def s_wtr_ln_check(self, in_feature_class):
        """QC check of S_Wtr_Ln"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'WTR_LN_ID', 'WTR_NM', 'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'SHOWN_FIRM', 'SHOWN_INDX')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_Wtr_Ln_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                shown_firm = row[1]
                shown_indx = row[2]

                # Perform these checks on specific MIP Tasks only
                if self.mip_task in ['Develop Final Mapping Products Data Capture',
                                     'Draft FIRM Database Data Capture',
                                     'Produce Preliminary Products Data Capture']:
                    # SHOWN_FIRM should not be 'U'
                    if shown_firm == 'U':
                        self.errors.append([unique_id, "SHOWN_FIRM should not be 'U'"])

                    # SHOWN_INDX should not be 'U'
                    if shown_indx == 'U':
                        self.errors.append([unique_id, "SHOWN_INDX should not be 'U'"])

    def s_xs_check(self, in_feature_class):
        """QC check of S_XS"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'XS_LN_ID', 'WTR_NM', 'STREAM_STN', 'START_ID',
                  'XS_LN_TYP', 'WSEL_REG', 'STRMBED_EL', 'LEN_UNIT', 'V_DATUM', 'MODEL_ID',
                  'SOURCE_CIT')

        # Fields for specific checks
        spec_list = (fields[2], 'WTR_NM', 'STREAM_STN', 'START_ID', 'XS_LTR', 'XS_LN_TYP',
                     'LEN_UNIT', 'V_DATUM', 'WSEL_REG', 'STRMBED_EL')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_feature_class, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_feature_class, fields[2],
                                                  S_XS_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in L_XS_Elev.XS_LN_ID
        for error in self.__id_table_check(in_feature_class, 'XS_LN_ID',
                                           self.workspace + '\\L_XS_Elev' + self.dbf_ext,
                                           'XS_LN_ID'):
            self.errors.append(error)

        # Check for matching ID value in S_Stn_Start.START_ID
        for error in self.__id_table_check(
                in_feature_class, 'START_ID',
                self.workspace + self.dataset + '\\S_Stn_Start' + self.shp_ext, 'START_ID'):
            self.errors.append(error)

        # Perform specific feature class checks
        with SearchCursor(in_feature_class, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                wtr_nm = row[1]
                stream_stn = row[2]
                start_id = row[3]
                xs_ltr = row[4]
                xs_ln_typ = row[5]
                len_unit = row[6]
                v_datum = row[7]
                wsel_reg = row[8]
                strmbed_el = row[9]

                # WTR_NM should not be 'NP'
                if wtr_nm == 'NP':
                    self.errors.append([unique_id, "WTR_NM should not be 'NP'"])

                # STREAM_STN should not be '-8888'
                if stream_stn == -8888:
                    self.errors.append([unique_id, "STREAM_STN should not be '-8888'"])

                # START_ID should not be 'NP'
                if start_id == 'NP':
                    self.errors.append([unique_id, "START_ID should not be 'NP'"])

                # XS_LTR is empty or has a NULL when XS_LN_TYP is 'LETTERED'
                if xs_ltr is None:
                    if xs_ln_typ in ['1010', 'LETTERED, MAPPED']:
                        self.errors.append(
                            [unique_id,
                             "XS_LTR is empty or has a NULL when XS_LN_TYP is 'LETTERED'"])
                elif xs_ltr.strip() == '' and xs_ln_typ in ['1010', 'LETTERED, MAPPED']:
                    self.errors.append(
                        [unique_id, "XS_LTR is empty or has a NULL when XS_LN_TYP is 'LETTERED'"])

                # XS_LTR should not be populated when XS_LN_TYP is not 'LETTERED'
                if xs_ltr is not None:
                    if xs_ltr.strip() != '' and xs_ln_typ not in ['1010', 'LETTERED, MAPPED']:
                        self.errors.append(
                            [unique_id,
                             "XS_LTR should not be populated when XS_LN_TYP is not 'LETTERED'"])

                # LEN_UNIT should not be 'NP'
                if len_unit == 'NP':
                    self.errors.append([unique_id, "LEN_UNIT should not be 'NP'"])

                # V_DATUM does not match the V_DATUM value in Study_Info
                if v_datum != self.v_datum and self.v_datum != '' and \
                        self.mip_task in ['Develop Final Mapping Products Data Capture',
                                          'Draft FIRM Database Data Capture',
                                          'Produce Preliminary Products Data Capture']:
                    self.errors.append([unique_id, "V_DATUM value of " + str(row[7]) +
                                        " does not match the V_DATUM value in Study_Info"])

                # Is STRMBED_EL greater than the WSEL_REG
                if wsel_reg and strmbed_el:
                    if wsel_reg != -8888 and strmbed_el != -8888:
                        if wsel_reg < strmbed_el:
                            self.errors.append(
                                [unique_id, "STRMBED_EL should be greater or equal to WSEL_REG"])

    def study_info_check(self, in_table):
        """QC check of Study_Info"""
        # Required fields
        fields = ['DFIRM_ID', 'VERSION_ID', 'STD_NFO_ID', 'STUDY_NM', 'STATE_NM', 'CNTY_NM',
                  'LG_PAN_NO', 'OPP_TF', 'H_DATUM', 'V_DATUM', 'PROJECTION', 'PROJ_ZONE',
                  'PROJ_UNIT', 'LANDWD_VAL', 'CW_TF', 'RTROFT_TF', 'META_NM', 'FIS_NM',
                  'LOGO_NM', 'INDX_EFFDT', 'DBREV_DT']

        # Fields for specific checks
        spec_list = [fields[2], 'LG_PAN_NO', 'OPP_TF', 'RTROFT_TF', 'META_NM', 'FIS_NM',
                     'H_DATUM', 'V_DATUM', 'PROJECTION', 'PROJ_ZONE', 'PROJ_UNIT',
                     'PROJ_SUNIT', 'PROJ_SECND', 'PROJ_SZONE', 'CW_TF']

        # Add field for the 2021 Updates.
        if self.schema == "2021":
            spec_list.append("INDX_SUFFX")

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], Study_Info_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                lg_pan_no = row[1]
                opp_tf = row[2]
                rtroft_tf = row[3]
                meta_nm = row[4]
                fis_nm = row[5]
                h_datum = row[6]
                v_datum = row[7]
                projection = row[8]
                proj_zone = row[9]
                proj_unit = row[10]
                proj_sunit = row[11]
                proj_secnd = row[12]
                proj_szone = row[13]
                cw_tf = row[14]

                if self.schema == "2021":
                    index_suffix = row[15]

                # S_FIRM_PAN path
                firm_pan = self.workspace + self.dataset + '\\S_FIRM_Pan' + self.shp_ext

                # LG_PAN_NO value does not match largest panel number
                largest_panel = 0
                if arcpy.Exists(firm_pan):
                    panel_list = list([row[0] for row in SearchCursor(firm_pan, "PANEL")])
                    largest_panel = max(panel_list)
                if lg_pan_no != largest_panel:
                    self.errors.append([unique_id,
                                        "LG_PAN_NO value does not match largest panel number"])

                # OPP_TF should not be 'U'
                if opp_tf == 'U':
                    self.errors.append([unique_id, "OPP_TF should not be 'U'"])

                # OPP_TF value F with only 1 printed panel
                printed_panel_count = 0
                if arcpy.Exists(firm_pan):
                    panel_typ = arcpy.AddFieldDelimiters(firm_pan, "PANEL_TYP")
                    printed_panel_types = ('1000', 'Countywide, Panel Printed',
                                           '1020', 'Community Based, Panel Printed',
                                           '1050', 'Statewide, Panel Printed')
                    panel_list = list([row[0] for row in
                                       SearchCursor(firm_pan, "PANEL",
                                                    panel_typ + " IN " + str(printed_panel_types))])
                    printed_panel_count = len(panel_list)
                if opp_tf == 'F' and printed_panel_count == 1:
                    self.errors.append([unique_id, "OPP_TF value F with only 1 printed panel"])

                # OPP_TF value T with multiple printed panels
                if opp_tf == 'T' and printed_panel_count > 1:
                    self.errors.append([unique_id, "OPP_TF value T with multiple printed panels"])

                # CW_TF should not be 'U'
                if cw_tf == 'U':
                    self.errors.append([unique_id, "CW_TF should not be 'U'"])

                # RTROFT_TF should not be 'U'
                if rtroft_tf == 'U':
                    self.errors.append([unique_id, "RTROFT_TF should not be 'U'"])

                # META_NM field does not end with '.txt' or '.xml'
                if str(meta_nm)[-4:] not in ['.txt', '.xml']:
                    self.errors.append([unique_id,
                                        "META_NM field does not end with '.txt' or '.xml'"])

                # Multiple rows with the STUDY_INFO table
                result = arcpy.GetCount_management(in_table)
                count = int(result[0])
                if count > 1:
                    self.errors.append(
                        [unique_id, "META_NM field not properly formatted for Effective studies"])

                # FIS_NM field does not end with '.pdf'
                if str(fis_nm)[-4:] not in ['.pdf']:
                    self.errors.append([unique_id, "FIS_NM field does not end with '.pdf'"])

                # H_DATUM should not be 'NP'
                if h_datum == 'NP':
                    self.errors.append([unique_id, "H_DATUM should not be 'NP'"])

                # V_DATUM should not be 'NP'
                if v_datum == 'NP':
                    self.errors.append([unique_id, "V_DATUM should not be 'NP'"])

                # PROJECTION should not be 'NP'
                if projection == 'NP':
                    self.errors.append([unique_id, "PROJECTION should not be 'NP'"])

                # PROJ_ZONE should not be 'NP' when PROJECTION is not 'GCS'
                if str(projection) not in ['GCS', 'GEOGRAPHIC COORDINATE SYSTEM'] \
                        and proj_zone == 'NP':
                    self.errors.append(
                        [unique_id, "PROJ_ZONE should not be 'NP' when PROJECTION is not 'GCS'"])

                # PROJ_UNIT should not be 'NP'
                if proj_unit == 'NP':
                    self.errors.append([unique_id, "PROJ_UNIT should not be 'NP'"])

                # PROJ_SUNIT should be populated if PROJ_SECND is populated
                if proj_secnd is not None:
                    if proj_secnd.strip() != '':
                        if proj_sunit is None:
                            self.errors.append(
                                [unique_id,
                                 "PROJ_SUNIT should be populated if PROJ_SECND is populated"])
                        elif proj_sunit.strip() == '' or proj_sunit == 'NP':
                            self.errors.append(
                                [unique_id,
                                 "PROJ_SUNIT should be populated if PROJ_SECND is populated"])

                # PROJ_SUNIT should not be populated if PROJ_SECND is not populated
                if proj_secnd is None:
                    if proj_sunit is not None:
                        if proj_sunit.strip() != '':
                            self.errors.append(
                                [unique_id,
                                 "PROJ_SUNIT should not be populated if "
                                 "PROJ_SECND is not populated"])
                elif proj_secnd.strip() == '':
                    if proj_sunit is not None:
                        if proj_sunit.strip() != '':
                            self.errors.append(
                                [unique_id,
                                 "PROJ_SUNIT should not be populated if "
                                 "PROJ_SECND is not populated"])

                # PROJ_SZONE should be not be 'NP' when PROJ_SECND is not 'GCS'
                if proj_secnd is not None:
                    if proj_secnd.strip() != '':
                        if str(proj_secnd) not in ['GCS', 'GEOGRAPHIC COORDINATE SYSTEM'] \
                                and proj_szone == 'NP':
                            self.errors.append(
                                [unique_id,
                                 "PROJ_SZONE should not be 'NP' when PROJ_SECND is not 'GCS'"])

                # PROJ_SZONE should not be populated when PROJ_SECND is not populated
                if proj_secnd is None:
                    if proj_szone is not None:
                        if proj_szone.strip() != '':
                            self.errors.append(
                                [unique_id,
                                 "PROJ_SZONE should not be populated if PROJ_SECND"
                                 " is not populated"])
                elif proj_secnd.strip() == '':
                    if proj_szone is not None:
                        if proj_szone.strip() != '':
                            self.errors.append(
                                [unique_id,
                                 "PROJ_SZONE should not be populated if "
                                 "PROJ_SECND is not populated"])

                # Check that INDX_SUFFX is populated.  This is a required field, however, it is
                # only one character wide therefore 'NP' cannot be used.
                if self.schema == "2021":
                    # Create a list of acceptable values.  List should include A-Z, all uppercase and
                    # excluding letters I and O.
                    letters = list(string.ascii_uppercase)
                    letters.remove('I')
                    letters.remove('O')

                    if index_suffix is None:
                        self.errors.append(
                            [unique_id, "INDX_SUFFX should be populated"])
                    elif index_suffix.strip() == "":
                        self.errors.append(
                            [unique_id, "INDX_SUFFX should be populated"])
                    elif index_suffix not in letters:
                        self.errors.append(
                            [unique_id, "INDX_SUFFX should be a value from A-Z exluding I and O"])

    def l_comm_info_check(self, in_table):
        """QC check of L_Comm_Info"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'COM_NFO_ID', 'REPOS_ADR1', 'REPOS_CITY', 'REPOS_ST',
                  'REPOS_ZIP', 'IN_ID_DAT', 'IN_NFIP_DT', 'IN_FHBM_DT', 'IN_FRM_DAT', 'FST_CW_EFF',
                  'FST_CW_FIS', 'REVISIONS', 'MULTICO_TF', 'FLOODPRONE', 'FIS_INCLUD',
                  'RECENT_FIS')

        # Fields for specific checks
        spec_list = (fields[2], 'REPOS_ZIP', 'IN_ID_DAT', 'IN_FRM_DAT', 'FST_CW_EFF', 'RECENT_DAT',
                     'RECENT_FIS', 'IN_NFIP_DT', 'MULTICO_TF', 'FLOODPRONE', 'FIS_INCLUD',
                     'FST_CW_FIS', 'IN_FHBM_DT', 'REVISIONS')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_Comm_Info_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_Pol_Ar.COM_NFO_ID
        for error in self.__id_table_check(
                in_table, 'COM_NFO_ID',
                self.workspace + self.dataset + '\\S_Pol_Ar' + self.shp_ext, 'COM_NFO_ID'):
            self.errors.append(error)

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                repos_zip = row[1]
                in_id_dat = row[2]
                in_frm_dat = row[3]
                fst_cw_eff = row[4]
                recent_dat = row[5]
                recent_fis = row[6]
                in_nfip_dt = row[7]
                multico_tf = row[8]
                floodprone = row[9]
                fis_includ = row[10]
                fst_cw_fis = row[11]
                in_fhbm_dt = row[12]
                revisions = row[13]

                # REPOS_ZIP value not 5 or 9 characters
                if repos_zip:
                    if not (len(repos_zip) == 5 or len(repos_zip) == 9):
                        self.errors.append([unique_id, "REPOS_ZIP value not 5 or 9 characters"])

                # IN_ID_DAT should be earlier than IN_FRM_DAT
                if in_id_dat and in_frm_dat:
                    if in_id_dat > in_frm_dat and '8888' not in str(in_frm_dat):
                        self.errors.append(
                            [unique_id, "IN_ID_DAT should be earlier than IN_FRM_DAT"])

                # IN_ID_DAT should be earlier than FST_CW_EFF
                if in_id_dat and fst_cw_eff:
                    if in_id_dat > fst_cw_eff and '8888' not in str(fst_cw_eff):
                        self.errors.append(
                            [unique_id, "IN_ID_DAT should be earlier than FST_CW_EFF"])

                # IN_ID_DAT should be earlier than FST_CW_FIS
                if in_id_dat and fst_cw_fis:
                    if in_id_dat > fst_cw_fis and '8888' not in str(fst_cw_fis):
                        self.errors.append(
                            [unique_id, "IN_ID_DAT should be earlier than FST_CW_FIS"])

                # IN_ID_DAT should be earlier than RECENT_DAT
                if in_id_dat and recent_dat:
                    if in_id_dat > recent_dat and '8888' not in str(recent_dat):
                        self.errors.append(
                            [unique_id, "IN_ID_DAT should be earlier than RECENT_DAT"])

                # IN_ID_DAT should be earlier than RECENT_FIS
                if in_id_dat and recent_fis:
                    if in_id_dat > recent_fis and '8888' not in str(recent_fis):
                        self.errors.append(
                            [unique_id, "IN_ID_DAT should be earlier than RECENT_FIS"])

                # IN_ID_DAT should not be 8/8/8888
                if '8888' in str(in_id_dat):
                    self.errors.append([unique_id, "IN_ID_DAT should not be 8/8/8888"])

                # IN_NFIP_DT should be earlier or equal to IN_FRM_DAT
                if in_nfip_dt and in_frm_dat:
                    if in_nfip_dt > in_frm_dat and '8888' not in str(in_frm_dat):
                        self.errors.append(
                            [unique_id, "IN_NFIP_DT should be earlier than IN_FRM_DAT"])

                # IN_NFIP_DT should be earlier or equal to FST_CW_EFF
                if in_nfip_dt and fst_cw_eff:
                    if in_nfip_dt > fst_cw_eff and '8888' not in str(fst_cw_eff):
                        self.errors.append(
                            [unique_id, "IN_NFIP_DT should be earlier than FST_CW_EFF"])

                # IN_NFIP_DT should be earlier or equal to FST_CW_FIS
                if in_nfip_dt and fst_cw_fis:
                    if in_nfip_dt > fst_cw_fis and '8888' not in str(fst_cw_fis):
                        self.errors.append(
                            [unique_id, "IN_NFIP_DT should be earlier than FST_CW_FIS"])

                # IN_NFIP_DT should be earlier or equal to RECENT_DAT
                if in_nfip_dt and recent_dat:
                    if in_nfip_dt > recent_dat and '8888' not in str(recent_dat):
                        self.errors.append(
                            [unique_id, "IN_NFIP_DT should be earlier than RECENT_DAT"])

                # IN_NFIP_DT should be earlier or equal to RECENT_FIS
                if in_nfip_dt and recent_fis:
                    if in_nfip_dt > recent_fis and '8888' not in str(recent_fis):
                        self.errors.append(
                            [unique_id, "IN_NFIP_DT should be earlier than RECENT_FIS"])

                # IN_NFIP_DT should not be 8/8/8888
                if '8888' in str(in_nfip_dt):
                    self.errors.append(
                        [unique_id, "WARNING: Verify if the IN_NFIP_DT should be populated"])

                # IN_FRM_DAT should be later than IN_FHBM_DT
                if in_frm_dat and in_fhbm_dt:
                    if in_frm_dat < in_fhbm_dt and '8888' not in str(in_fhbm_dt):
                        self.errors.append(
                            [unique_id, "IN_FRM_DAT should be later than IN_FHBM_DT"])

                # IN_FRM_DAT should be earlier or equal to FST_CW_EFF
                if in_frm_dat and fst_cw_eff:
                    if in_frm_dat > fst_cw_eff and '8888' not in str(fst_cw_eff):
                        self.errors.append(
                            [unique_id, "IN_FRM_DAT should be earlier than FST_CW_EFF"])

                # IN_FRM_DAT should be earlier or equal to FST_CW_FIS
                if in_frm_dat and fst_cw_fis:
                    if in_frm_dat > fst_cw_fis and '8888' not in str(fst_cw_fis):
                        self.errors.append(
                            [unique_id, "IN_FRM_DAT should be earlier than FST_CW_FIS"])

                # IN_FRM_DAT should be earlier or equal to RECENT_DAT
                if in_frm_dat and recent_dat:
                    if in_frm_dat > recent_dat and '8888' not in str(recent_dat):
                        self.errors.append(
                            [unique_id, "IN_FRM_DAT should be earlier than RECENT_DAT"])

                # IN_FRM_DAT should be earlier or equal to RECENT_FIS
                if in_frm_dat and recent_fis:
                    if in_frm_dat > recent_fis and '8888' not in str(recent_fis):
                        self.errors.append(
                            [unique_id, "IN_FRM_DAT should be earlier than RECENT_FIS"])

                # FST_CW_EFF should be earlier or equal to RECENT_DAT
                if fst_cw_eff and recent_dat:
                    if fst_cw_eff > recent_dat and '9999' not in str(recent_dat):
                        self.errors.append(
                            [unique_id, "FST_CW_EFF should be earlier than RECENT_DAT"])

                # FST_CW_EFF should be earlier or equal to RECENT_FIS
                if fst_cw_eff and recent_fis:
                    if fst_cw_eff > recent_fis and '8888' not in str(recent_fis):
                        self.errors.append(
                            [unique_id, "FST_CW_EFF should be earlier than RECENT_FIS"])

                # REVISIONS should not be 'U'
                if revisions == 'U':
                    self.errors.append([unique_id, "REVISIONS should not be 'U'"])

                # MULTICO_TF should not be 'U'
                if multico_tf == 'U':
                    self.errors.append([unique_id, "MULTICO_TF should not be 'U'"])

                # FLOODPRONE should not be 'U'
                if floodprone == 'U':
                    self.errors.append([unique_id, "FLOODPRONE should not be 'U'"])

                # FIS_INCLUD should not be 'U'
                if fis_includ == 'U':
                    self.errors.append([unique_id, "FIS_INCLUD should not be 'U'"])

    def l_comm_revis_check(self, in_table):
        """QC check of L_Comm_Revis"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'COM_REV_ID', 'COM_NFO_ID', 'REVIS_DATE')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_Comm_Revis_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in L_Comm_Info.COM_NFO_ID
        for error in self.__id_table_check(in_table, 'COM_NFO_ID',
                                           self.workspace + '\\L_Comm_Info' + self.dbf_ext,
                                           'COM_NFO_ID'):
            self.errors.append(error)

    def l_cst_model_check(self, in_table):
        """QC check of L_Cst_Model"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'CST_MDL_ID', 'HUC8', 'WTR_NM', 'LIMIT_FROM',
                  'LIMIT_TO', 'EROS_TF', 'PFD_TF', 'HAZARDEVAL')

        # Fields for specific checks
        spec_list = (fields[2], 'HUC8', 'WTR_NM', 'LIMIT_FROM', 'LIMIT_TO')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Change the domains if 2017 spec is used
        if self.schema == '2017':
            domains = L_Cst_Model_domains_2017
        else:
            domains = L_Cst_Model_domains

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check if the CST_MDL_ID can't be found in S_Cst_Gage, S_Cst_Tsct_Ln,
        # S_Submittal_Info or in S_Tsct_Basln
        gages_id_list = []
        if arcpy.Exists(self.workspace + self.dataset + '\\S_Cst_Gage' + self.shp_ext):
            gages_id_list = sorted(
                list(set(
                    [str(row[0]) for row in
                     SearchCursor(self.workspace + '\\S_Cst_Gage' + self.shp_ext,
                                  'CST_MDL_ID')])))

        transect_id_list = []
        if arcpy.Exists(self.workspace + self.dataset + '\\S_Cst_Tsct_Ln' + self.shp_ext):
            transect_id_list = sorted(
                list(set([str(row[0]) for row in
                          SearchCursor(self.workspace + '\\S_Cst_Tsct_Ln' + self.shp_ext,
                                       'CST_MDL_ID')])))

        subbasin_id_list = []
        if arcpy.Exists(self.workspace + self.dataset + '\\S_Submittal_Info' + self.shp_ext):
            subbasin_id_list = sorted(
                list(set([str(row[0]) for row in
                          SearchCursor(self.workspace + '\\S_Submittal_Info' + self.shp_ext,
                                       'CST_MDL_ID')])))

        node_id_list = sorted(
            list(set([str(row[0]) for row in SearchCursor(in_table, 'CST_MDL_ID')])))

        foreign_id_list = gages_id_list + transect_id_list + subbasin_id_list

        for node_id in node_id_list:
            if node_id not in foreign_id_list:
                self.errors.append(
                    [node_id, "ID not found in the CST_MDL_ID field in either S_Cst_Gage, "
                              "S_Cst_Tsct_Ln, S_Submittal_Info or S_Tsct_Basln"])

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                huc8 = row[1]
                wtr_nm = row[2]
                limit_from = row[3]
                limit_to = row[4]

                # HUC8 should not be 'NP'
                if huc8 == 'NP':
                    self.errors.append([unique_id, "HUC8 should not be 'NP'"])

                    # WTR_NM should not be 'NP'
                if wtr_nm == 'NP':
                    self.errors.append([unique_id, "WTR_NM should not be 'NP'"])

                # LIMIT_FROM should not be 'NP'
                if limit_from == 'NP':
                    self.errors.append([unique_id, "LIMIT_FROM should not be 'NP'"])

                # LIMIT_TO should not be 'NP'
                if limit_to == 'NP':
                    self.errors.append([unique_id, "LIMIT_TO should not be 'NP'"])

    def l_cst_struct_check(self, in_table):
        """QC check of L_Cst_Struct"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'CST_STR_ID', 'STRUCT_ID', 'WTR_NM', 'CERT_STAT',
                  'STRUCT_LEN', 'LEN_UNIT', 'STRUCT_MTL')

        # Fields for specific checks
        spec_list = (fields[2], 'WTR_NM', 'STRUCT_LEN', 'LEN_UNIT')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Change the domains to correct spec
        if self.schema == '2020':
            domains = L_XS_Struct_domains_2020
        else:
            domains = L_XS_Struct_domains

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_Gen_Struct.STRUCT_ID
        for error in self.__id_table_check(
                in_table, 'STRUCT_ID',
                self.workspace + self.dataset + '\\S_Gen_Struct' + self.shp_ext, 'STRUCT_ID'):
            self.errors.append(error)

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                wtr_nm = row[1]
                struct_len = row[2]
                len_unit = row[3]

                # WTR_NM should not be 'NP'
                if wtr_nm == 'NP':
                    self.errors.append([unique_id, "WTR_NM should not be 'NP'"])

                # STRUCT_LEN should be greater than 0
                if struct_len:
                    if struct_len <= 0:
                        self.errors.append([unique_id, "STRUCT_LEN should be greater than 0"])

                # LEN_UNIT should not be NP
                if len_unit == 'NP':
                    self.errors.append([unique_id, "LEN_UNIT should not be 'NP'"])

    def l_cst_tsct_elev_check(self, in_table):
        """QC check of L_Cst_Tsct_Elev"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'CT_INFO_ID', 'TRAN_LN_ID', 'EVENT_TYP')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_Cst_Tsct_Elev_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_Cst_Tsct_Ln.TRAN_LN_ID
        for error in self.__id_table_check(
                in_table, 'TRAN_LN_ID',
                self.workspace + self.dataset + '\\S_Cst_Tsct_Ln' + self.shp_ext, 'TRAN_LN_ID'):
            self.errors.append(error)

    def l_manningsn_check(self, in_table):
        """QC check of L_ManningsN"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'MANN_ID', 'WTR_NM', 'CHANNEL_N',
                  'OVERBANK_N', 'LANDCOVER')

        # Fields for specific checks
        spec_list = (fields[2], 'WTR_NM', 'CHANNEL_N', 'OVERBANK_N', 'LANDCOVER')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_ManningsN_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                wtr_nm = row[1]
                channel_n = row[2]
                overbank_n = row[3]
                landcover = row[4]

                # WTR_NM should not be 'NP'
                if wtr_nm == 'NP':
                    self.errors.append([unique_id, "WTR_NM should not be 'NP'"])

                # CHANNEL_N should not be 'NP'
                if channel_n == 'NP':
                    self.errors.append([unique_id, "CHANNEL_N should not be 'NP'"])

                # OVERBANK_N should not be 'NP'
                if overbank_n == 'NP':
                    self.errors.append([unique_id, "OVERBANK_N should not be 'NP'"])

                # LANDCOVER should not be 'NP'
                if landcover == 'NP':
                    self.errors.append([unique_id, "LANDCOVER should not be 'NP'"])

    def l_meetings_check(self, in_table):
        """QC check of L_Meetings"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'MTG_ID', 'COM_NFO_ID', 'MTG_TYP', 'MTG_DATE',
                  'MTG_LOC', 'MTG_PURP', 'FIS_EFF_DT')

        # Fields for specific checks
        spec_list = (fields[2], 'MTG_TYP', 'MTG_DATE', 'MTG_LOC', 'MTG_PURP')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2],
                                                  L_Meetings_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in L_Mtg_POC.MTG_ID
        for error in self.__id_table_check(in_table, 'MTG_ID',
                                           self.workspace + '\\L_Mtg_POC' + self.dbf_ext, 'MTG_ID'):
            self.errors.append(error)

        # Check for matching ID value in L_Comm_Info.COM_NFO_ID
        for error in self.__id_table_check(in_table, 'COM_NFO_ID',
                                           self.workspace + '\\L_Comm_Info' + self.dbf_ext,
                                           'COM_NFO_ID'):
            self.errors.append(error)

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                mtg_typ = row[1]
                mtg_date = row[2]
                mtg_loc = row[3]
                mtg_purp = row[4]

                # MTG_TYP should not be 'NP'
                if mtg_typ == 'NP':
                    self.errors.append([unique_id, "MTG_TYP should not be 'NP'"])

                # MTG_DATE should not be '8/8/8888'
                if '8888' in str(mtg_date):
                    self.errors.append([unique_id, "MTG_DATE should not be '8/8/8888'"])

                # MTG_LOC should not be 'NP'
                if mtg_loc == 'NP':
                    self.errors.append([unique_id, "MTG_LOC should not be 'NP'"])

                # MTG_PURP should not be 'NP'
                if mtg_purp == 'NP':
                    self.errors.append([unique_id, "MTG_PURP should not be 'NP'"])

    def l_mt2_lomr_check(self, in_table):
        """QC check of L_Mt2_Lomr"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'LOMR_ID', 'CASE_NO', 'EFF_DATE', 'WTR_NM',
                  'FIRM_PAN', 'STATUS', 'SCALE')

        # Fields for specific checks
        spec_list = (fields[2], 'CASE_NO', 'EFF_DATE', 'WTR_NM', 'STATUS')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_MT2_LOMR_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in FIRM_PAN.FIRM_PAN
        for error in self.__id_table_check(
                in_table, 'FIRM_PAN',
                self.workspace + self.dataset + '\\S_FIRM_Pan' + self.shp_ext, 'FIRM_PAN'):
            self.errors.append(error)

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                case_no = row[1]
                eff_date = row[2]
                wtr_nm = row[3]
                status = row[4]

                # CASE_NO should not be 'NP'
                if case_no == 'NP':
                    self.errors.append([unique_id, "CASE_NO should not be 'NP'"])

                # EFF_DATE should not be '8/8/8888'
                if '8888' in str(eff_date):
                    self.errors.append([unique_id, "EFF_DATE should not be '8/8/8888'"])

                # WTR_NM should not be 'NP'
                if wtr_nm == 'NP':
                    self.errors.append([unique_id, "WTR_NM should not be 'NP'"])

                # STATUS should not be 'NP'
                if status == 'NP':
                    self.errors.append([unique_id, "STATUS should not be 'NP'"])

    def l_mtg_poc_check(self, in_table):
        """QC check of L_Mtg_Poc"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'POC_ID', 'MTG_ID', 'POC_NAME', 'FIRST_NAME',
                  'LAST_NAME', 'AGENCY', 'CEO', 'FPA', 'SHMO', 'GIS')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_Mtg_POC_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in L_Meetings.MTG_ID
        for error in self.__id_table_check(in_table, 'MTG_ID',
                                           self.workspace + '\\L_Meetings' + self.dbf_ext,
                                           'MTG_ID'):
            self.errors.append(error)

    def l_pan_revis_check(self, in_table):
        """QC check of L_Pan_Revis"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'REVIS_ID', 'FIRM_PAN', 'REVIS_DATE', 'REVIS_NOTE')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_Pan_Revis_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_FIRM_Pan.FIRM_PAN
        for error in self.__id_table_check(
                in_table, 'FIRM_PAN',
                self.workspace + self.dataset + '\\S_FIRM_Pan' + self.shp_ext, 'FIRM_PAN'):
            self.errors.append(error)

    def l_pol_fhbm_check(self, in_table):
        """QC check of L_Pol_FHBM"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'FHBM_ID', 'COM_NFO_ID', 'FHBM_DATE', 'FHBM_NOTE')

        # Fields for specific checks
        spec_list = ()

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_Pol_FHBM_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in L_Comm_Info.COM_NFO_ID
        for error in self.__id_table_check(in_table, 'COM_NFO_ID',
                                           self.workspace + '\\L_Comm_Info' + self.dbf_ext,
                                           'COM_NFO_ID'):
            self.errors.append(error)

    def l_profil_bkwtr_el_check(self, in_table):
        """QC check of L_Profil_Bkwtr_El"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'PROF_BW_ID', 'WTR_NM',
                  'EVENT_TYP', 'BKWTR_WSEL', 'LEN_UNIT', 'V_DATUM')

        # Fields for specific checks
        spec_list = (fields[2], 'LEN_UNIT', 'V_DATUM')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_Profil_Bkwtr_El_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching WTR_NM value in S_Profil_Basln.WTR_NM
        for error in self.__id_table_check(
                in_table, 'WTR_NM',
                self.workspace + self.dataset + '\\S_Profil_Basln' + self.shp_ext, 'WTR_NM',
                error_message="WTR_NM in L_Profil_Bkwtr_El can not be found in S_Profil_Basln"):
            self.errors.append(error)

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                len_unit = row[1]
                v_datum = row[2]

                # LEN_UNIT should not be 'NP'
                if len_unit == 'NP':
                    self.errors.append([unique_id, "LEN_UNIT should not be 'NP'"])

                # V_DATUM does not match the V_DATUM value in Study_Info
                if v_datum != self.v_datum and self.v_datum != '' and \
                        self.mip_task in ['Develop Final Mapping Products Data Capture',
                                          'Draft FIRM Database Data Capture',
                                          'Produce Preliminary Products Data Capture']:
                    self.errors.append([unique_id, "V_DATUM value of " + str(row[2]) +
                                        " does not match the V_DATUM value in Study_Info"])

        # BKWTR_WSEL values does not match cross section elevation and
        # all backwater values are less than the L_XS_ELEV WSEL values
        if arcpy.Exists(self.workspace + self.dataset + '\\S_XS' + self.shp_ext) and \
                arcpy.Exists(self.workspace + '\\L_XS_Elev' + self.dbf_ext):

            # Make the needed table views
            xs_lyr = arcpy.MakeTableView_management(
                self.workspace + self.dataset + '\\S_XS' + self.shp_ext, "xs_lyr")
            xs_elev = arcpy.MakeTableView_management(self.workspace + '\\L_XS_Elev' + self.dbf_ext,
                                                     "xs_elev")

            # Perform the join between s_xs and l_xs_elev
            xs_join = arcpy.AddJoin_management(xs_elev, "XS_LN_ID", xs_lyr, "XS_LN_ID")

            # Create a tuple of wtr_nm, event_typ and wsel from the joined table
            join_tuple = tuple([(str(row[0]), str(row[1]), str(round(row[2], 1)))
                                for row in
                                SearchCursor(xs_join, ['S_XS.WTR_NM', 'L_XS_Elev.EVENT_TYP',
                                                       'L_XS_Elev.WSEL'])])

            # Remove the join
            arcpy.RemoveJoin_management(xs_elev)

            # Remove the table views
            arcpy.Delete_management("xs_lyr")
            arcpy.Delete_management("xs_elev")

            # Create a tuple of the prof_bw_id, wtr_nm, event_typ and
            # bkwtr_wsel from l_profil_bkwter_elev
            bkwtr_dict = {}

            with SearchCursor(in_table,
                              ['PROF_BW_ID', 'WTR_NM', 'EVENT_TYP', 'BKWTR_WSEL']) as cursor:
                for row in cursor:
                    bkwtr_dict[str(row[0])] = (str(row[1]), str(row[2]), str(round(row[3], 1)))

            # Look for missing entries in the join_tuple that exist in the bkwtr_dict
            for key in bkwtr_dict:
                if bkwtr_dict[key] not in join_tuple:
                    self.errors.append([key,
                                        "WARNING: Backwater elevation for the " + bkwtr_dict[key][
                                            1] + " event was not found in L_XS_ELEV"])

            # Check that no L_XS_ELEV are less than the backwater value
            for key in bkwtr_dict:
                water_name = bkwtr_dict[key][0]  # Backwater water name
                event = bkwtr_dict[key][1]  # Backwater event
                wsel = bkwtr_dict[key][2]  # Backwater wsel value

                for join_water_name, join_event, join_wsel in join_tuple:
                    if water_name == join_water_name and \
                            event == join_event and \
                            float(wsel) > float(join_wsel):
                        self.errors.append(
                            [key, "WARNING: The WSEL for the " + bkwtr_dict[key][1] +
                             " event was greater than a L_XS_ELEV value. "
                             "'Check for Flooding Controlled By...'"])

    def l_profil_label_check(self, in_table):
        """QC check of L_Profil_Label"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'PROFLBL_ID', 'WTR_NM', 'STREAM_STN', 'ELEV', 'DESCR',
                  'ORIENT', 'ADJUSTED', 'UNDERLINE', 'LEN_UNIT', 'V_DATUM')

        # Fields for specific checks
        spec_list = (fields[2], 'UNDERLINE', 'LEN_UNIT', 'V_DATUM', 'ELEV')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_Profil_Label_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                underline = row[1]
                len_unit = row[2]
                v_datum = row[3]
                elev = row[4]

                # UNDERLINE should not be 'U'
                if underline == 'U':
                    self.errors.append([unique_id, "UNDERLINE should not be 'U'"])

                # LEN_UNIT should not be 'NP'
                if len_unit == 'NP':
                    self.errors.append([unique_id, "LEN_UNIT should not be 'NP'"])

                # V_DATUM does not match the V_DATUM value in Study_Info
                if v_datum != self.v_datum and self.v_datum != '' and \
                        self.mip_task in ['Develop Final Mapping Products Data Capture',
                                          'Draft FIRM Database Data Capture',
                                          'Produce Preliminary Products Data Capture']:
                    self.errors.append([unique_id, "V_DATUM value of " + str(row[3]) +
                                        " does not match the V_DATUM value in Study_Info"])

                # ELEV should not be -8888 or less than 0
                if (elev < 0 and elev != -8888) or elev == -8888:
                    self.errors.append([unique_id, "ELEV should not be -8888 or less than 0"])

    def l_profil_panel_check(self, in_table):
        """QC check of L_Profil_Panel"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'PROFPAN_ID', 'WTR_NM', 'PANEL_NO', 'FIS_PAN_NO',
                  'START_STN', 'END_STN', 'START_ELEV', 'END_ELEV', 'ORIGIN_X', 'ORIGIN_Y',
                  'H_SCALE', 'V_SCALE', 'LEN_UNIT', 'V_DATUM')

        # Fields for specific checks
        spec_list = (fields[2], 'PANEL_NO', 'START_STN', 'END_STN', 'START_ELEV', 'END_ELEV',
                     'ORIGIN_X', 'ORIGIN_Y', 'H_SCALE', 'LEN_UNIT', 'V_DATUM')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_Profil_Panel_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                panel_no = row[1]
                start_stn = row[2]
                end_stn = row[3]
                start_elev = row[4]
                end_elev = row[5]
                origin_x = row[6]
                origin_y = row[7]
                h_scale = row[8]
                len_unit = row[9]
                v_datum = row[10]

                # PANEL_NO should not be -8888 or less than 0
                if (panel_no < 0 and panel_no != -8888) or panel_no == -8888:
                    self.errors.append([unique_id, "PANEL_NO should not be -8888 or less than 0"])

                # START_STN should not be -8888
                if start_stn == -8888:
                    self.errors.append([unique_id, "START_STN should not be -8888"])

                # END_STN should not be -8888 or less than 0
                if (end_stn < 0 and end_stn != -8888) or end_stn == -8888:
                    self.errors.append([unique_id, "END_STN should not be -8888 or less than 0"])

                # START_ELEV should not be -8888
                if start_elev == -8888:
                    self.errors.append([unique_id, "START_ELEV should not be -8888"])

                # END_ELEV should not be -8888 or less than 0
                if (end_elev < 0 and end_elev != -8888) or end_elev == -8888:
                    self.errors.append([unique_id, "END_ELEV should not be -8888 or less than 0"])

                # ORIGIN_X should not be -8888
                if origin_x == -8888:
                    self.errors.append([unique_id, "ORIGIN_X should not be -8888"])

                # ORIGIN_Y should not be -8888
                if origin_y == -8888:
                    self.errors.append([unique_id, "ORIGIN_Y should not be -8888"])

                # H_SCALE should not be -8888 or less than 0
                if (h_scale < 0 and h_scale != -8888) or h_scale == -8888:
                    self.errors.append([unique_id, "H_SCALE should not be -8888 or less than 0"])

                # LEN_UNIT should not be 'NP'
                if len_unit == 'NP':
                    self.errors.append([unique_id, "LEN_UNIT should not be 'NP'"])

                # V_DATUM does not match the V_DATUM value in Study_Info
                if v_datum != self.v_datum and self.v_datum != '' and \
                        self.mip_task in ['Develop Final Mapping Products Data Capture',
                                          'Draft FIRM Database Data Capture',
                                          'Produce Preliminary Products Data Capture']:
                    self.errors.append([unique_id, "V_DATUM value of " + str(row[3]) +
                                        " does not match the V_DATUM value in Study_Info"])

    def l_source_cit_check(self, in_table):
        """QC check of L_Source_Cit"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'SOURCE_CIT', 'CITATION', 'PUBLISHER', 'TITLE',
                  'PUB_DATE', 'MEDIA')

        # Fields for specific checks
        spec_list = (fields[2], 'SRC_SCALE')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_Source_Cit_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for extra SOURCE_CIT values compared to all other tables
        remaining_sources = [source for source in self.source_citations]
        fc_list = arcpy.ListFeatureClasses("*", "All", self.dataset)  # List of the feature classes
        for feature_class in fc_list:
            # Get a list of fields for the current feature class
            fc_field_list = arcpy.ListFields(os.path.join(self.workspace, feature_class))
            for field in fc_field_list:
                # Only look for feature classes with a SOURCE_CIT field
                if field.name == "SOURCE_CIT":
                    # Get a list of source_cits for the current feature classes
                    sources = \
                        list(set([str(fc_row[0]) for fc_row in SearchCursor(
                            os.path.join(self.workspace, feature_class), 'SOURCE_CIT')]))
                    # If the source exists in the L_Source_Cit table, remove it from the list
                    for source in sources:
                        if source in remaining_sources:
                            remaining_sources.remove(source)

        # Any remaining sources are an error
        for source in remaining_sources:
            if source[0:3] != 'REF':  # Skip the REF source citations
                self.errors.append([source, "Source Citation is not found in any spatial tables"])

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]

                src_scale = row[1]

                # SRC_SCALE format should be like '1:24000'
                if src_scale:
                    src_scale = src_scale.strip()
                    if src_scale[0:2] != '1:' and unique_id[0:3].lower() != 'ref':
                        self.errors.append([unique_id, "SRC_SCALE format should be like '1:24000'"])
                    if unique_id[0:3].lower() == 'ref' and src_scale[0:2].strip() != "":
                        self.errors.append([unique_id, "WARNING: SRC_SCALE usually not" +
                                            " populated for REF sources"])

    def l_summary_discharges_check(self, in_table):
        """QC check of L_Summary_Discharges"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'SUMDSCH_ID', 'NODE_ID', 'NODE_DESC', 'DRAIN_AREA',
                  'AREA_UNIT', 'EVENT_TYP', 'DISCH', 'DISCH_UNIT', 'SHOWN_FIS')

        # Fields for specific checks
        spec_list = (fields[2], 'AREA_UNIT', 'DISCH_UNIT', 'WSEL_UNIT', 'SHOWN_FIS', 'V_DATUM')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2],
                                                  L_Summary_Discharges_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_Nodes.NODE_ID
        for error in self.__id_table_check(
                in_table, 'NODE_ID',
                self.workspace + self.dataset + '\\S_Nodes' + self.shp_ext, 'NODE_ID'):
            self.errors.append(error)

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                area_unit = row[1]
                disch_unit = row[2]
                wsel_unit = row[3]
                shown_fis = row[4]
                v_datum = row[5]

                # AREA_UNIT should not be 'NP'
                if area_unit == 'NP':
                    self.errors.append([unique_id, "AREA_UNIT should not be 'NP'"])

                # DISCH_UNIT should not be 'NP'
                if disch_unit == 'NP':
                    self.errors.append([unique_id, "DISCH_UNIT should not be 'NP'"])

                # WSEL_UNIT should not be 'NP'
                if wsel_unit == 'NP':
                    self.errors.append([unique_id, "WSEL_UNIT should not be 'NP'"])

                # SHOWN_FIS should not be 'U'
                if shown_fis == 'U':
                    self.errors.append([unique_id, "SHOWN_FIS should not be 'U'"])

                # V_DATUM does not match value in Study_Info
                if v_datum:
                    if v_datum != self.v_datum and v_datum.strip() != '' and \
                            self.v_datum != '' and \
                            self.mip_task in ['Develop Final Mapping Products Data Capture',
                                              'Draft FIRM Database Data Capture',
                                              'Produce Preliminary Products Data Capture']:
                        self.errors.append([unique_id, "V_DATUM value of " + str(row[3]) +
                                            " does not match the V_DATUM value in Study_Info"])

    def l_summary_elevations_check(self, in_table):
        """QC check of L_Summary_Elevations"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'SUMELEV_ID', 'NODE_ID', 'EVENT_TYP', 'WSEL',
                  'WSEL_UNIT', 'V_DATUM', 'SHOWN_FIS')

        # Fields for specific checks
        spec_list = (fields[2], 'V_DATUM', 'SHOWN_FIS', 'WSEL', 'WSEL_UNIT')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2],
                                                  L_Summary_Elevations_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # Check for matching ID value in S_Nodes.NODE_ID
        for error in self.__id_table_check(
                in_table, 'NODE_ID',
                self.workspace + self.dataset + '\\S_Nodes' + self.shp_ext, 'NODE_ID'):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                v_datum = row[1]
                shown_fis = row[2]
                wsel = row[3]
                wsel_unit = row[4]

                # V_DATUM does not match value in Study_Info
                if v_datum != self.v_datum and self.v_datum != '' and \
                        self.mip_task in ['Develop Final Mapping Products Data Capture',
                                          'Draft FIRM Database Data Capture',
                                          'Produce Preliminary Products Data Capture']:
                    self.errors.append([unique_id, "V_DATUM value of " + str(row[3]) +
                                        " does not match the V_DATUM value in Study_Info"])

                # SHOWN_FIS should not be 'U'
                if shown_fis == 'U':
                    self.errors.append([unique_id, "SHOWN_FIS should not be 'U'"])

                # WSEL should not be 'NP'
                if wsel == 'NP':
                    self.errors.append([unique_id, "WSEL should not be 'NP'"])

                # WSEL_UNIT should not be 'NP'
                if wsel_unit == 'NP':
                    self.errors.append([unique_id, "WSEL_UNIT should not be 'NP'"])

    def l_survey_pt_check(self, in_table):
        """QC check of L_Survey_Pt"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'SURVPT_ID', 'SURVSTR_ID', 'SURV_CODE', 'STRUCTDESC',
                  'NORTHING', 'EASTING', 'ELEV', 'ELEV_UNIT', 'H_DATUM', 'V_DATUM', 'PROJECTION',
                  'PROJ_ZONE', 'PROJ_UNIT')

        # Fields for specific checks
        spec_list = (fields[2], 'ELEV_UNIT', 'H_DATUM', 'V_DATUM', 'PROJECTION', 'PROJ_UNIT')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_Survey_Pt_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                elev_unit = row[1]
                h_datum = row[2]
                v_datum = row[3]
                projection = row[4]
                proj_unit = row[5]

                # ELEV_UNIT should not be NP
                if elev_unit == 'NP':
                    self.errors.append([unique_id, "ELEV_UNIT should not be NP"])

                # H_DATUM should not be NP
                if h_datum == 'NP':
                    self.errors.append([unique_id, "H_DATUM should not be NP"])

                # V_DATUM does not match value in Study_Info
                if v_datum != self.v_datum and self.v_datum != '' and \
                        self.mip_task in ['Develop Final Mapping Products Data Capture',
                                          'Draft FIRM Database Data Capture',
                                          'Produce Preliminary Products Data Capture']:
                    self.errors.append([unique_id, "V_DATUM value of " + str(row[3]) +
                                        " does not match the V_DATUM value in Study_Info"])

                # PROJECTION should not be NP
                if projection == 'NP':
                    self.errors.append([unique_id, "PROJECTION should not be NP"])

                # PROJ_UNIT should not be NP
                if proj_unit == 'NP':
                    self.errors.append([unique_id, "PROJ_UNIT should not be NP"])

    def l_xs_elev_check(self, in_table):
        """QC check of L_XS_Elev"""
        # Required fields
        fields = ['DFIRM_ID', 'VERSION_ID', 'XS_ELEV_ID', 'XS_LN_ID', 'EVENT_TYP', 'WSEL',
                  'LEN_UNIT', 'V_DATUM', 'LEVEE_TF', 'CALC_WO_BW']

        # Fields for specific checks
        spec_list = [fields[2], 'FW_WIDTH', 'FW_WIDTHIN', 'NE_WIDTH_L', 'NE_WIDTH_R', 'XS_AREA',
                     'AREA_UNIT', 'VELOCITY', 'VEL_UNIT', 'EVENT_TYP', 'WSEL', 'WSEL_WOFWY',
                     'WSEL_FLDWY', 'WSEL_INCRS', 'LEN_UNIT', 'V_DATUM', 'LEVEE_TF', 'LVSCENARIO',
                     'CALC_WO_BW']

        # Add the EVAL_LN field if using 2020 spec
        if self.schema == '2020':
            fields.append("EVAL_LN")
            spec_list.append("EVAL_LN")

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_XS_Elev_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_XS.XS_LN_ID
        for error in self.__id_table_check(in_table, 'XS_LN_ID',
                                           self.workspace + self.dataset + '\\S_XS' + self.shp_ext,
                                           'XS_LN_ID'):
            self.errors.append(error)

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                fw_width = row[1]
                fw_widthin = row[2]
                ne_width_l = row[3]
                ne_width_r = row[4]
                xs_area = row[5]
                area_unit = row[6]
                velocity = row[7]
                vel_unit = row[8]
                event_typ = row[9]
                wsel = row[10]
                wsel_wofwy = row[11]
                wsel_fldwy = row[12]
                wsel_incrs = row[13]
                len_unit = row[14]
                v_datum = row[15]
                levee_tf = row[16]
                lvscenario = row[17]
                calc_wo_bc = row[18]

                if self.schema == "2020":
                    eval_ln = row[19]

                # FW_WIDTH should not be populated if the EVENT_TYP is not for
                # the 1% Annual Return Period
                if fw_width != -9999 and event_typ not in ('01pct', '1 Percent Chance'):
                    self.errors.append(
                        [unique_id, "FW_WIDTH should not be populated for this Return Period"])

                # FW_WIDTHIN should not be populated if the EVENT_TYP is not for
                # the 1% Annual Return Period
                if fw_widthin != -9999 and event_typ not in ('01pct', '1 Percent Chance'):
                    self.errors.append(
                        [unique_id, "FW_WIDTHIN should not be populated for this Return Period"])

                # NE_WIDTH_L should not be populated if the EVENT_TYP is not for
                # the 1% Annual Return Period
                if ne_width_l != -9999 and event_typ not in ('01pct', '1 Percent Chance'):
                    self.errors.append(
                        [unique_id, "NE_WIDTH_L should not be populated for this Return Period"])

                # NE_WIDTH_R should not be populated if the EVENT_TYP is not for
                # the 1% Annual Return Period
                if ne_width_r != -9999 and event_typ not in ('01pct', '1 Percent Chance'):
                    self.errors.append(
                        [unique_id, "NE_WIDTH_R should not be populated for this Return Period"])

                # XS_AREA should not be populated if the EVENT_TYP is not for
                # the 1% Annual Return Period
                if xs_area != -9999 and event_typ not in ('01pct', '1 Percent Chance'):
                    self.errors.append(
                        [unique_id, "XS_AREA should not be populated for this Return Period"])

                # AREA_UNIT should not be populated if the EVENT_TYP is not for
                # the 1% Annual Return Period
                if str(area_unit) not in ("", " ", "None") and event_typ not in (
                        '01pct', '1 Percent Chance'):
                    self.errors.append(
                        [unique_id, "AREA_UNIT should not be populated for this Return Period"])

                # VELOCITY should not be populated if the EVENT_TYP is not for
                # the 1% Annual Return Period
                if velocity != -9999 and event_typ not in ('01pct', '1 Percent Chance'):
                    self.errors.append(
                        [unique_id, "VELOCITY should not be populated for this Return Period"])

                # VEL_UNIT should not be populated if the EVENT_TYP is not for
                # the 1% Annual Return Period
                if str(vel_unit) not in ("", " ", "None") and event_typ not in (
                        '01pct', '1 Percent Chance'):
                    self.errors.append(
                        [unique_id, "VEL_UNIT should not be populated for this Return Period"])

                # WSEL_WOFWY should not be populated if the EVENT_TYP is not for
                # the 1% Annual Return Period
                if wsel_wofwy != -9999 and event_typ not in ('01pct', '1 Percent Chance'):
                    self.errors.append(
                        [unique_id, "WSEL_WOFWY should not be populated for this Return Period"])

                # WSEL_FLDWY should not be populated if the EVENT_TYP is not for
                # the 1% Annual Return Period
                if wsel_fldwy != -9999 and event_typ not in ('01pct', '1 Percent Chance'):
                    self.errors.append(
                        [unique_id, "WSEL_FLDWY should not be populated for this Return Period"])

                # WSEL_INCRS should not be populated if the EVENT_TYP is not for
                # the 1% Annual Return Period
                if wsel_incrs != -9999 and event_typ not in ('01pct', '1 Percent Chance'):
                    self.errors.append(
                        [unique_id, "WSEL_INCRS should not be populated for this Return Period"])

                    # XS_AREA should be populated when FW_WIDTH is populated
                if xs_area == -9999 and fw_width != -9999:
                    self.errors.append(
                        [unique_id, "XS_AREA should be populated when FW_WIDTH is populated"])

                # AREA_UNIT should be populated when XS_AREA is not -9999
                if area_unit is None:
                    if xs_area != -9999:
                        self.errors.append(
                            [unique_id, "AREA_UNIT should be populated when XS_AREA is not -9999"])
                elif area_unit.strip() in ['', 'NP'] and xs_area != -9999:
                    self.errors.append(
                        [unique_id, "AREA_UNIT should be populated when XS_AREA is not -9999"])

                # VELOCITY should be populated when FW_WIDTH is populated
                if velocity == -9999 and fw_width != -9999:
                    self.errors.append(
                        [unique_id, "VELOCITY should be populated when FW_WIDTH is populated"])

                # VEL_UNIT should be populated when VELOCITY is populated
                if vel_unit is None:
                    if velocity != -9999:
                        self.errors.append(
                            [unique_id, "VEL_UNIT should be populated when VELOCITY is populated"])
                elif vel_unit.strip() in ['', 'NP'] and velocity != -9999:
                    self.errors.append(
                        [unique_id, "VEL_UNIT should be populated when VELOCITY is populated"])

                # LEN_UNIT should be populated when FW_WIDTH is populated
                if len_unit is None:
                    if fw_width != -9999:
                        self.errors.append(
                            [unique_id, "LEN_UNIT should be populated when FW_WIDTH is populated"])
                elif len_unit.strip() in ['', 'NP'] and fw_width != -9999:
                    self.errors.append(
                        [unique_id, "LEN_UNIT should be populated when FW_WIDTH is populated"])

                # LEN_UNIT should be populated when WSEL is populated
                if len_unit is None:
                    if wsel != -9999:
                        self.errors.append(
                            [unique_id, "LEN_UNIT should be populated when WSEL is populated"])
                elif len_unit.strip() in ['', 'NP'] and wsel != -9999:
                    self.errors.append(
                        [unique_id, "LEN_UNIT should be populated when WSEL is populated"])

                # V_DATUM checks
                if v_datum:
                    # V_DATUM does not match value in Study_Info
                    if v_datum != self.v_datum and self.v_datum != '' and \
                            self.mip_task in ['Develop Final Mapping Products Data Capture',
                                              'Draft FIRM Database Data Capture',
                                              'Produce Preliminary Products Data Capture']:
                        self.errors.append([unique_id, "V_DATUM value of " + str(row[3]) +
                                            " does not match the V_DATUM value in Study_Info"])

                # LEVEE_TF should not be 'U'
                if levee_tf == 'U':
                    self.errors.append([unique_id, "LEVEE_TF should not be 'U'"])

                # LVSCENARIO should be populated if LEVEE_TF is 'T'
                if lvscenario is None:
                    if levee_tf == 'T':
                        self.errors.append(
                            [unique_id, "LVSCENARIO should be populated if LEVEE_TF is 'T'"])
                elif lvscenario.strip() == '' and levee_tf == 'T':
                    self.errors.append(
                        [unique_id, "LVSCENARIO should be populated if LEVEE_TF is 'T'"])

                # WSEL_INCRS does not equal WSEL_FLDWY - WSEL_WOFWY
                if wsel_incrs and wsel_fldwy and wsel_wofwy:
                    if wsel_fldwy != -9999 and wsel_wofwy != -9999:
                        # Convert to integers to improve math opeartions
                        wsel_incrs_int = int(round(wsel_incrs, 1) * 10)
                        wsel_fldwy_int = int(round(wsel_fldwy, 1) * 10)
                        wsel_wofwy_int = int(round(wsel_wofwy, 1) * 10)

                        if wsel_incrs_int != wsel_fldwy_int - wsel_wofwy_int:
                            self.errors.append([unique_id,
                                                "WSEL_INCRS does not equal WSEL_FLDWY - WSEL_WOFWY "
                                                "(try re-rounding the value)"])

                # CALC_WO_BW should be 'T' if WSEL does not equal WSEL_WOFWY
                if calc_wo_bc != 'T' and wsel != wsel_wofwy and wsel_wofwy != -9999:
                    self.errors.append(
                        [unique_id, "CALC_WO_BW should be 'T' if WSEL does not equal WSEL_WOFWY"])

                # EVAL_LN should not be 'U', if using 2020 spec
                if self.schema == "2020":
                    if eval_ln == 'U':
                        self.errors.append([unique_id, "EVAL_LN should not be 'U'"])

    def l_xs_struct_check(self, in_table):
        """QC check of L_XS_Struct"""
        # Required fields
        fields = ('DFIRM_ID', 'VERSION_ID', 'XS_STR_ID', 'XS_LN_ID', 'STRUCT_TYP', 'WTR_NM',
                  'STRUC_FACE', 'STR_STN', 'LO_CHRD_EL', 'HI_CHRD_EL', 'STRMBED_EL',
                  'LEN_UNIT', 'V_DATUM')

        # Fields for specific checks
        spec_list = (fields[2], 'STRUC_FACE', 'V_DATUM')

        # Get a dictionary of applicable and required fields
        required_fields, applicable_fields = self.__get_field_dict(in_table, fields)

        # Perform the standard table checks
        for error in self.__standard_table_checks(in_table, fields[2], L_XS_Struct_domains,
                                                  required_fields, applicable_fields,
                                                  set(fields + spec_list)):
            self.errors.append(error)

        # If fields were missing, return
        if self.missing_field:
            return

        # Check for matching ID value in S_XS.XS_LN_ID
        for error in self.__id_table_check(in_table, 'XS_LN_ID',
                                           self.workspace + self.dataset + '\\S_XS' + self.shp_ext,
                                           'XS_LN_ID'):
            self.errors.append(error)

        # Perform specific table checks
        with SearchCursor(in_table, spec_list) as cursor:
            for row in cursor:
                unique_id = row[0]
                struc_face = row[1]
                v_datum = row[2]

                # STRUC_FACE value should not be 'UNK' or 'Unknown'
                if str(struc_face) in ['U', 'Unknown']:
                    self.errors.append(
                        [unique_id, "STRUC_FACE value should not be 'UNK' or 'Unknown'"])

                # V_DATUM checks
                if v_datum:
                    # V_DATUM does not match value in Study_Info
                    if v_datum != self.v_datum and self.v_datum != '' and \
                            self.mip_task in ['Develop Final Mapping Products Data Capture',
                                              'Draft FIRM Database Data Capture',
                                              'Produce Preliminary Products Data Capture']:
                        self.errors.append([unique_id, "V_DATUM value of " + str(row[2]) +
                                            " does not match the V_DATUM value in Study_Info"])

    def write_out_errors_dbf(self, in_errors, table_name):
        """Writes out the errors to a DBF file"""
        if not in_errors:
            self.__printer("\tNo errors found")
            return

        if not self.excel_export:
            self.__printer("\t" + str(len(in_errors)) + " error(s) found")
            self.total_errors += len(in_errors)

        # Write out to DBF file
        out_filename = self.out_folder + '\\' + table_name.lower() + '_errors.dbf'
        if arcpy.Exists(out_filename):
            arcpy.Delete_management(out_filename)

        # Create the table
        arcpy.CreateTable_management(self.out_folder, table_name.lower() + '_errors.dbf')

        # Add the fields
        arcpy.AddField_management(out_filename, "Unique_ID", "TEXT", field_length=25)
        arcpy.AddField_management(out_filename, "Error", "TEXT", field_length=254)
        arcpy.AddField_management(out_filename, "Comment", "TEXT", field_length=254)
        arcpy.AddField_management(out_filename, "Response", "TEXT", field_length=254)

        # Drop the extra field that's created
        arcpy.DeleteField_management(out_filename, "Field1")

        # Add the errors to the table
        rows = arcpy.InsertCursor(out_filename)

        for error in in_errors:
            row = rows.newRow()
            row.setValue("Unique_ID", error[0])
            row.setValue("Error", error[1])
            rows.insertRow(row)

        # Delete the cursor/row objects
        del row
        del rows

    def write_out_errors_exel(self, in_errors, table_name):
        """Writes out the errors to a DBF file"""
        # Return if no errors are present
        if not in_errors:
            self.__printer("\tNo errors found")
            return

        self.__printer("\t" + str(len(in_errors)) + " error(s) found")
        self.total_errors += len(in_errors)

        # Excel file to write to
        out_filename = self.out_folder + '\\Errors.xlsx'

        # Check to see if the Excel file exists, if not create it.
        if not os.path.exists(out_filename):
            new_workbook = openpyxl.Workbook()
            new_workbook.save(out_filename)

        # Open the workbook for writing
        workbook = openpyxl.load_workbook(out_filename)

        # Get a list of worksheets
        sheet_names = workbook.get_sheet_names()

        # If the worksheet for the current table doesn't exist, create it
        if table_name not in sheet_names:
            workbook.create_sheet(title=table_name)

        # Remove the first sheet
        if 'Sheet' in sheet_names:
            workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))

        # Set the current sheet to active
        sheet = workbook.get_sheet_by_name(table_name)

        # Add the Header
        sheet['A1'].font = Font(size=12, bold=True)
        sheet['B1'].font = Font(size=12, bold=True)
        sheet['C1'].font = Font(size=12, bold=True)
        sheet['D1'].font = Font(size=12, bold=True)
        sheet['A1'].border = Border(bottom=Side(border_style="thick", color="00000f"))
        sheet['B1'].border = Border(bottom=Side(border_style="thick", color="00000f"))
        sheet['C1'].border = Border(bottom=Side(border_style="thick", color="00000f"))
        sheet['D1'].border = Border(bottom=Side(border_style="thick", color="00000f"))
        sheet['A1'].fill = PatternFill("solid", fgColor="DDDDDD")
        sheet['B1'].fill = PatternFill("solid", fgColor="DDDDDD")
        sheet['C1'].fill = PatternFill("solid", fgColor="DDDDDD")
        sheet['D1'].fill = PatternFill("solid", fgColor="DDDDDD")
        sheet['A1'].alignment = Alignment(horizontal="center")
        sheet['B1'].alignment = Alignment(horizontal="center")
        sheet['C1'].alignment = Alignment(horizontal="center")
        sheet['D1'].alignment = Alignment(horizontal="center")
        sheet['A1'] = "ID From Table"
        sheet['B1'] = "Error Found"
        sheet['C1'] = "Comment"
        sheet['D1'] = "Response"

        # Freeze top row
        sheet.freeze_panes = "A2"

        # Set dimensions
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 100
        sheet.column_dimensions['C'].width = 50
        sheet.column_dimensions['D'].width = 50

        # Set the initial row
        row = 2

        # Write the errors to the Worksheet
        for error in in_errors:
            sheet['A' + str(row)].alignment = Alignment(horizontal="center")
            sheet['A' + str(row)] = error[0]
            sheet['B' + str(row)] = error[1]
            row += 1

        # Save the file
        workbook.save(out_filename)


if __name__ == '__main__':

    try:
        workspace = sys.argv[1]
        output_folder = sys.argv[2]
        mip_task = sys.argv[3]
        schema = sys.argv[4]
        tables = sys.argv[5]
        coded_check = sys.argv[6]
        shapefile_export = sys.argv[7]
        excel_export = sys.argv[8]

        qc_check = QCChecks(workspace, output_folder, mip_task, schema, tables,
                            coded_check, shapefile_export, excel_export)
        qc_check.iterate_tables()

    except arcpy.ExecuteError:
        arcpy.AddError(arcpy.GetMessages(2))
        print(arcpy.GetMessages(2))

    finally:
        # Delete the feature layer
        if arcpy.Exists("fc_lyr"):
            arcpy.Delete_management("fc_lyr")
        if arcpy.Exists("xs_lyr"):
            arcpy.Delete_management("xs_lyr")
        if arcpy.Exists("xs_elev"):
            arcpy.Delete_management("xs_elev")

